"""
shared/sheets_integration.py — Google Drive + Sheets integration.

Lets the app read from / write to the user's Google Sheets directly so
they don't have to download / re-upload Excel files. Uses an OAuth2
desktop flow on first use (one-time consent), then caches a token in
config/sheets_token.json that's shared across all sheet-aware features.

Public API
----------
status(resources_path) -> dict
    {configured, has_token, scopes}

reauthorize(resources_path, port=8600) -> dict
    Run the desktop OAuth flow. Opens system browser, waits for consent,
    writes config/sheets_token.json. Idempotent.

list_spreadsheets(resources_path, query='') -> list[dict]
    Newest-first list of every spreadsheet the user can see in Drive
    (filtered by the optional name fragment).

get_tabs(resources_path, spreadsheet_id) -> list[dict]
    Tabs (sheets) inside a spreadsheet, with row/col counts.

read_sheet(resources_path, spreadsheet_id, tab_name, range_a1='') -> list[list]
    2-D array of cell values.

write_range(resources_path, spreadsheet_id, range_a1, values) -> dict
    Overwrite a range with the given 2-D array. Range form: "TabName!A2:B10".

update_cell(resources_path, spreadsheet_id, tab_name, row, col, value) -> dict
    1-indexed row + column.

ensure_column(resources_path, spreadsheet_id, tab_name, header) -> int
    Find a column in row 1 by exact header (case-insensitive). If not
    present, append it to the right and return the new 1-based column
    index. Used when an op needs a "Status" column to write into.

batch_update_status(resources_path, spreadsheet_id, tab_name, status_col,
                    row_values) -> dict
    Write many { row_index → status_text } entries in a single API call.

Notes
-----
- Scope set kept minimal: Drive metadata.readonly (to list files) plus
  Spreadsheets full read/write.
- All errors are logged via shared.logger.print and surfaced as a dict
  {success: False, message: ...} so callers don't have to wrap.
"""

from __future__ import annotations

import json
import os
import threading
from pathlib import Path

# Scopes:
#   drive.metadata.readonly  — list spreadsheet files in the user's Drive
#   spreadsheets             — full read/write on the chosen spreadsheets
SCOPES = [
    'https://www.googleapis.com/auth/drive.metadata.readonly',
    'https://www.googleapis.com/auth/spreadsheets',
]

TOKEN_FILENAME = 'sheets_token.json'
CRED_FILENAME = 'gdrive_credentials.json'   # reuse the existing OAuth client

# ─────────────────────────────────────────────────────────────────────────────
# Retry helper — exponential backoff for Google API 429 / 503 errors
# ─────────────────────────────────────────────────────────────────────────────

import time as _time

def _api_call_with_retry(fn, *args, max_retries: int = 3, base_delay: float = 1.0, **kwargs):
    """Execute a Google API callable with exponential backoff.

    Retries on HTTP 429 (rate limit) and HTTP 503 (transient server error).
    Uses jittered backoff: delay = base_delay * 2^attempt + rand(0, 1).

    Args:
        fn:          A zero-arg callable that performs the API request,
                     typically a bound .execute() call wrapped in a lambda.
        max_retries: Max number of additional attempts after the first failure.
        base_delay:  Starting delay in seconds before first retry.

    Returns:
        The result of fn() on success.

    Raises:
        The last exception if all retries are exhausted.
    """
    import random as _rnd
    last_exc = None
    for attempt in range(max_retries + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_exc = e
            # Only retry on rate-limit (429) or transient server errors (500, 503)
            retryable = False
            try:
                from googleapiclient.errors import HttpError
                if isinstance(e, HttpError):
                    status = int(getattr(getattr(e, 'resp', None), 'status', 0))
                    retryable = status in (429, 500, 503)
            except Exception:
                pass
            if not retryable or attempt >= max_retries:
                raise
            delay = base_delay * (2 ** attempt) + _rnd.uniform(0, 1)
            print(f"[SHEETS] API error (attempt {attempt + 1}/{max_retries + 1}), "
                  f"retrying in {delay:.1f}s: {type(e).__name__}", flush=True)
            _time.sleep(delay)
    raise last_exc  # should never reach here



def _token_path(resources_path) -> Path:
    return Path(resources_path) / 'config' / TOKEN_FILENAME


def _cred_path(resources_path) -> Path:
    return Path(resources_path) / 'config' / CRED_FILENAME


_load_error: dict = {'msg': ''}   # populated when _load_creds fails

def _load_creds(resources_path):
    """Load + refresh the cached OAuth credentials. Returns Credentials
    object or None on any failure (with a human-readable reason cached
    in _load_error)."""
    _load_error['msg'] = ''
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
    except Exception as e:
        _load_error['msg'] = f'google libs missing: {type(e).__name__}: {e}'
        return None
    p = _token_path(resources_path)
    if not p.exists():
        _load_error['msg'] = 'No sheets token — click Connect Google Sheets first.'
        return None
    try:
        from shared.credential_vault import decrypt_json_file, encrypt_json_file, is_file_encrypted
        raw = decrypt_json_file(str(p))
        if not raw:
            raise ValueError("Token file empty or decryption failed")
        import json
        info = json.loads(raw)
        creds = Credentials.from_authorized_user_info(info, SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            p.write_text(creds.to_json(), encoding='utf-8')
            encrypt_json_file(str(p))
        elif not is_file_encrypted(str(p)):
            encrypt_json_file(str(p))
        return creds
    except Exception as e:
        _load_error['msg'] = f'token load failed: {type(e).__name__}: {e}'
        return None


_build_error: dict = {'msg': ''}


def _format_error(e: Exception) -> str:
    """Squeeze a useful one-liner out of any googleapiclient / google-auth
    exception. HttpError instances often have empty str() but carry their
    real reason in .resp.status / .content / .error_details."""
    try:
        from googleapiclient.errors import HttpError
    except Exception:
        HttpError = None   # type: ignore
    parts = [type(e).__name__]
    if HttpError is not None and isinstance(e, HttpError):
        try:
            status = getattr(getattr(e, 'resp', None), 'status', None)
            if status is not None:
                parts.append(f'HTTP {status}')
        except Exception:
            pass
        # error_details is a list of dicts when populated
        try:
            details = getattr(e, 'error_details', None)
            if details:
                parts.append(str(details)[:160])
        except Exception:
            pass
        # Body bytes — last resort, often has a 'message' field
        try:
            content = getattr(e, 'content', None)
            if content:
                txt = content.decode('utf-8', errors='replace') if isinstance(content, (bytes, bytearray)) else str(content)
                # Find the "message": "..." substring if present
                import json as _j
                try:
                    body = _j.loads(txt)
                    err_obj = (body or {}).get('error') or {}
                    if err_obj.get('message'):
                        parts.append(err_obj['message'][:200])
                    elif err_obj.get('status'):
                        parts.append(err_obj['status'])
                except Exception:
                    if 'message' in txt:
                        parts.append(txt[:200])
        except Exception:
            pass
    else:
        s = str(e) or repr(e)
        if s:
            parts.append(s[:200])
    return ': '.join(parts)

def _drive(resources_path):
    """Build a Google Drive v3 service."""
    _build_error['msg'] = ''
    creds = _load_creds(resources_path)
    if not creds:
        return None
    try:
        from googleapiclient.discovery import build
        return build('drive', 'v3', credentials=creds, cache_discovery=False)
    except Exception as e:
        _build_error['msg'] = f'Drive build: {_format_error(e)}'
        return None


def _sheets(resources_path):
    """Build a Google Sheets v4 service."""
    _build_error['msg'] = ''
    creds = _load_creds(resources_path)
    if not creds:
        return None
    try:
        from googleapiclient.discovery import build
        return build('sheets', 'v4', credentials=creds, cache_discovery=False)
    except Exception as e:
        _build_error['msg'] = f'Sheets build: {_format_error(e)}'
        return None


def _why_no_service() -> str:
    """Human-readable reason the Drive/Sheets service couldn't be built.
    Falls back to a generic 'not authorised' if both error slots empty."""
    return (_build_error['msg'] or _load_error['msg']
            or 'Sheets not authorized — click Connect Google Sheets.')


# ─────────────────────────────────────────────────────────────────────────────
# Auth status / consent flow
# ─────────────────────────────────────────────────────────────────────────────

def status(resources_path) -> dict:
    p = _token_path(resources_path)
    cred = _cred_path(resources_path)
    has_token = p.exists()
    scopes_ok = False
    if has_token:
        try:
            from shared.credential_vault import decrypt_json_file
            raw = decrypt_json_file(str(p))
            data = json.loads(raw) if raw else {}
            scopes = set(data.get('scopes') or [])
            scopes_ok = all(s in scopes for s in SCOPES)
        except Exception:
            pass
    return {
        'configured': has_token and scopes_ok and cred.exists(),
        'has_token': has_token,
        'has_credentials': cred.exists(),
        'scopes_ok': scopes_ok,
    }


def reauthorize(resources_path, port: int = 8600) -> dict:
    """Run the desktop OAuth flow for the Sheets + Drive scopes."""
    cred = _cred_path(resources_path)
    if not cred.exists():
        return {'success': False,
                'message': 'config/gdrive_credentials.json missing'}
    try:
        from google_auth_oauthlib.flow import InstalledAppFlow
    except Exception as e:
        return {'success': False,
                'message': f'google-auth-oauthlib not available: {e}'}
    try:
        flow = InstalledAppFlow.from_client_secrets_file(str(cred), SCOPES)
        creds = flow.run_local_server(port=port, prompt='consent', open_browser=True)
        token_p = _token_path(resources_path)
        token_p.write_text(creds.to_json(), encoding='utf-8')
        from shared.credential_vault import encrypt_json_file
        encrypt_json_file(str(token_p))
        return {'success': True,
                'message': 'Google Sheets re-authorized successfully'}
    except Exception as e:
        return {'success': False, 'message': f'Re-auth failed: {_format_error(e)}'}


# ─────────────────────────────────────────────────────────────────────────────
# Drive: list spreadsheets
# ─────────────────────────────────────────────────────────────────────────────

def list_spreadsheets(resources_path, query: str = '', limit: int = 50) -> dict:
    drive = _drive(resources_path)
    if not drive:
        return {'success': False,
                'message': _why_no_service(),
                'sheets': []}
    try:
        q_parts = [
            "mimeType = 'application/vnd.google-apps.spreadsheet'",
            'trashed = false',
        ]
        if query.strip():
            # name contains, escape single quotes
            esc = query.replace("'", "\\'")
            q_parts.append(f"name contains '{esc}'")
        q = ' and '.join(q_parts)
        # No spaces after commas — Google's field-mask parser rejects them.
        res = drive.files().list(
            q=q, orderBy='modifiedTime desc', pageSize=limit,
            fields='files(id,name,modifiedTime,owners(displayName,emailAddress))',
        ).execute()
        files = res.get('files') or []
        out = []
        for f in files:
            owners = f.get('owners') or []
            owner = owners[0].get('displayName', '') if owners else ''
            out.append({
                'id': f['id'],
                'name': f['name'],
                'modified': f.get('modifiedTime'),
                'owner': owner,
            })
        return {'success': True, 'sheets': out}
    except Exception as e:
        return {'success': False, 'message': f'Drive list failed: {_format_error(e)}', 'sheets': []}


# ─────────────────────────────────────────────────────────────────────────────
# Sheets: tabs / read / write / append / column helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_tabs(resources_path, spreadsheet_id: str) -> dict:
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service(), 'tabs': []}
    try:
        # NOTE: Google's field-mask parser rejects spaces after commas.
        # Use the exact form 'properties.title,sheets.properties' — no whitespace.
        meta = s.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields='properties.title,sheets.properties',
        ).execute()
        tabs = []
        for sh in meta.get('sheets') or []:
            p = sh.get('properties') or {}
            grid = p.get('gridProperties') or {}
            tabs.append({
                'id': p.get('sheetId'),
                'title': p.get('title'),
                'rows': grid.get('rowCount'),
                'cols': grid.get('columnCount'),
            })
        return {'success': True,
                'spreadsheet_title': meta.get('properties', {}).get('title'),
                'tabs': tabs}
    except Exception as e:
        return {'success': False,
                'message': f'Sheets read failed: {_format_error(e)}',
                'tabs': []}


def _quote_tab(tab: str) -> str:
    """Single-quote a tab name for A1 ranges (escape inner quotes)."""
    safe = tab.replace("'", "''")
    return f"'{safe}'"


def read_sheet(resources_path, spreadsheet_id: str, tab_name: str,
               range_a1: str = '') -> dict:
    """Read a 2-D array of values. range_a1 should be the cell range
    WITHOUT the tab prefix (e.g. 'A1:Z' or '' for the whole tab)."""
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        if range_a1:
            r = f'{_quote_tab(tab_name)}!{range_a1}'
        else:
            r = _quote_tab(tab_name)
        res = s.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=r,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
        return {'success': True, 'values': res.get('values') or []}
    except Exception as e:
        return {'success': False, 'message': f'Read failed: {_format_error(e)}'}


def write_range(resources_path, spreadsheet_id: str, range_a1: str,
                values: list[list]) -> dict:
    """Overwrite *range_a1* (must include the tab name, e.g.
    "Sheet1!A2:B10") with the given 2-D array."""
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        body = {'values': values}
        res = _api_call_with_retry(
            lambda: s.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id, range=range_a1,
                valueInputOption='USER_ENTERED', body=body,
            ).execute()
        )
        return {'success': True,
                'updated_cells': res.get('updatedCells', 0),
                'updated_range': res.get('updatedRange')}
    except Exception as e:
        return {'success': False, 'message': f'Write failed: {_format_error(e)}'}


def update_cell(resources_path, spreadsheet_id: str, tab_name: str,
                row: int, col: int, value) -> dict:
    """Set a single cell (row + col are 1-based)."""
    from openpyxl.utils import get_column_letter
    a1 = f'{_quote_tab(tab_name)}!{get_column_letter(col)}{row}'
    return write_range(resources_path, spreadsheet_id, a1, [[value]])


def append_rows(resources_path, spreadsheet_id: str, tab_name: str,
                values: list[list]) -> dict:
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        res = s.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=_quote_tab(tab_name),
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body={'values': values},
        ).execute()
        return {'success': True,
                'updates': res.get('updates', {})}
    except Exception as e:
        return {'success': False, 'message': f'Append failed: {_format_error(e)}'}


def ensure_column(resources_path, spreadsheet_id: str, tab_name: str,
                  header: str) -> dict:
    """Find a column by exact header (case-insensitive) anywhere in the sheet.
    Scans ALL rows (not just row 1) so it works when headers are in row 11
    or any other row.  If the header is not found, appends it to the SAME
    row as other headers (the first row that has the most non-empty cells).
    Returns {success, col (1-based), header_row (1-based), created}."""
    from openpyxl.utils import get_column_letter
    full_res = read_sheet(resources_path, spreadsheet_id, tab_name)
    if not full_res.get('success'):
        return {'success': False, 'message': full_res.get('message')}
    all_rows = full_res.get('values') or []
    header_l = header.strip().lower()

    # Search every row for the exact header cell
    for ri, row in enumerate(all_rows):
        for ci, cell in enumerate(row or []):
            if str(cell or '').strip().lower() == header_l:
                return {'success': True, 'col': ci + 1,
                        'header_row': ri + 1, 'created': False}

    # Header not found — find the "main header row" (the row with the most
    # non-empty cells) and append the new header there.
    if not all_rows:
        header_row_idx = 0          # empty sheet → use row 1
    else:
        header_row_idx = max(range(len(all_rows)),
                             key=lambda i: sum(1 for c in (all_rows[i] or []) if c))
    header_row = header_row_idx + 1  # 1-based sheet row
    existing_row = all_rows[header_row_idx] if header_row_idx < len(all_rows) else []
    new_col = len(existing_row) + 1
    a1 = f'{_quote_tab(tab_name)}!{get_column_letter(new_col)}{header_row}'
    res = write_range(resources_path, spreadsheet_id, a1, [[header]])
    if not res.get('success'):
        return {'success': False, 'message': res.get('message')}
    return {'success': True, 'col': new_col,
            'header_row': header_row, 'created': True}


def batch_update_status(resources_path, spreadsheet_id: str, tab_name: str,
                        status_col: int, row_values: dict[int, str]) -> dict:
    """Push many status-cell updates in a single API call.

    row_values: {row_index_1based: 'Live'|'Missing'|'Error'|...}
    """
    if not row_values:
        return {'success': True, 'updated': 0}
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(status_col)
    data = []
    for row, val in row_values.items():
        a1 = f'{_quote_tab(tab_name)}!{col_letter}{int(row)}'
        data.append({'range': a1, 'values': [[val]]})
    try:
        res = s.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'valueInputOption': 'USER_ENTERED', 'data': data},
        ).execute()
        return {'success': True, 'updated': res.get('totalUpdatedCells', 0)}
    except Exception as e:
        return {'success': False, 'message': f'Batch update failed: {_format_error(e)}'}


# ─────────────────────────────────────────────────────────────────────────────
# Convenience helper for ops: read a column by exact header
# ─────────────────────────────────────────────────────────────────────────────

def _parse_status_rows(rows: list, target_status: str, tab_name: str) -> dict:
    """Pure parser: given a 2-D values array (already fetched), find the
    Status/Email/Link columns and return matching rows. Shared by both
    the single-tab `read_rows_by_status` and the batched
    `batch_read_rows_by_status` so they stay behaviourally identical.
    """
    if not rows:
        return {'success': True, 'rows': [], 'status_col': None,
                'email_col': None, 'link_col': None,
                'status_header_used': '', 'unique_status_values': []}

    h_idx = None
    st_keys = {'status', 'appeal status', 'review status', 'live status'}
    em_keys = {'email', 'gmail', 'account', 'profile', 'user'}
    lk_keys = {'review live link', 'live link', 'review link', 'link', 'url', 'direct review link'}

    for ri, row in enumerate(rows):
        if not row: continue
        row_vals = [str(c or '').strip().lower() for c in row]
        has_em = any(k in row_vals for k in em_keys)
        has_st = any(k in row_vals for k in st_keys)
        if has_em and has_st:
            h_idx = ri
            break

    if h_idx is None:
        for ri, row in enumerate(rows):
            if not row: continue
            row_vals = [str(c or '').strip().lower() for c in row]
            if any(k in row_vals for k in st_keys):
                h_idx = ri
                break

    if h_idx is None:
        likely_h_row = max(range(len(rows)), key=lambda i: sum(1 for c in (rows[i] or []) if c))
        header_sample = [str(h) for h in (rows[likely_h_row] or []) if h]
        return {'success': False,
                'message': f"Header 'Status' not found in tab '{tab_name}'",
                'available_headers': header_sample[:15]}

    headers = rows[h_idx]
    header_vals = [str(h or '').strip().lower() for h in headers]

    def _get_idx(keys):
        for i, val in enumerate(header_vals):
            if val in keys: return i
        for i, val in enumerate(header_vals):
            if any(k in val for k in keys): return i
        return None

    status_idx = _get_idx(st_keys)
    email_idx = _get_idx(em_keys)
    link_idx = _get_idx(lk_keys)

    if email_idx is None or status_idx is None:
        return {'success': False,
                'message': f"Required headers (Status/Email) not found in row {h_idx+1}",
                'available_headers': [str(h) for h in headers if h]}

    target = target_status.strip().lower()
    out = []
    value_counts: dict[str, int] = {}

    for ri_zero, row in enumerate(rows[h_idx + 1:]):
        ri_one = h_idx + 2 + ri_zero
        st_raw = str(row[status_idx] or '').strip() if status_idx < len(row) else ''
        if st_raw:
            value_counts[st_raw] = value_counts.get(st_raw, 0) + 1
        if st_raw.lower() != target:
            continue
        em = str(row[email_idx] or '').strip() if email_idx < len(row) else ''
        if not em:
            continue
        link = ''
        if link_idx is not None and link_idx < len(row):
            link = str(row[link_idx] or '').strip()
        out.append({'row': ri_one, 'email': em, 'link': link})

    sample_values = sorted(
        ({'value': k, 'count': v} for k, v in value_counts.items()),
        key=lambda x: -x['count'],
    )[:12]

    return {
        'success': True,
        'status_col': status_idx + 1,
        'email_col': email_idx + 1,
        'link_col': (link_idx + 1) if link_idx is not None else None,
        'rows': out,
        'status_header_used': str(headers[status_idx]),
        'unique_status_values': sample_values,
    }


def read_rows_by_status(resources_path, spreadsheet_id: str, tab_name: str,
                        target_status: str = 'missing') -> dict:
    """Find rows in *tab_name* whose 'Status' column matches
    *target_status* (case-insensitive). Returns each match with its
    row index, email, and review live link.

    Uses a targeted header scan: looks for the first row containing BOTH
    an Email-like column and a Status-like column.
    """
    res = read_sheet(resources_path, spreadsheet_id, tab_name)
    if not res.get('success'):
        return {'success': False, 'message': res.get('message')}
    return _parse_status_rows(res.get('values') or [], target_status, tab_name)


def batch_read_rows_by_status(resources_path, spreadsheet_id: str,
                              tab_names: list,
                              target_status: str = 'missing') -> dict:
    """Read many tabs in a SINGLE batchGet API call and parse each one
    for rows matching *target_status*. Used by the Appeal preview so
    selecting 10 tabs costs ONE API call instead of ten sequential ones.

    Output:
      {
        'success': True,
        'tabs': [
          {'tab': 'Tab A', 'success': True, 'rows': [...],
           'status_header_used': '...', 'unique_status_values': [...],
           'status_col': N, 'email_col': N, 'link_col': N},
          {'tab': 'Tab B', 'success': False, 'message': '...'},
          ...
        ]
      }
    """
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    if not tab_names:
        return {'success': True, 'tabs': []}
    ranges = [_quote_tab(t) for t in tab_names]
    try:
        res = s.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id, ranges=ranges,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
    except Exception as e:
        return {'success': False,
                'message': f'Batch read failed: {_format_error(e)}'}
    value_ranges = res.get('valueRanges') or []
    tabs_out: list[dict] = []
    for tab, vr in zip(tab_names, value_ranges):
        rows = vr.get('values') or []
        parsed = _parse_status_rows(rows, target_status, tab)
        parsed['tab'] = tab
        tabs_out.append(parsed)
    return {'success': True, 'tabs': tabs_out}


def read_column_by_header(resources_path, spreadsheet_id: str, tab_name: str,
                          header: str) -> dict:
    """Find a column by exact header (case-insensitive) anywhere in the sheet, 
    then return the values from the rows below it as [(row_index_1based, value)].
    Empty cells are skipped."""
    res = read_sheet(resources_path, spreadsheet_id, tab_name)
    if not res.get('success'):
        return {'success': False, 'message': res.get('message')}
    rows = res.get('values') or []
    if not rows:
        return {'success': True, 'col': None, 'rows': []}
    
    # Robust scan: find the header in ANY row (not just row 0)
    h_row = None
    col_idx = None
    header_l = header.strip().lower()
    
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row or []):
            if str(cell or '').strip().lower() == header_l:
                h_row = ri
                col_idx = ci
                break
        if h_row is not None:
            break
            
    if col_idx is None:
        # Pick the most likely header row for the diagnostic
        likely_h_row = max(range(len(rows)), key=lambda i: sum(1 for c in (rows[i] or []) if c))
        headers = rows[likely_h_row] if likely_h_row < len(rows) else []
        return {'success': False, 'message': f"Header '{header}' not found in tab '{tab_name}'",
                'available_headers': [str(h) for h in headers if h]}
                
    out = []
    for ri, row in enumerate(rows[h_row + 1:], start=h_row + 2):   # rows below header
        v = row[col_idx] if col_idx < len(row) else None
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        out.append((ri, s))
    return {'success': True, 'col': col_idx + 1, 'rows': out}


def batch_read_links(resources_path, spreadsheet_id: str,
                     tab_names: list, header: str) -> dict:
    """Read link columns from many tabs in a SINGLE batchGet API call.

    Returns:
      {
        'success': True,
        'tabs': {
          tab_name: {
            'success': bool,
            'all_rows': [(sheet_row_idx, url), ...],  # 1-indexed, all rows
            'link_col': int,   # 1-indexed column of the link header
          }
          | {'success': False, 'message': str}
        }
      }
    """
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    if not tab_names:
        return {'success': True, 'tabs': {}}
    ranges = [_quote_tab(t) for t in tab_names]
    try:
        res = s.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=ranges,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
    except Exception as e:
        return {'success': False, 'message': f'Batch read failed: {_format_error(e)}'}

    header_l = header.strip().lower()
    value_ranges = res.get('valueRanges') or []
    out: dict = {}

    for tab, vr in zip(tab_names, value_ranges):
        rows = vr.get('values') or []
        if not rows:
            out[tab] = {'success': False, 'message': 'Tab is empty'}
            continue

        # Find header row + column
        h_row, h_col = None, None
        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row or []):
                if str(cell or '').strip().lower() == header_l:
                    h_row, h_col = ri, ci
                    break
            if h_row is not None:
                break

        if h_row is None:
            out[tab] = {'success': False, 'message': f"Header '{header}' not found"}
            continue

        all_rows = []
        for offset, row in enumerate(rows[h_row + 1:]):
            sheet_row = h_row + 2 + offset   # 1-indexed sheet row
            v = row[h_col] if h_col < len(row) else None
            if v is None:
                continue
            sv = str(v).strip()
            if not sv:
                continue
            all_rows.append((sheet_row, sv))

        out[tab] = {
            'success': True,
            'all_rows': all_rows,
            'link_col': h_col + 1,
        }

    return {'success': True, 'tabs': out}


def batch_read_links_and_status(resources_path, spreadsheet_id: str,
                                tab_names: list,
                                link_header: str = 'Review Live Link',
                                status_header: str = 'Status') -> dict:
    """Single batchGet that returns BOTH link rows AND Status column info
    for every tab. Lets the multi-tab live-check stream per-URL writes
    in real time without paying for N×2 extra reads upfront.

    Returns:
      {
        'success': True,
        'tabs': {
          tab_name: {
            'success': bool,
            'all_rows': [(sheet_row_idx, url), ...],   # link col, 1-indexed
            'link_col': int,
            'status_exists': bool,
            'status_col': int|None,                    # 1-indexed
            'status_header_row': int|None,             # 1-indexed
            'existing_status': {sheet_row_idx: str},   # current cell values
          }
        }
      }
    """
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    if not tab_names:
        return {'success': True, 'tabs': {}}
    ranges = [_quote_tab(t) for t in tab_names]
    try:
        res = s.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id, ranges=ranges,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
    except Exception as e:
        return {'success': False, 'message': f'Batch read failed: {_format_error(e)}'}

    link_l = link_header.strip().lower()
    status_l = status_header.strip().lower()
    value_ranges = res.get('valueRanges') or []
    out: dict = {}

    for tab, vr in zip(tab_names, value_ranges):
        rows = vr.get('values') or []
        if not rows:
            out[tab] = {'success': False, 'message': 'Tab is empty'}
            continue

        # Locate the Link header (first matching cell) and the Status
        # header (any matching cell — typically same row as Link).
        link_row = link_col = None
        status_row = status_col = None
        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row or []):
                cv = str(cell or '').strip().lower()
                if link_row is None and cv == link_l:
                    link_row, link_col = ri, ci
                if status_row is None and cv == status_l:
                    status_row, status_col = ri, ci
            if link_row is not None and status_row is not None:
                break

        if link_row is None:
            out[tab] = {'success': False,
                        'message': f"Header '{link_header}' not found"}
            continue

        all_rows = []
        for offset, row in enumerate(rows[link_row + 1:]):
            sheet_row = link_row + 2 + offset   # 1-indexed
            v = row[link_col] if link_col < len(row) else None
            if v is None:
                continue
            sv = str(v).strip()
            if not sv:
                continue
            all_rows.append((sheet_row, sv))

        existing_status: dict[int, str] = {}
        if status_row is not None and status_col is not None:
            for offset, row in enumerate(rows[status_row + 1:]):
                sheet_row = status_row + 2 + offset
                v = row[status_col] if status_col < len(row) else None
                if v is None:
                    continue
                sv = str(v).strip()
                if sv:
                    existing_status[sheet_row] = sv

        out[tab] = {
            'success': True,
            'all_rows': all_rows,
            'link_col': link_col + 1,
            'status_exists': status_row is not None,
            'status_col': (status_col + 1) if status_col is not None else None,
            'status_header_row': (status_row + 1) if status_row is not None else None,
            'existing_status': existing_status,
        }

    return {'success': True, 'tabs': out}


def batch_count_links_by_header(resources_path, spreadsheet_id: str,
                                 tab_names: list, header: str,
                                 status_header: str = 'Status') -> dict:
    """Probe many tabs in a SINGLE API call (spreadsheets.values.batchGet)
    and return, for each tab, whether the requested header exists and
    how many URL-like values sit beneath it. Used by the Live Check
    picker so it doesn't fire N parallel reads and trip the per-minute
    quota.

    Also reads the `Status` column (when present) and returns a per-
    lowercased-value row count so the Live Check modal can compute the
    actual "rows that match the chosen status filter" count without
    a second round-trip.

    Output:
      {
        'success': True,
        'tabs': [
          {
            tab, success,
            header_row?, col?, link_count?,
            status_col_found?,           # bool — was 'Status' header present
            status_counts?: {lower: int}, # row count per lowercased status
            message?,
          },
          ...
        ]
      }
    """
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    if not tab_names:
        return {'success': True, 'tabs': []}
    ranges = [_quote_tab(t) for t in tab_names]
    try:
        res = s.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=ranges,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
    except Exception as e:
        return {'success': False,
                'message': f'Batch read failed: {_format_error(e)}'}
    value_ranges = res.get('valueRanges') or []
    header_l = header.strip().lower()
    status_l = (status_header or '').strip().lower()
    tabs_out: list[dict] = []
    for tab, vr in zip(tab_names, value_ranges):
        rows = vr.get('values') or []
        if not rows:
            tabs_out.append({'tab': tab, 'success': False,
                             'message': 'Tab is empty'})
            continue
        h_row, h_col = None, None
        s_row, s_col = None, None
        for ri, row in enumerate(rows):
            for ci, h in enumerate(row or []):
                cv = str(h or '').strip().lower()
                if h_row is None and cv == header_l:
                    h_row, h_col = ri, ci
                if status_l and s_row is None and cv == status_l:
                    s_row, s_col = ri, ci
            if h_row is not None and (not status_l or s_row is not None):
                break
        if h_row is None:
            tabs_out.append({'tab': tab, 'success': False,
                             'message': f"Header '{header}' not found"})
            continue
        seen: set[str] = set()
        link_count = 0
        status_counts: dict = {}
        # Walk rows starting AFTER whichever header sits earlier — keeps
        # the row pairing intact even if Status sits above Link or vice
        # versa. We re-look-up Status by absolute row index.
        data_start = h_row + 1
        for offset, row in enumerate(rows[data_start:]):
            v = row[h_col] if h_col < len(row) else None
            if v is None:
                continue
            sv = str(v).strip()
            sv_l = sv.lower()
            if not (sv_l.startswith('http://') or sv_l.startswith('https://')):
                continue
            if sv_l in seen:
                continue
            seen.add(sv_l)
            link_count += 1
            # Per-status row count (uses lowercased Status cell value).
            # Blank/empty status rows are bucketed as '' so callers can
            # see them when investigating "filter excluded too many"
            # situations.
            if s_col is not None and s_row is not None:
                # Map the link's row offset back to a Status cell — the
                # two columns share the same data rows.
                abs_row = data_start + offset
                if abs_row > s_row:
                    s_offset = abs_row - (s_row + 1)
                    s_rows_slice = rows[s_row + 1:]
                    if 0 <= s_offset < len(s_rows_slice):
                        s_row_data = s_rows_slice[s_offset]
                        sv_status = (s_row_data[s_col]
                                     if s_col < len(s_row_data) else None)
                        key = str(sv_status or '').strip().lower()
                        status_counts[key] = status_counts.get(key, 0) + 1
        tabs_out.append({
            'tab': tab,
            'success': True,
            'header_row': h_row + 1,
            'col': h_col + 1,
            'link_count': link_count,
            'status_col_found': s_col is not None,
            'status_counts': status_counts,
        })
    return {'success': True, 'tabs': tabs_out}
