"""
shared/live_status_check.py — Bulk-check whether GMB reviews are LIVE.

Reads an Excel file that has a "Review Live Link" (or similar) column,
visits each URL in a headless browser, and writes a NEW workbook out
with a "Live Status" column populated as Live / Missing / Error.

Detection logic ported from E:/mailexus-advanced/step4/operations/live_check.py
— same selectors and missing-text indicators as the proven implementation.

Standalone op: NO login, NO profile manager, NO NST. The browser is a
fresh headless Chromium spun up via Playwright. Its user-data folder is
created in a dedicated temp directory and DELETED after the run finishes
so the host disk doesn't accumulate junk. Profile-manager data folders
are never touched.
"""

from __future__ import annotations

import asyncio
import os
import shutil
import tempfile
import threading
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Detection rules (copied from mailexus-advanced/step4/operations/live_check.py)
# ─────────────────────────────────────────────────────────────────────────────

LIVE_SELECTORS = [
    'div.Upo0Ec',                        # Like/Share container
    'button[aria-label="Like"]',
    'button[aria-label="Share"]',
    'button.gllhef[data-review-id]',
    'span.wiI7pd',                       # review text
    'div.MyEned',                        # review card
    'div.jftiEf',                        # review body
    'button[data-tooltip="Like"]',
    'div.DUwDvf',                        # review header
    'span.RfnDt',                        # reviewer name
]

MISSING_INDICATORS = [
    'this content is no longer available',
    "content isn't available",
    'page not found',
    "couldn't find",
    'no longer exists',
    'has been removed',
    'violates our policies',
]

JS_CHECK = """
(() => {
    const selectors = [
        'div.Upo0Ec', 'button[aria-label="Like"]', 'button[aria-label="Share"]',
        'span.wiI7pd', 'div.MyEned', 'div.jftiEf', 'button[data-tooltip="Like"]',
    ];
    for (const sel of selectors) {
        const el = document.querySelector(sel);
        if (el && el.offsetParent !== null) return true;
    }
    const hasStars = document.querySelectorAll('span.fzvQIb, span.kvMYJc').length > 0;
    const hasReview = document.querySelectorAll('div.MyEned, div.jftiEf, span.wiI7pd').length > 0;
    return hasStars && hasReview;
})()
"""


# ─────────────────────────────────────────────────────────────────────────────
# Module state
# ─────────────────────────────────────────────────────────────────────────────

_status: dict = {
    'running': False, 'started_at': '', 'finished_at': '',
    'total': 0, 'done': 0, 'live': 0, 'not_live': 0, 'errors': 0,
    'current_url': '', 'report_path': '',
}
_status_lock = threading.Lock()
_cancel = threading.Event()


# ─────────────────────────────────────────────────────────────────────────────
# Status filter — maps UI chip names → all string variants found in the
# wild (case-insensitive). When the user wants "Live check only" we limit
# the run to rows whose Status column already holds one of these values.
# 'all' means no filter (every row gets checked — original behaviour).
# ─────────────────────────────────────────────────────────────────────────────
_STATUS_SYNONYMS: dict = {
    'live':     {'live'},
    # Appealed includes the misspellings/variants the codebase has been
    # seeing in user sheets for years — keep these in sync with the
    # APPEAL_LIKE set inside _worker_sheet / _worker_sheet_tabs.
    'appealed': {'appeal', 'appealed', 'applead', 'applied'},
    'done':     {'done'},
    'missing':  {'missing'},
    'disabled': {'disabled', 'disable'},
}


def _build_status_allow_set(status_filter) -> set | None:
    """Return a lowercased set of Status values to allow, or None for
    'no filter' (the All-check case).

    A None return means the caller should run the original unfiltered
    path — important so callers can branch without paying the filter
    cost (existing_status read, per-row check) on every run.
    """
    if not status_filter:
        return None
    allow: set = set()
    for s in status_filter:
        sl = str(s or '').strip().lower()
        if not sl or sl == 'all':
            return None  # 'all' overrides any specific picks
        allow |= _STATUS_SYNONYMS.get(sl, {sl})
    return allow if allow else None
# Live worker handles so cancel() can forcibly close every browser.
# 'browsers' → list of Playwright BrowserContexts (empty when idle)
# 'loop'     → the asyncio loop the worker is using (so cancel can
#              schedule the async close on the right loop from a
#              different thread)
_live_browser_ref: dict = {'browsers': [], 'loop': None}


def get_status() -> dict:
    with _status_lock:
        return dict(_status)


def cancel():
    """Hard cancel: signal the run to stop AND forcibly close every
    live browser so any in-flight goto/wait calls abort immediately.
    Without the close, navigation timeouts can keep the workers busy
    for another 20-30s after Stop is pressed."""
    _cancel.set()
    browsers = list(_live_browser_ref.get('browsers') or [])
    loop = _live_browser_ref.get('loop')
    if not browsers or loop is None:
        return
    # Schedule b.close() on the worker's own asyncio loop. We can call
    # this from any thread thanks to call_soon_threadsafe; awaiting
    # from elsewhere would crash because Playwright objects are bound
    # to the loop that created them.
    def _kill_all():
        for b in browsers:
            try:
                asyncio.ensure_future(b.close())
            except Exception:
                pass
    try:
        loop.call_soon_threadsafe(_kill_all)
    except Exception:
        pass


def start(file_path: str, num_workers: int = 5, timeout_sec: int = 20,
          resources_path: Path | None = None, show_browser: bool = False,
          status_filter: list | None = None) -> dict:
    with _status_lock:
        if _status['running']:
            return {'success': False, 'message': 'Already running'}

    if not Path(file_path).exists():
        return {'success': False, 'message': f'File not found: {file_path}'}

    _cancel.clear()
    threading.Thread(
        target=_worker,
        args=(file_path, num_workers, timeout_sec, resources_path, show_browser),
        kwargs={'status_filter': status_filter},
        daemon=True, name='live-status-check',
    ).start()
    return {'success': True}


def start_from_sheet(sheet_id: str, tab_name: str, num_workers: int = 5,
                     timeout_sec: int = 20,
                     resources_path: Path | None = None,
                     show_browser: bool = False,
                     status_filter: list | None = None) -> dict:
    """Live-check the URLs in a Google Sheet tab. Status is written back
    to the same tab's Status column in real time."""
    with _status_lock:
        if _status['running']:
            return {'success': False, 'message': 'Already running'}
    if not sheet_id or not tab_name:
        return {'success': False, 'message': 'sheet_id + tab_name required'}
    _cancel.clear()
    threading.Thread(
        target=_worker_sheet,
        args=(sheet_id, tab_name, num_workers, timeout_sec,
              resources_path, show_browser),
        kwargs={'status_filter': status_filter},
        daemon=True, name='live-status-check-sheet',
    ).start()
    return {'success': True}


def start_from_sheet_tabs(sheet_id: str, tabs: list, num_workers: int = 5,
                          timeout_sec: int = 20,
                          resources_path: Path | None = None,
                          show_browser: bool = False,
                          status_filter: list | None = None) -> dict:
    """Live-check URLs across multiple Google Sheet tabs in one pass.
    Reads all tabs, deduplicates URLs, runs checks, then writes
    verdicts back to every tab's Status column."""
    with _status_lock:
        if _status['running']:
            return {'success': False, 'message': 'Already running'}
    if not sheet_id or not tabs:
        return {'success': False, 'message': 'sheet_id + tabs required'}
    _cancel.clear()
    threading.Thread(
        target=_worker_sheet_tabs,
        args=(sheet_id, list(tabs), num_workers, timeout_sec,
              resources_path, show_browser),
        kwargs={'status_filter': status_filter},
        daemon=True, name='live-status-check-sheet-tabs',
    ).start()
    return {'success': True}


# ─────────────────────────────────────────────────────────────────────────────
# Worker
# ─────────────────────────────────────────────────────────────────────────────

def _worker_sheet(sheet_id: str, tab_name: str, num_workers: int,
                  timeout_sec: int, resources_path: Path | None,
                  show_browser: bool, status_filter: list | None = None):
    """Same as _worker, but reads URLs from a Google Sheet tab and
    writes the verdicts back to that tab's Status column."""
    from shared import sheets_integration as _si
    global _status

    with _status_lock:
        _status.update({
            'running': True, 'started_at': datetime.utcnow().isoformat() + 'Z',
            'finished_at': '', 'total': 0, 'done': 0, 'live': 0,
            'not_live': 0, 'errors': 0, 'current_url': '', 'report_path': '',
        })

    # Sweep stale temp folders
    try:
        _tmp_parent = Path(tempfile.gettempdir())
        for stale in _tmp_parent.glob('nst_live_check_*'):
            try:
                shutil.rmtree(stale, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass
    tmp_root = Path(tempfile.mkdtemp(prefix='nst_live_check_'))

    try:
        # 1. Read the Review Live Link column from the sheet
        col_res = _si.read_column_by_header(
            resources_path, sheet_id, tab_name, 'Review Live Link',
        )
        if not col_res.get('success'):
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = (
                    f"FATAL: {col_res.get('message', 'header not found')}"
                )
            return

        # Build (row, url) lists; dedup by URL but preserve all rows for
        # writing the status back to every occurrence.
        all_rows: list[tuple[int, str]] = []
        items: list[tuple[int, str]] = []
        seen: set[str] = set()
        for ri, v in col_res.get('rows') or []:
            url = v.strip()
            ul = url.lower()
            if not (ul.startswith('http://') or ul.startswith('https://')):
                continue
            all_rows.append((ri, url))
            if ul not in seen:
                seen.add(ul)
                items.append((ri, url))

        with _status_lock:
            _status['total'] = len(items)

        # 2. Make sure a Status column exists on the sheet
        st_res = _si.ensure_column(
            resources_path, sheet_id, tab_name, 'Status',
        )
        if not st_res.get('success'):
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = (
                    f"FATAL: {st_res.get('message', 'could not create Status column')}"
                )
            return
        status_col = st_res['col']
        header_row  = st_res.get('header_row', 1)
        data_start  = header_row + 1

        # 2b. Read the EXISTING value of the Status column for every row
        # so we can promote 'Applead' → 'Done' when the live check
        # confirms the review is still live.
        from openpyxl.utils import get_column_letter
        status_col_letter = get_column_letter(status_col)
        existing_status: dict[int, str] = {}
        try:
            sd = _si.read_sheet(resources_path, sheet_id, tab_name,
                                f'{status_col_letter}{data_start}:{status_col_letter}')
            if sd.get('success'):
                for offset, row in enumerate(sd.get('values') or []):
                    if not row:
                        continue
                    val = str(row[0] if row else '').strip()
                    if val:
                        existing_status[data_start + offset] = val
        except Exception:
            pass

        # 2c. STATUS FILTER — when the user picked specific chips (Live,
        # Appealed, Done, Missing, Disabled) in the modal, drop every row
        # whose existing Status doesn't match. Has to run AFTER existing
        # status is read, and we re-dedup `items` from the filtered
        # `all_rows` so a URL doesn't get checked just because its FIRST
        # occurrence (kept by the original dedup) happened to be a row
        # that the filter would have excluded.
        allow_set = _build_status_allow_set(status_filter)
        if allow_set is not None:
            def _row_passes(ri: int) -> bool:
                cur = (existing_status.get(ri, '') or '').strip().lower()
                return cur in allow_set
            all_rows = [(ri, url) for (ri, url) in all_rows if _row_passes(ri)]
            seen2: set[str] = set()
            items = []
            for ri, url in all_rows:
                ul = url.lower()
                if ul not in seen2:
                    seen2.add(ul)
                    items.append((ri, url))
            with _status_lock:
                _status['total'] = len(items)
            if not items:
                with _status_lock:
                    _status['running'] = False
                    _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                    _status['report_path'] = (
                        f"No rows matched status filter "
                        f"({', '.join(sorted(allow_set))}) — nothing to check"
                    )
                return

        # 3. Set up the per-result sheet writer BEFORE running the checks
        # so each worker can push verdicts to the sheet in real time as
        # they finish (user policy: don't wait for the whole batch to
        # write — they want the sheet up-to-date URL by URL).
        #
        # Status-transition rules (per user spec):
        #     existing 'Live'               + check='Missing' → 'Missing'
        #     existing 'Appealed'           + check='Live'    → 'Done'
        #     existing 'Appealed'           + check='Missing' → unchanged ('Appealed')
        #     existing 'Done'               + check='Missing' → 'Missing'
        #     existing 'Done'               + check='Live'    → 'Done' (unchanged)
        #     anything else                                   → verdict as-is
        # Only the 'Status' column is updated — no other columns are touched.
        APPEAL_LIKE = {'appeal', 'appealed', 'applead', 'applied'}
        DONE_LIKE   = {'done'}

        def _final_status(row_idx: int, verdict: str) -> str:
            cur_raw = existing_status.get(row_idx, '') or ''
            cur = cur_raw.strip().lower()
            if cur in APPEAL_LIKE:
                if verdict == 'Live':
                    return 'Done'
                if verdict == 'Missing':
                    # Don't downgrade an appealed row when the link is
                    # not visible yet — leave the existing value alone.
                    return cur_raw
            if cur in DONE_LIKE:
                if verdict == 'Missing':
                    return 'Missing'
                # Done + Live → stays Done
                return cur_raw
            return verdict

        # Lock guards the existing_status cache + serialises sheet writes
        # across workers. Sheet writes themselves are rate-limited by
        # Google (60 writes/min/user) so contention isn't the bottleneck.
        _sheet_lock = threading.Lock()
        _live_writes = {'updated': 0, 'failed': 0}

        def _on_result_sync(row_idx: int, url: str, verdict: str) -> None:
            """Called from each worker right after a URL verdict is known.
            Mirrors the verdict to every row that holds this URL and
            pushes ONLY the changed cells to the sheet immediately."""
            cells_to_write: dict[int, str] = {}
            url_l = url.lower()
            with _sheet_lock:
                for ri, u in all_rows:
                    if u.lower() != url_l:
                        continue
                    new_val = _final_status(ri, verdict)
                    cur = (existing_status.get(ri, '') or '').strip()
                    if new_val == cur:
                        continue
                    cells_to_write[ri] = new_val
                if not cells_to_write:
                    return
                try:
                    bu_one = _si.batch_update_status(
                        resources_path, sheet_id, tab_name,
                        status_col, cells_to_write,
                    )
                    if bu_one.get('success'):
                        # Update local cache only after a confirmed write
                        # so the final safety-net pass below skips
                        # already-written cells and retries failed ones.
                        for ri, new_val in cells_to_write.items():
                            existing_status[ri] = new_val
                        _live_writes['updated'] += len(cells_to_write)
                    else:
                        _live_writes['failed'] += len(cells_to_write)
                except Exception:
                    _live_writes['failed'] += len(cells_to_write)

        # 4. Run the checks — workers now stream per-URL writes via on_result
        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            results = loop.run_until_complete(
                _run_checks(items, num_workers, timeout_sec, tmp_root,
                            show_browser, on_result=_on_result_sync)
            )
        finally:
            try: loop.close()
            except Exception: pass

        # 5. Final safety-net batch — retries anything per-row writes failed
        # on, and covers verdicts produced AFTER cancellation (queued URLs
        # that get auto-marked 'Cancelled'). Idempotent: cells already
        # written above are skipped because existing_status is up to date.
        verdict_by_url = {url.lower(): v for (_, url, v) in results}
        row_updates: dict[int, str] = {}
        for row_idx, url in all_rows:
            verdict = verdict_by_url.get(url.lower(), 'Error')
            new_val = _final_status(row_idx, verdict)
            cur = (existing_status.get(row_idx, '') or '').strip()
            if new_val == cur:
                continue
            row_updates[row_idx] = new_val

        report_msg = (f"Live: streamed {_live_writes['updated']} cells "
                      f"as URLs were checked")
        if row_updates:
            bu = _si.batch_update_status(
                resources_path, sheet_id, tab_name, status_col, row_updates,
            )
            if bu.get('success'):
                report_msg += (f"; final safety-net wrote {bu.get('updated', 0)} "
                               f"additional cells")
            else:
                report_msg += f"; final batch failed: {bu.get('message')}"
        if _live_writes['failed']:
            report_msg += f"; {_live_writes['failed']} per-row writes failed (retried)"

        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['report_path'] = report_msg
    except Exception as e:
        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['errors'] = (_status.get('errors') or 0) + 1
            _status['current_url'] = f'FATAL: {e}'
    finally:
        try: shutil.rmtree(tmp_root, ignore_errors=True)
        except Exception: pass


def _worker_sheet_tabs(sheet_id: str, tabs: list, num_workers: int,
                       timeout_sec: int, resources_path: Path | None,
                       show_browser: bool, status_filter: list | None = None):
    """Multi-tab variant of _worker_sheet.

    Uses ONE batchGet API call to read all tabs at once (instead of N
    individual reads) so it doesn't trip Google Sheets quota limits.
    Runs checks on deduplicated URLs then writes verdicts back per tab.
    """
    from shared import sheets_integration as _si
    global _status

    with _status_lock:
        _status.update({
            'running': True, 'started_at': datetime.utcnow().isoformat() + 'Z',
            'finished_at': '', 'total': 0, 'done': 0, 'live': 0,
            'not_live': 0, 'errors': 0, 'current_url': 'Reading tabs…',
            'report_path': '',
        })

    try:
        _tmp_parent = Path(tempfile.gettempdir())
        for stale in _tmp_parent.glob('nst_live_check_*'):
            try:
                shutil.rmtree(stale, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass
    tmp_root = Path(tempfile.mkdtemp(prefix='nst_live_check_'))

    try:
        # ── Step 1: SINGLE batchGet — link rows + Status col + existing values
        # for every tab in one round-trip. Lets us start streaming verdict
        # writes from the very first URL check (no end-of-run batch wait).
        batch_res = _si.batch_read_links_and_status(
            resources_path, sheet_id, tabs,
            link_header='Review Live Link', status_header='Status',
        )
        if not batch_res.get('success'):
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = (
                    f"FATAL: {batch_res.get('message', 'batch read failed')}"
                )
            return

        tabs_data = batch_res.get('tabs') or {}

        # ── Step 2: dedup URLs across tabs, build url → [(tab, row)] index
        # When status_filter is set (Live / Appealed / Done / Missing /
        # Disabled chips active in the modal) we drop every row whose
        # existing Status doesn't match BEFORE adding it to any of the
        # tracking maps — so the URL never gets checked AND no verdict
        # gets written back to it. allow_set=None means "All" → original
        # unfiltered behaviour.
        allow_set = _build_status_allow_set(status_filter)
        print(f"[LIVE CHECK] status_filter={status_filter!r} → "
              f"allow_set={sorted(allow_set) if allow_set else 'None (All)'}")
        seen_urls: set[str] = set()
        unique_items: list[tuple[int, str]] = []
        # url_locations[url_lower] = [(tab_name, sheet_row_idx), ...]
        url_locations: dict[str, list] = {}
        # raw_tab_rows[tab_name] = [(row_idx, url), ...] — every row incl. dupes
        raw_tab_rows: dict[str, list] = {}
        # Per-tab counters for diagnostic logging — surface mismatches
        # like "the Status column has 'appealed' but the header was
        # spelled 'Statuses' so existing_status came back empty".
        diag_rows_seen = 0
        diag_rows_kept = 0
        diag_filter_value_counts: dict = {}

        for tab_name in tabs:
            td = tabs_data.get(tab_name) or {}
            if not td.get('success'):
                continue
            es = td.get('existing_status') or {}
            tab_rows: list[tuple[int, str]] = []
            tab_seen = 0
            tab_kept = 0
            for ri, v in td.get('all_rows') or []:
                url = str(v).strip()
                ul = url.lower()
                if not (ul.startswith('http://') or ul.startswith('https://')):
                    continue
                tab_seen += 1
                if allow_set is not None:
                    cur_raw = (es.get(ri, '') or '')
                    cur = str(cur_raw).strip().lower()
                    diag_filter_value_counts[cur or '<blank>'] = (
                        diag_filter_value_counts.get(cur or '<blank>', 0) + 1
                    )
                    if cur not in allow_set:
                        continue
                tab_kept += 1
                tab_rows.append((ri, url))
                url_locations.setdefault(ul, []).append((tab_name, ri))
                if ul not in seen_urls:
                    seen_urls.add(ul)
                    unique_items.append((ri, url))
            diag_rows_seen += tab_seen
            diag_rows_kept += tab_kept
            print(f"[LIVE CHECK] tab='{tab_name}' urls_seen={tab_seen} "
                  f"kept_after_filter={tab_kept} "
                  f"status_col_found={td.get('status_exists')} "
                  f"existing_status_entries={len(es)}")
            if tab_rows:
                raw_tab_rows[tab_name] = tab_rows

        if allow_set is not None:
            top = sorted(diag_filter_value_counts.items(),
                         key=lambda x: -x[1])[:8]
            print(f"[LIVE CHECK] FILTER SUMMARY: kept {diag_rows_kept}/"
                  f"{diag_rows_seen} URL rows. allow={sorted(allow_set)}. "
                  f"Most-common existing-status values seen: {top}")
            # Surface the filter outcome in `current_url` so the user can
            # see WHY a filter run ended up too big / too small — the UI
            # progress panel reads this field. Most useful when the Status
            # column was missing or named differently (existing_status
            # would be empty for that tab, so no row would match).
            top_str = ', '.join(f"{v!r}={c}" for v, c in top) if top else 'none'
            with _status_lock:
                _status['current_url'] = (
                    f"Filter {sorted(allow_set)} kept {diag_rows_kept}/"
                    f"{diag_rows_seen} rows. Values seen: {top_str}"
                )

        with _status_lock:
            _status['total'] = len(unique_items)
            _status['current_url'] = ''

        if not unique_items:
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['report_path'] = (
                    'No Review Live Links found in selected tabs'
                )
            return

        # ── Step 3: ensure every participating tab has a Status column.
        # Already-present columns (the common case) cost ZERO API calls
        # since batch_read_links_and_status returned status_col + existing
        # values for them. Only tabs missing the column trigger ensure_column.
        for tab_name in list(raw_tab_rows.keys()):
            td = tabs_data[tab_name]
            if td.get('status_exists'):
                continue
            st_res = _si.ensure_column(
                resources_path, sheet_id, tab_name, 'Status',
            )
            if not st_res.get('success'):
                print(f"[LIVE CHECK] ✗ {tab_name}: could not create Status "
                      f"column — {st_res.get('message')}")
                td['status_exists'] = False
                continue
            td['status_exists'] = True
            td['status_col'] = st_res['col']
            td['status_header_row'] = st_res.get('header_row', 1)
            td.setdefault('existing_status', {})

        APPEAL_LIKE = {'appeal', 'appealed', 'applead', 'applied'}
        DONE_LIKE   = {'done'}

        def _final_status(tab_name: str, row_idx: int, verdict: str) -> str:
            es = tabs_data[tab_name].get('existing_status') or {}
            cur_raw = es.get(row_idx, '') or ''
            cur = cur_raw.strip().lower()
            if cur in APPEAL_LIKE:
                if verdict == 'Live':
                    return 'Done'
                if verdict == 'Missing':
                    return cur_raw   # keep Appealed unchanged
            if cur in DONE_LIKE:
                if verdict == 'Missing':
                    return 'Missing'
                return cur_raw       # Done + Live → stays Done
            return verdict

        # Serialise sheet writes + guard the existing_status caches across
        # workers. Google Sheets API caps writes at ~60/min/user, so the
        # browser-check cost (~1-5s per URL) keeps us well inside quota.
        _sheet_lock = threading.Lock()
        _live_writes = {'updated': 0, 'failed': 0}

        def _on_result_sync(row_idx: int, url: str, verdict: str) -> None:
            """Called from each worker the moment a URL verdict is known.
            Pushes the verdict to EVERY tab+row holding that URL in real
            time — one batch_update_status per tab so the UI sees the
            sheet light up as each check completes."""
            url_l = url.lower()
            with _sheet_lock:
                # Group changes by tab so each tab is one API call max.
                per_tab: dict[str, dict[int, str]] = {}
                for tab_name, ri in url_locations.get(url_l, []):
                    td = tabs_data.get(tab_name) or {}
                    if not td.get('status_exists'):
                        continue
                    new_val = _final_status(tab_name, ri, verdict)
                    es = td.setdefault('existing_status', {})
                    cur = (es.get(ri, '') or '').strip()
                    if new_val == cur:
                        continue
                    per_tab.setdefault(tab_name, {})[ri] = new_val
                if not per_tab:
                    return
                for tab_name, changes in per_tab.items():
                    td = tabs_data[tab_name]
                    try:
                        bu = _si.batch_update_status(
                            resources_path, sheet_id, tab_name,
                            td['status_col'], changes,
                        )
                        if bu.get('success'):
                            es = td.setdefault('existing_status', {})
                            for ri, nv in changes.items():
                                es[ri] = nv
                            _live_writes['updated'] += len(changes)
                        else:
                            _live_writes['failed'] += len(changes)
                            print(f"[LIVE CHECK] stream write to {tab_name} "
                                  f"failed: {bu.get('message')}")
                    except Exception as e:
                        _live_writes['failed'] += len(changes)
                        print(f"[LIVE CHECK] stream write to {tab_name} "
                              f"crashed: {e}")

        # ── Step 4: run URL checks with streaming writes ──────────────────
        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            results = loop.run_until_complete(
                _run_checks(unique_items, num_workers, timeout_sec,
                            tmp_root, show_browser,
                            on_result=_on_result_sync)
            )
        finally:
            try: loop.close()
            except Exception: pass

        # ── Step 5: safety-net batch — retries failed stream writes and
        # covers verdicts produced AFTER cancellation. Idempotent: cells
        # whose existing_status cache already matches the new value get
        # skipped, so already-streamed cells are untouched.
        verdict_by_url = {url.lower(): v for (_, url, v) in results}
        safety_writes = 0
        total_tabs = len(raw_tab_rows)
        for tab_name, tab_rows in raw_tab_rows.items():
            td = tabs_data[tab_name]
            if not td.get('status_exists'):
                continue
            row_updates: dict[int, str] = {}
            es = td.setdefault('existing_status', {})
            for row_idx, url in tab_rows:
                verdict = verdict_by_url.get(url.lower(), 'Error')
                new_val = _final_status(tab_name, row_idx, verdict)
                cur = (es.get(row_idx, '') or '').strip()
                if new_val != cur:
                    row_updates[row_idx] = new_val
            if row_updates:
                bu = _si.batch_update_status(
                    resources_path, sheet_id, tab_name,
                    td['status_col'], row_updates,
                )
                if bu.get('success'):
                    n = bu.get('updated', 0)
                    safety_writes += n
                    for ri, nv in row_updates.items():
                        es[ri] = nv

        summary = (f"Streamed {_live_writes['updated']} cells live; "
                   f"safety-net wrote {safety_writes} more "
                   f"across {total_tabs} tab(s)")
        if _live_writes['failed']:
            summary += f"; {_live_writes['failed']} stream writes retried"
        print(f"[LIVE CHECK] Done. {summary}")
        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['current_url'] = ''
            _status['report_path'] = summary
    except Exception as e:
        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['errors'] = (_status.get('errors') or 0) + 1
            _status['current_url'] = f'FATAL: {e}'
    finally:
        try: shutil.rmtree(tmp_root, ignore_errors=True)
        except Exception: pass


def _worker(file_path: str, num_workers: int, timeout_sec: int,
            resources_path: Path | None, show_browser: bool = False,
            status_filter: list | None = None):
    import openpyxl
    global _status

    with _status_lock:
        _status.update({
            'running': True, 'started_at': datetime.utcnow().isoformat() + 'Z',
            'finished_at': '', 'total': 0, 'done': 0,
            'live': 0, 'not_live': 0, 'errors': 0,
            'current_url': '', 'report_path': '',
        })

    # Sweep any leftover temp folders from previously crashed runs first
    # (so chromium garbage doesn't accumulate even when the app was killed
    # mid-check). We're scoped to our own prefix only — never touches
    # other apps' temp data or the profile-manager folder.
    try:
        _tmp_parent = Path(tempfile.gettempdir())
        for stale in _tmp_parent.glob('nst_live_check_*'):
            try:
                shutil.rmtree(stale, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass

    # Dedicated temp dir for the browser's user-data — wiped after the run.
    # NEVER points at the profile-manager data folder.
    tmp_root = Path(tempfile.mkdtemp(prefix='nst_live_check_'))
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [str(c.value or '').strip() for c in ws[1]]
        # ONLY accept the exact "Review Live Link" header per user spec.
        link_col_idx = None
        for i, h in enumerate(headers, 1):
            if h.strip().lower() == 'review live link':
                link_col_idx = i
                break
        if link_col_idx is None:
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = ("FATAL: header 'Review Live Link' not found. "
                                          "Add a column named exactly 'Review Live Link'.")
            return

        # Use a SINGLE 'Status' column. Reuse existing one if the file
        # already has 'Status' or 'Live Status', otherwise append it as
        # the next column. No other columns are added — the output keeps
        # every original cell intact and only this one is updated.
        status_col_idx = None
        for i, h in enumerate(headers, 1):
            if h.lower() in ('status', 'live status'):
                status_col_idx = i
                break
        if status_col_idx is None:
            status_col_idx = ws.max_column + 1
            ws.cell(row=1, column=status_col_idx, value='Status')

        # Walk every row once and remember which row had which URL.
        # We CHECK each unique URL only once, but at the end we write
        # the verdict back to EVERY row that carried that URL — so the
        # output keeps every original row intact (no dedup data loss).
        # Status filter (Live / Appealed / Done / Missing / Disabled chips)
        # also reads the current Status cell so we can drop rows whose
        # status doesn't match. allow_set=None → All-check (unfiltered).
        allow_set = _build_status_allow_set(status_filter)
        all_rows: list[tuple[int, str]] = []      # (row_idx, url) every row
        items: list[tuple[int, str]] = []         # (row_idx, url) unique
        first_row_for_url: dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=link_col_idx).value
            if v is None:
                continue
            url = str(v).strip()
            if not url:
                continue
            # Skip cells that aren't actually URLs (notes, junk, etc.)
            ul = url.lower()
            if not (ul.startswith('http://') or ul.startswith('https://')):
                continue
            if allow_set is not None:
                cur_raw = ws.cell(row=r, column=status_col_idx).value
                cur = str(cur_raw or '').strip().lower()
                if cur not in allow_set:
                    continue
            all_rows.append((r, url))
            key = url.lower()
            if key not in first_row_for_url:
                first_row_for_url[key] = r
                items.append((r, url))

        with _status_lock:
            _status['total'] = len(items)

        # Bail out cleanly when the filter excluded every row — saves
        # spinning up Playwright for a zero-URL run.
        if not items:
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['report_path'] = (
                    f"No rows matched status filter "
                    f"({', '.join(sorted(allow_set))}) — nothing to check"
                    if allow_set is not None
                    else 'No Review Live Links found in file'
                )
            wb.close()
            return

        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            results = loop.run_until_complete(
                _run_checks(items, num_workers, timeout_sec, tmp_root, show_browser)
            )
        finally:
            try: loop.close()
            except Exception: pass

        # Build a url → verdict map then write only the Status cell on
        # every row that had this URL. Nothing else in the workbook is
        # changed.
        verdict_by_url: dict[str, str] = {
            url.lower(): verdict for (_, url, verdict) in results
        }
        for row_idx, url in all_rows:
            verdict = verdict_by_url.get(url.lower(), 'Error')
            ws.cell(row=row_idx, column=status_col_idx, value=verdict)

        in_path = Path(file_path)
        out_name = (
            f"{in_path.stem}_live_status_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        # Save to RESOURCES_PATH/output so the report appears in the in-app
        # Report Ledger automatically (alongside batch-login reports etc.).
        # Fallback to next to the input file if resources_path is missing.
        out_dir = None
        if resources_path is not None:
            try:
                out_dir = Path(resources_path) / 'output'
                out_dir.mkdir(parents=True, exist_ok=True)
            except Exception:
                out_dir = None
        if out_dir is None:
            out_dir = in_path.parent
        out_path = out_dir / out_name
        wb.save(out_path)
        wb.close()

        with _status_lock:
            _status['report_path'] = str(out_path)
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'

    except Exception as e:
        with _status_lock:
            _status['running'] = False
            _status['errors'] += 1
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['current_url'] = f'FATAL: {e}'

    finally:
        # Wipe the dedicated temp dir — keeps host disk clean. This is OUR
        # temp folder, not anything inside MailNexusPro / browser_profiles,
        # so profile data is safe.
        try:
            shutil.rmtree(tmp_root, ignore_errors=True)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Async runner
# ─────────────────────────────────────────────────────────────────────────────

async def _run_checks(items: list[tuple[int, str]], workers: int,
                      timeout_sec: int, tmp_root: Path,
                      show_browser: bool = False,
                      on_result=None) -> list[tuple[int, str, str]]:
    """Spawn N independent Chromium browsers (one per worker) and pull
    URLs off a shared queue. Each browser has its own user_data_dir, so
    each gets its own cookies / consent state / process — Google doesn't
    see one browser hammering 100s of review URLs in a row, which is
    what was producing the rate-limit / captcha pattern."""
    from playwright.async_api import async_playwright

    out: list[tuple[int, str, str]] = []
    out_lock = asyncio.Lock()

    n_workers = max(1, workers)

    async def _launch(p, slot: int):
        ud = tmp_root / f'browser_{slot}'
        ud.mkdir(parents=True, exist_ok=True)
        return await p.chromium.launch_persistent_context(
            user_data_dir=str(ud),
            headless=not show_browser,
            locale='en-US',
            user_agent=('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                        'AppleWebKit/537.36 (KHTML, like Gecko) '
                        'Chrome/132.0.0.0 Safari/537.36'),
            extra_http_headers={'Accept-Language': 'en-US,en;q=0.9'},
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-features=Translate',
                '--no-default-browser-check',
                '--lang=en-US',
            ],
        )

    async def _warmup(browser):
        try:
            page = await browser.new_page()
            try:
                await page.goto('https://maps.google.com/?hl=en',
                                wait_until='domcontentloaded', timeout=20000)
                await page.wait_for_timeout(1500)
                for sel in [
                    'button[aria-label*="Accept all" i]',
                    'button:has-text("Accept all")',
                    'button:has-text("I agree")',
                    'button:has-text("Tout accepter")',
                    'form[action*="consent"] button[type="submit"]',
                ]:
                    try:
                        b = page.locator(sel).first
                        if await b.count() > 0 and await b.is_visible(timeout=600):
                            await b.click()
                            await page.wait_for_timeout(800)
                            break
                    except Exception:
                        continue
            finally:
                try: await page.close()
                except Exception: pass
        except Exception:
            pass

    async with async_playwright() as p:
        # Launch all workers' browsers in parallel for fast startup
        browsers = await asyncio.gather(
            *[_launch(p, i) for i in range(n_workers)],
            return_exceptions=False,
        )

        # Publish all browsers + loop so cancel() can close them
        # immediately from a different thread.
        _live_browser_ref['browsers'] = list(browsers)
        _live_browser_ref['loop'] = asyncio.get_event_loop()

        # Warm up every browser in parallel
        await asyncio.gather(*[_warmup(b) for b in browsers],
                             return_exceptions=True)

        # Shared work queue
        q: asyncio.Queue = asyncio.Queue()
        for it in items:
            q.put_nowait(it)

        async def _worker_loop(slot: int, browser):
            while True:
                if _cancel.is_set():
                    return
                try:
                    row_idx, url = q.get_nowait()
                except asyncio.QueueEmpty:
                    return

                with _status_lock:
                    _status['current_url'] = url[:120]
                page = None
                verdict = 'Error'
                try:
                    page = await browser.new_page()
                    # Close any popups that open (Google auth, cookie consent, etc.)
                    # so they don't accumulate as about:blank windows.
                    page.on('popup', lambda p: asyncio.ensure_future(p.close()))
                    verdict = await _check_url(page, url, timeout_sec)
                except Exception as e:
                    if _cancel.is_set():
                        verdict = 'Cancelled'
                    else:
                        verdict = f'Error: {str(e)[:60]}'
                finally:
                    try:
                        if page: await page.close()
                    except Exception:
                        pass

                async with out_lock:
                    out.append((row_idx, url, verdict))
                with _status_lock:
                    _status['done'] += 1
                    if verdict == 'Live':
                        _status['live'] += 1
                    elif verdict == 'Missing':
                        _status['not_live'] += 1
                    else:
                        _status['errors'] += 1

                # Real-time per-verdict callback (sheet writer). Runs in
                # the default thread-pool executor so the sync Google
                # Sheets API call doesn't block this worker's event loop
                # — other workers keep checking URLs while one writes.
                if on_result is not None:
                    try:
                        await asyncio.get_event_loop().run_in_executor(
                            None, on_result, row_idx, url, verdict,
                        )
                    except Exception:
                        # Sheet-write failures are tracked in the
                        # callback's own counter and retried by the
                        # outer safety-net batch.
                        pass

        await asyncio.gather(
            *[_worker_loop(i, b) for i, b in enumerate(browsers)],
            return_exceptions=True,
        )

        # If we were cancelled mid-run, mark every still-queued URL as
        # 'Cancelled' so the report reflects user intent.
        if _cancel.is_set():
            while True:
                try:
                    row_idx, url = q.get_nowait()
                except asyncio.QueueEmpty:
                    break
                async with out_lock:
                    out.append((row_idx, url, 'Cancelled'))

        # Close all browsers
        for b in browsers:
            try: await b.close()
            except Exception: pass

        _live_browser_ref['browsers'] = []
        _live_browser_ref['loop'] = None

    out.sort(key=lambda x: x[0])
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Per-URL check (mailexus logic)
# ─────────────────────────────────────────────────────────────────────────────

async def _check_url(page, url: str, timeout_sec: int) -> str:
    """Per user spec: a posted Google Maps review renders the star-rating
    container with id 'DU9Pgb'. If that element exists on the page,
    treat the review as Live. Anything else is Missing.

    The previous text/selector heuristics are kept ONLY as a tail-end
    fallback in case Google ever changes the id; the id check is what
    decides the verdict in normal operation.
    """
    # Step 1: navigate
    try:
        await page.goto(url, wait_until='domcontentloaded',
                        timeout=timeout_sec * 1000)
    except Exception:
        try:
            await page.goto(url, wait_until='commit', timeout=10000)
        except Exception:
            pass

    # If both gotos left the page at about:blank (redirect intercepted by
    # bot-detection or OAuth setup), do one more direct attempt.
    if page.url in ('about:blank', ''):
        try:
            await page.goto(url, wait_until='commit', timeout=10000)
        except Exception:
            pass

    await asyncio.sleep(3)

    # PRIMARY CHECK — the star-rating container of a posted review.
    # A live review can carry ANY rating (1, 2, 3, 4 or 5 stars), so
    # the verdict isn't tied to a specific star count — we just need
    # to confirm the rating widget is on the page.
    #
    # Selectors below cover both signals the user shared, and they
    # accept any rating because the attribute selector matches the
    # SUBSTRING "star" (covers both "1 star" and "2/3/4/5 stars").
    LIVE_SIGNAL_SELECTORS = [
        # 1) Minified class token on the rating container
        '.DU9Pgb',
        '[class*="DU9Pgb"]',
        '[jsname="DU9Pgb"]',
        # 2) The star-rating widget itself —
        #    <span class="kvMYJc" role="img" aria-label="N stars">
        'span.kvMYJc[role="img"][aria-label*="star"]',
        'span[role="img"][aria-label*="star"]',
    ]
    for _ in range(20):             # up to ~10s
        for sel in LIVE_SIGNAL_SELECTORS:
            try:
                count = await page.locator(sel).count()
                if count > 0:
                    return 'Live'
            except Exception:
                continue
        await asyncio.sleep(0.5)

    # FALLBACK — explicit "review removed / not available" copy.
    try:
        body = (await page.inner_text('body')).lower()
        for ind in MISSING_INDICATORS:
            if ind in body:
                return 'Missing'
    except Exception:
        pass

    return 'Missing'
