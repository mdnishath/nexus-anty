"""
routes/sheets.py — Google Sheets Blueprint

Handles:
  GET  /api/sheets/status                 OAuth status check
  POST /api/sheets/authorize              Run desktop OAuth consent flow
  GET  /api/sheets/list                   List spreadsheets in Drive
  GET  /api/sheets/<sheet_id>/tabs        List tabs inside a spreadsheet
  POST /api/sheets/<sheet_id>/preview     Count Review Live Links in a tab
  POST /api/sheets/<sheet_id>/preview-batch  Probe multiple tabs at once

All routes previously lived inline in server.py.
Registered via create_sheets_blueprint().
"""

from __future__ import annotations

import threading
from pathlib import Path

from flask import Blueprint, jsonify, request


def create_sheets_blueprint(resources_path: Path) -> Blueprint:
    """Factory that returns the sheets Blueprint.

    Args:
        resources_path: The server-level RESOURCES_PATH (Path object).
    """
    bp = Blueprint('sheets', __name__)

    from shared import sheets_integration as _sheets_int

    @bp.route('/api/sheets/status', methods=['GET'])
    def sheets_status():
        return jsonify(_sheets_int.status(resources_path))

    @bp.route('/api/sheets/authorize', methods=['POST'])
    def sheets_authorize():
        """Open the OAuth consent flow in the user's default browser.
        Runs in a background thread with a 3-min wall-clock cap."""
        state = {'done': False, 'result': None}

        def _worker():
            try:
                state['result'] = _sheets_int.reauthorize(resources_path)
            except Exception as e:
                state['result'] = {'success': False, 'message': str(e)}
            state['done'] = True

        t = threading.Thread(target=_worker, daemon=True, name='sheets-auth')
        t.start()
        t.join(timeout=180)
        if not state['done']:
            return jsonify({'success': False, 'message': 'OAuth timeout'})
        return jsonify(state['result'])

    @bp.route('/api/sheets/list', methods=['GET'])
    def sheets_list():
        q = request.args.get('q', '').strip()
        return jsonify(_sheets_int.list_spreadsheets(resources_path, q))

    @bp.route('/api/sheets/<sheet_id>/tabs', methods=['GET'])
    def sheets_get_tabs(sheet_id):
        return jsonify(_sheets_int.get_tabs(resources_path, sheet_id))

    @bp.route('/api/sheets/<sheet_id>/preview', methods=['POST'])
    def sheets_preview(sheet_id):
        """Return how many 'Review Live Link' URLs are in the chosen tab."""
        body     = request.get_json(silent=True) or {}
        tab_name = (body.get('tab_name') or '').strip()
        if not tab_name:
            return jsonify({'success': False, 'message': 'tab_name required'}), 400

        res = _sheets_int.read_column_by_header(
            resources_path, sheet_id, tab_name, 'Review Live Link',
        )
        if not res.get('success'):
            return jsonify(res)

        rows       = res.get('rows') or []
        seen       = set()
        duplicates = 0
        first_row  = None
        last_row   = None

        for ri, val in rows:
            v = val.lower().strip()
            if not (v.startswith('http://') or v.startswith('https://')):
                continue
            if first_row is None:
                first_row = ri
            last_row = ri
            if v in seen:
                duplicates += 1
            else:
                seen.add(v)

        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(res['col']) if res.get('col') else None

        return jsonify({
            'success':       True,
            'total_links':   sum(1 for _, v in rows if v.lower().startswith(('http://', 'https://'))),
            'unique_links':  len(seen),
            'duplicates':    duplicates,
            'header_column': 'Review Live Link',
            'column_letter': col_letter,
            'first_row':     first_row,
            'last_row':      last_row,
        })

    @bp.route('/api/sheets/<sheet_id>/preview-batch', methods=['POST'])
    def sheets_preview_batch(sheet_id):
        """Probe many tabs in a single Sheets API call.
        Body: {tabs: [name,...]}. Prevents N parallel reads from blowing the quota."""
        body = request.get_json(silent=True) or {}
        tabs = body.get('tabs') or []
        if not isinstance(tabs, list):
            return jsonify({'success': False, 'message': 'tabs must be a list'}), 400
        clean = [str(t).strip() for t in tabs if str(t or '').strip()]
        return jsonify(_sheets_int.batch_count_links_by_header(
            resources_path, sheet_id, clean, 'Review Live Link',
        ))

    return bp
