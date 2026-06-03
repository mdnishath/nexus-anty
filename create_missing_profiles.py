"""
Create the 11 missing profiles in NST + add them to profiles.json with
their original credentials (from xlsx exports) and a fresh proxy
session ID per profile.
"""
import json, os, glob, secrets, sys, time, shutil
from pathlib import Path
from datetime import datetime
from collections import Counter
import openpyxl, requests

sys.stdout.reconfigure(encoding='utf-8')

NST_KEY  = json.loads(open('config/browser.json', encoding='utf-8').read())['nst_api_key']
NST_BASE = 'http://localhost:8848/api/v2'
HEADERS  = {'x-api-key': NST_KEY, 'Content-Type': 'application/json'}

PF = Path(os.environ['APPDATA']) / 'MailNexusPro' / 'profiles' / 'profiles.json'

TARGET = {'beccaobergefell.com', 'kiteesurf.com',
          'banglaminute.com', 'cozblog.com'}

# Proxy template — only the sessid-XXXX part is replaced per profile
PROXY_USER_TEMPLATE = '437dee9f2a569a71ecf2__cr.fr__sessid-{sess}'
PROXY_PASS = '9182e8e0f3678824'
PROXY_HOST = '74.81.81.81'
PROXY_PORT = '10000'


def fresh_session_id() -> str:
    return secrets.token_hex(6)        # 12-hex-char session id, fresh each time


# ── 1. Collect creds for ALL 4-domain profiles from xlsx ───────────────────
print('Scanning XLSX for all domain credentials...')
creds_by_email = {}
for d in (r'C:\Users\nisha\OneDrive\Desktop', r'C:\Users\nisha\Downloads'):
    for path in glob.glob(d + r'\**\*.xlsx', recursive=True):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            try:
                headers = [str(c).lower() if c else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
            except StopIteration:
                continue
            em_idx = pw_idx = totp_idx = bc_idx = grp_idx = -1
            for i, h in enumerate(headers):
                if em_idx < 0 and 'email' in h: em_idx = i
                if pw_idx < 0 and 'password' in h: pw_idx = i
                if totp_idx < 0 and ('totp' in h or 'secret' in h): totp_idx = i
                if bc_idx < 0 and 'backup' in h: bc_idx = i
                if grp_idx < 0 and 'group' in h: grp_idx = i
            if em_idx < 0 or pw_idx < 0: continue
            for r in ws.iter_rows(min_row=2, values_only=True):
                em = r[em_idx] if em_idx < len(r) else None
                pw = r[pw_idx] if pw_idx < len(r) else None
                if not em or not pw: continue
                em = str(em).strip().lower()
                if '@' not in em: continue
                if em.split('@',1)[1] not in TARGET: continue
                if em in creds_by_email and creds_by_email[em].get('totp'): continue
                creds_by_email[em] = {
                    'email':    str(r[em_idx]).strip(),
                    'password': str(pw).strip() if pw else '',
                    'totp':     (str(r[totp_idx]).strip() if 0 <= totp_idx < len(r) and r[totp_idx] else ''),
                    'backup':   (str(r[bc_idx]).strip() if 0 <= bc_idx < len(r) and r[bc_idx] else ''),
                    'group':    (str(r[grp_idx]).strip() if 0 <= grp_idx < len(r) and r[grp_idx] else 'Default'),
                }
print(f'  Total credentials in xlsx: {len(creds_by_email)}')

# ── 2. Determine which are missing from NST ───────────────────────────────
existing = json.loads(PF.read_text('utf-8'))
existing_emails = {(p.get('email') or '').strip().lower() for p in existing}
missing = [em for em in creds_by_email if em not in existing_emails]
print(f'  Already in profiles.json: {len(creds_by_email) - len(missing)}')
print(f'  Missing (need to create): {len(missing)}')

if not missing:
    print('Nothing to create. Done.')
    sys.exit(0)

# ── 3. Sanity: track existing proxy session IDs to guarantee freshness ────
used_sessions = set()
for p in existing:
    px = p.get('proxy') or {}
    u = px.get('username', '')
    if 'sessid-' in u:
        used_sessions.add(u.split('sessid-')[-1].split(':')[0])
print(f'  Existing proxy sessions tracked: {len(used_sessions)}')


def unique_session() -> str:
    while True:
        s = fresh_session_id()
        if s not in used_sessions:
            used_sessions.add(s)
            return s


# ── 4. Helper: NST profile creation ───────────────────────────────────────
def create_nst(name: str, note: str, proxy_url: str) -> dict | None:
    body = {
        'name': name,
        'platform': 'windows',
        'kernelMilestone': '146',
        'note': note,
        'fingerprint': {
            'flags': {
                'audio': 'Noise', 'battery': 'Masked', 'canvas': 'Noise',
                'clientRect': 'Noise', 'fonts': 'Masked',
                'geolocation': 'Custom', 'geolocationPopup': 'Prompt',
                'gpu': 'Allow', 'localization': 'Custom',
                'mediaDevices': 'Real', 'screen': 'Custom',
                'speech': 'Masked', 'timezone': 'Custom',
                'webgl': 'Noise', 'webrtc': 'Masked',
            },
            'screen': {'width': 1366, 'height': 768},
            'deviceMemory': 8,
            'hardwareConcurrency': 8,
            'navigator': {'language': 'fr-FR', 'languages': ['fr-FR', 'en-US']},
        },
        'dnsServer': '',
        'args': {'--disable-features': 'DnsOverHttps,AsyncDns',
                 '--dns-over-https-mode': 'off'},
        'proxy': proxy_url,
    }
    try:
        r = requests.post(f'{NST_BASE}/profiles', json=body, headers=HEADERS, timeout=30)
        d = r.json()
        if d.get('code') == 200 and d.get('msg', '').lower() == 'success':
            return d['data']
        print(f'  NST error for {name}: {d.get("msg")} (code={d.get("code")})')
    except Exception as e:
        print(f'  NST exception for {name}: {e}')
    return None


# ── 5. Build local profile dict ───────────────────────────────────────────
def to_local(nst: dict, cred: dict, proxy_dict: dict) -> dict:
    pid = nst.get('profileId') or nst['_id']
    name = nst.get('name', '') or pid[:8]

    bc = (cred.get('backup') or '').replace('\n', ' ').strip()
    bc_list = [c.strip() for c in bc.split(',') if c.strip()] if ',' in bc else (
        [bc] if bc else []
    )
    return {
        'id': pid,
        'nst_profile_id': pid,
        'engine': 'nst',
        'name': name,
        'email': cred['email'],
        'group': cred.get('group') or 'Default',
        'status': 'not_logged_in',
        'created_at': nst.get('createdAt') or datetime.now().isoformat(timespec='seconds'),
        'last_used': None,
        'tags': [],
        'notes': cred['email'],
        'profile_dir': str(Path(os.environ['APPDATA']) / 'MailNexusPro' / 'profiles' / pid),
        'proxy': proxy_dict,
        'overview': {'name': name, 'group': cred.get('group') or 'Default', 'startup_urls': []},
        'fingerprint': {'id': nst.get('fingerprintId', '')} if nst.get('fingerprintId') else {},
        'advanced': {'save_tabs': True},
        'proxy_timezone': 'Europe/Paris',
        'password': cred['password'],
        'totp_secret': cred.get('totp') or '',
        'backup_codes': bc_list,
        'recovery_email': '',
        'recovery_phone': '',
        'address': '',
    }


# ── 6. Create each missing profile ────────────────────────────────────────
backup = PF.with_suffix(f'.json.bak.before_create_missing.{int(time.time())}')
shutil.copy2(PF, backup)
print(f'\nBackup: {backup.name}\n')

print('Creating missing profiles in NST...')
created = []
failed = []
for em in missing:
    cred = creds_by_email[em]
    sess = unique_session()
    user = PROXY_USER_TEMPLATE.format(sess=sess)
    proxy_url  = f'socks5://{user}:{PROXY_PASS}@{PROXY_HOST}:{PROXY_PORT}'
    proxy_dict = {
        'protocol': 'socks5',
        'host': PROXY_HOST,
        'port': PROXY_PORT,
        'username': user,
        'password': PROXY_PASS,
    }
    name = em.split('@', 1)[0]    # use the email prefix as profile name

    nst = create_nst(name, em, proxy_url)
    if not nst:
        failed.append(em)
        continue

    local = to_local(nst, cred, proxy_dict)
    created.append(local)
    print(f'  + {em}  →  {nst["profileId"][:8]}  sess={sess}')
    time.sleep(0.5)   # be gentle on NST API

print()
print(f'Created: {len(created)}  Failed: {len(failed)}')
if failed:
    print(f'  Failed emails: {failed}')

# ── 7. Append to profiles.json ────────────────────────────────────────────
existing.extend(created)
PF.write_text(json.dumps(existing, indent=2, default=str), 'utf-8')
print(f'\nWrote {len(existing)} profiles total to {PF}')

# ── 8. Verify proxy uniqueness across the WHOLE file ──────────────────────
sigs = []
for p in existing:
    px = p.get('proxy') or {}
    if px.get('host'):
        sigs.append(f"{px['protocol']}://{px['username']}:{px['password']}@{px['host']}:{px['port']}")
c = Counter(sigs)
dupes = {k: v for k, v in c.items() if v > 1}
print(f'Proxies: {len(sigs)} total, {len(c)} unique')
if dupes:
    print('DUPLICATE PROXIES:')
    for s, n in list(dupes.items())[:5]:
        print(f'  {n}x  {s[:80]}')
else:
    print('All proxies unique')
