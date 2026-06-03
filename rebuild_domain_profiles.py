"""
One-shot recovery: rebuild profiles.json for the 4 domain sets
(beccaobergefell, kiteesurf, banglaminute, cozblog) by merging
exported XLSX credentials with NST API metadata (proxy, profileId,
fingerprint).
"""
import json, os, glob, shutil, time, sys
from pathlib import Path
from datetime import datetime
import openpyxl, requests

sys.stdout.reconfigure(encoding='utf-8')

# ── 1. Load NST data (matches by email) ──────────────────────────────────────
NST_KEY = json.loads(Path('config/browser.json').read_text('utf-8'))['nst_api_key']
HEADERS = {'x-api-key': NST_KEY}

print('Fetching all NST profiles...')
nst_by_email = {}
nst_by_name = {}
page = 1
while True:
    r = requests.get('http://localhost:8848/api/v2/profiles/',
                     params={'page': page, 'pageSize': 100}, headers=HEADERS, timeout=30)
    docs = r.json().get('data', {}).get('docs') or []
    if not docs: break
    for d in docs:
        note = (d.get('note') or '').strip().lower()
        nm   = (d.get('name') or '').strip().lower()
        if '@' in note:
            nst_by_email[note] = d
        if nm:
            nst_by_name[nm] = d
    if len(docs) < 100: break
    page += 1
print(f'  NST profiles: by-email={len(nst_by_email)} by-name={len(nst_by_name)}')

# ── 2. Scan every xlsx for domain credentials ────────────────────────────────
TARGET_DOMAINS = {'beccaobergefell.com', 'kiteesurf.com',
                  'banglaminute.com', 'cozblog.com'}

print('\nScanning XLSX files for domain credentials...')
creds = {}
for d in (r'C:\Users\nisha\OneDrive\Desktop', r'C:\Users\nisha\Downloads'):
    for path in glob.glob(d + r'\**\*.xlsx', recursive=True):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            try:
                headers = [str(c).lower() if c else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
            except StopIteration:
                continue
            em_idx = pw_idx = totp_idx = bc_idx = grp_idx = pxy_idx = -1
            for i, h in enumerate(headers):
                if em_idx < 0 and ('email' in h):    em_idx = i
                if pw_idx < 0 and ('password' in h): pw_idx = i
                if totp_idx < 0 and ('totp' in h or 'secret' in h): totp_idx = i
                if bc_idx < 0 and ('backup' in h):   bc_idx = i
                if grp_idx < 0 and ('group' in h):   grp_idx = i
                if pxy_idx < 0 and ('proxy' in h):   pxy_idx = i
            if em_idx < 0 or pw_idx < 0:
                continue
            for r in ws.iter_rows(min_row=2, values_only=True):
                em = r[em_idx] if em_idx < len(r) else None
                pw = r[pw_idx] if pw_idx < len(r) else None
                if not em or not pw: continue
                em = str(em).strip().lower()
                if '@' not in em: continue
                domain = em.split('@', 1)[1]
                if domain not in TARGET_DOMAINS: continue
                # Already have it from a more-complete source? skip
                if em in creds and creds[em].get('totp'): continue
                creds[em] = {
                    'email':    str(r[em_idx]).strip(),
                    'password': str(pw).strip() if pw else '',
                    'totp':     (str(r[totp_idx]).strip() if 0 <= totp_idx < len(r) and r[totp_idx] else ''),
                    'backup':   (str(r[bc_idx]).strip() if 0 <= bc_idx < len(r) and r[bc_idx] else ''),
                    'group':    (str(r[grp_idx]).strip() if 0 <= grp_idx < len(r) and r[grp_idx] else 'Default'),
                    'source':   Path(path).name,
                }
        wb.close()

print(f'  Domain credentials collected: {len(creds)}')
from collections import Counter
print(f'  By domain: {Counter(k.split("@",1)[1] for k in creds)}')

# ── 3. Match credentials to NST profile data, build local profile entries ────
def map_to_local(nst, cred):
    pid  = nst.get('profileId') or nst['_id']
    name = nst.get('name', '') or pid[:8]
    pcfg = nst.get('proxyConfig') or {}
    proxy = None
    if pcfg.get('host'):
        proxy = {
            'protocol': pcfg.get('protocol', 'http'),
            'host': pcfg.get('host', ''),
            'port': pcfg.get('port', ''),
            'username': pcfg.get('username', ''),
            'password': pcfg.get('password', ''),
        }
    pxy_tz = ((nst.get('proxyResult') or {}).get('timezone')) or ''
    fp_id = nst.get('fingerprintId', '')

    # Backup codes string → list
    bc_str = (cred.get('backup') or '').replace('\n', ' ').strip()
    bc_list = [c.strip() for c in bc_str.split(',') if c.strip()] if ',' in bc_str else (
        [bc_str] if bc_str else []
    )

    return {
        'id': pid,
        'nst_profile_id': pid,
        'engine': 'nst',
        'name': name,
        'email': cred['email'],
        'group': cred.get('group') or (nst.get('group') or {}).get('name') or 'Default',
        'status': 'logged_in',
        'created_at': nst.get('createdAt') or datetime.now().isoformat(timespec='seconds'),
        'last_used': nst.get('lastLaunchedAt'),
        'tags': nst.get('tags') or [],
        'notes': nst.get('note') or '',
        'profile_dir': str(Path(os.environ['APPDATA']) / 'MailNexusPro' / 'profiles' / pid),
        'proxy': proxy,
        'overview': {'name': name,
                     'group': (nst.get('group') or {}).get('name') or 'Default',
                     'startup_urls': nst.get('startupUrls') or []},
        'fingerprint': {'id': fp_id} if fp_id else {},
        'advanced': {'save_tabs': True},
        'proxy_timezone': pxy_tz,
        'password': cred['password'],
        'totp_secret': cred.get('totp') or '',
        'backup_codes': bc_list,
        'recovery_email': '',
        'recovery_phone': '',
        'address': '',
    }

matched = []
unmatched_credentials = []
for em, cred in creds.items():
    nst = nst_by_email.get(em)
    if not nst:
        # Try matching by name: email "alice.smith@x.com" → name "alice.smith"
        local = em.split('@')[0]
        nst = nst_by_name.get(local)
    if not nst:
        unmatched_credentials.append(em)
        continue
    matched.append(map_to_local(nst, cred))

print(f'\nMatched to NST: {len(matched)}')
print(f'Unmatched credentials (no NST profile found): {len(unmatched_credentials)}')
if unmatched_credentials[:5]:
    print(f'  e.g.: {unmatched_credentials[:5]}')

# ── 4. Merge into existing profiles.json ─────────────────────────────────────
PROFILES_FILE = Path(os.environ['APPDATA']) / 'MailNexusPro' / 'profiles' / 'profiles.json'

existing = []
if PROFILES_FILE.exists() and PROFILES_FILE.stat().st_size > 2:
    existing = json.loads(PROFILES_FILE.read_text('utf-8'))
existing_ids = {p.get('id') for p in existing}

# For each restored profile, REPLACE if same id exists, else APPEND
by_id = {p.get('id'): p for p in existing}
for r in matched:
    by_id[r['id']] = r

merged = list(by_id.values())

# Backup current profiles.json
backup = PROFILES_FILE.with_suffix(f'.json.bak.before_domain_restore.{int(time.time())}')
shutil.copy2(PROFILES_FILE, backup)
print(f'\nBackup saved: {backup.name}')

PROFILES_FILE.write_text(json.dumps(merged, indent=2, default=str), 'utf-8')
print(f'Wrote {len(merged)} total profiles to {PROFILES_FILE}')
print(f'  ({len(matched)} domain profiles with credentials restored)')

# ── 5. Verify ────────────────────────────────────────────────────────────────
print('\n=== Verification ===')
verify = json.loads(PROFILES_FILE.read_text('utf-8'))
domain_with_pw = [p for p in verify
                  if (p.get('email') or '').lower().split('@', 1)[-1] in TARGET_DOMAINS
                  and (p.get('password') or '').strip()]
print(f'Domain profiles with password set: {len(domain_with_pw)}')
groups = Counter((p.get('group') or 'default') for p in domain_with_pw)
for g, c in groups.most_common():
    print(f'  group "{g}": {c}')
