"""
Flask API Server for Electron Frontend
Provides REST API for Gmail Bot operations
"""

from flask import Flask, Blueprint, request, jsonify, Response, send_from_directory, send_file
from flask_cors import CORS
import logging
import secrets
import threading
import sys
import os
from datetime import datetime

# Force stdout/stderr to UTF-8 so unicode log chars (✓ → etc.) never crash on Windows cp1252
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import atexit
from pathlib import Path
import pandas as pd
import time
import json
import itertools

# ── Suppress Werkzeug HTTP access logs ────────────────────────────────────────
# Flask's dev server (Werkzeug) writes every request to stderr by default.
# Electron captures stderr as [ERR] and floods the UI log panel with noise like:
#   127.0.0.1 - - [08/Mar/2026 05:59:39] "GET /api/health HTTP/1.1" 200 -
# Only show genuine errors (500s, crashes), not routine 200-OK access logs.
logging.getLogger('werkzeug').setLevel(logging.ERROR)

# RESOURCES_PATH is set by main.js (Electron):
#   - In dev mode: the gmail_boat/ root folder (one above electron-app/)
#   - In packaged EXE: process.resourcesPath where bot scripts are bundled
# Fallback: go 3 levels up from this file (dev default)
RESOURCES_PATH = Path(os.environ.get(
    'RESOURCES_PATH',
    str(Path(__file__).parent.parent.parent)
))

sys.path.insert(0, str(RESOURCES_PATH))

# Screenshots / auth-key txt files directory
# Playwright saves screenshots with relative paths like "screenshots/...",
# which resolves to CWD/screenshots/.  In production (PyInstaller EXE),
# __file__ points to a temp _MEI* extraction dir, so we must use CWD instead.
if getattr(sys, 'frozen', False):
    # Production: backend.exe — CWD is the Electron app's install directory
    SCREENSHOTS_PATH = Path.cwd() / 'screenshots'
else:
    # Development: python server.py — CWD is typically electron-app/
    SCREENSHOTS_PATH = Path(__file__).parent.parent / 'screenshots'


def _cleanup_old_screenshots(max_age_days=7):
    """Delete screenshots older than max_age_days to prevent disk bloat."""
    try:
        if not SCREENSHOTS_PATH.exists():
            return
        cutoff = time.time() - (max_age_days * 86400)
        removed = 0
        for f in SCREENSHOTS_PATH.rglob('*.png'):
            try:
                if f.stat().st_mtime < cutoff:
                    f.unlink()
                    removed += 1
            except Exception:
                pass
        if removed:
            print(f"[CLEANUP] Removed {removed} screenshots older than {max_age_days} days", flush=True)
    except Exception:
        pass

# Run cleanup on import (server startup)
_cleanup_old_screenshots()

# Lazy-loaded: prepare_excel_with_common_settings (heavy pandas/openpyxl imports)
_prepare_excel_fn = None
def prepare_excel_with_common_settings(*args, **kwargs):
    global _prepare_excel_fn
    if _prepare_excel_fn is None:
        from prepare_excel_with_common_settings import prepare_excel_with_common_settings as _fn
        _prepare_excel_fn = _fn
    return _prepare_excel_fn(*args, **kwargs)

app = Flask(__name__)
CORS(app, supports_credentials=False)


# Always return JSON errors so the UI doesn't see HTML and complain
# about "Non-JSON response. Restart backend." Logs the full traceback
# server-side for debugging.
@app.errorhandler(404)
def _json_404(_e):
    return jsonify({
        'success': False,
        'message': f'Route not found: {request.method} {request.path}',
    }), 404


@app.errorhandler(500)
def _json_500(e):
    import traceback as _tb
    err_str = str(e) or repr(e) or type(e).__name__
    try:
        _tb.print_exc()
    except Exception:
        pass
    return jsonify({'success': False, 'message': f'Server error: {err_str}'}), 500


@app.errorhandler(Exception)
def _json_exc(e):
    """Catch-all so unexpected raises in routes still produce JSON."""
    import traceback as _tb
    err_str = str(e) or repr(e) or type(e).__name__
    try:
        _tb.print_exc()
    except Exception:
        pass
    return jsonify({
        'success': False,
        'message': f'{type(e).__name__}: {err_str}',
    }), 500

# ── Auth (Internal API Token) ────────────────────────────────────────────────
# Generates a random token on startup. Electron captures this from stdout.
# Protects against localhost CSRF/abuse by unauthorized scripts.

_INTERNAL_TOKEN = secrets.token_hex(32)
_PUBLIC_PATHS = {
    '/api/health', '/api/app/version', '/api/auth/status',
    '/api/license/info', '/api/license/activate', '/api/license/deactivate'
}

@app.before_request
def _gate_internal_api():
    """Require the internal token for all /api/ endpoints (except whitelist)."""
    p = request.path or ''
    if not p.startswith('/api/'):
        return None
        
    if p in _PUBLIC_PATHS:
        return None
        
    # Check header (standard fetch) or query param (for EventSource/SSE)
    provided = request.headers.get('X-Api-Token') or request.args.get('token')
    
    if not provided or provided != _INTERNAL_TOKEN:
        return jsonify({'success': False, 'message': 'Unauthorized (invalid internal token)'}), 401

# ── App Version & Update Check ───────────────────────────────────────────────
_FALLBACK_VERSION = '2.0.0'  # used when package.json isn't readable (frozen exe)


def _read_app_version():
    """Read version from package.json (stays in sync with build.bat).

    In the packaged exe __file__ points inside the PyInstaller temp dir, so the
    package.json lookup fails — fall back to the embedded version constant so
    the About page still shows the right number."""
    # Dev mode: read the real electron-app/package.json next to the backend.
    try:
        pkg = Path(__file__).resolve().parent.parent / 'package.json'
        if pkg.exists():
            return json.loads(pkg.read_text(encoding='utf-8')).get('version', _FALLBACK_VERSION)
    except Exception:
        pass
    # Packaged mode: try the config dir shipped alongside the exe.
    try:
        import os as _os
        res = _os.environ.get('RESOURCES_PATH', '')
        if res:
            pkg2 = Path(res) / 'package.json'
            if pkg2.exists():
                return json.loads(pkg2.read_text(encoding='utf-8')).get('version', _FALLBACK_VERSION)
    except Exception:
        pass
    return _FALLBACK_VERSION


APP_VERSION = _read_app_version()
_version_cache = {'last_check': 0, 'data': None}


def _get_version_manifest_url():
    """Read version_manifest_url from tools/gist_config.json (written by admin panel)."""
    cfg_path = RESOURCES_PATH / 'tools' / 'gist_config.json'
    try:
        if cfg_path.exists():
            cfg = json.loads(cfg_path.read_text(encoding='utf-8'))
            return cfg.get('version_manifest_url', '')
    except Exception:
        pass
    return ''



# Auto-incrementing log ID (no collisions unlike time*1000)
_log_id = itertools.count(1)

# Global state
processing_state = {
    'status': 'idle',  # idle, processing, completed, stopped
    'current': 0,
    'total': 0,
    'current_account': '',
    'recent_logs': [],
    'log_clear_id': 0,   # highest log id at the time of last clear; SSE only replays logs after this
    'file_path': '',  # Input file path
    'output_file_path': '',  # Output file path (for progress tracking)
    'step_name': '',   # Current step(s) being processed (e.g. 'step1', 'step2', 'step1+step2')
    'step_label': '',  # Human-readable step label for UI (e.g. 'Step 2 - Security')
    'operations': '',  # Current operations string (e.g. '1,4,5' or 'A3')
}

processing_thread = None
_processing_lock = threading.Lock()  # Prevent double-start race condition
stop_flag = threading.Event()

# ── In-memory progress counters ──────────────────────────────────────────────
# Updated by _spawn_and_stream() as it parses subprocess stdout lines.
# get_progress() reads from here during processing to avoid reading Excel.
_progress_counters = {
    'total': 0,
    'success': 0,
    'failed': 0,
}
_progress_lock = threading.Lock()  # Lightweight lock for counters only

# SMS code relay storage (received from phone app)
sms_codes = []


def _is_safe_child(child: Path, parent: Path) -> bool:
    """Return True if child is strictly inside parent (no traversal)."""
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _persist_output_dirs(output_file_path='', input_file_path=''):
    """Persist the latest output/input paths to disk so they survive server restarts.
    Called after processing completes (or whenever output_file_path is captured)."""
    try:
        state_file = RESOURCES_PATH / 'config' / 'last_output.json'
        state_file.parent.mkdir(parents=True, exist_ok=True)

        # Read existing state and merge (keep any previously saved dirs)
        existing = {}
        if state_file.exists():
            with open(state_file, 'r', encoding='utf-8') as f:
                existing = json.load(f)

        dirs = set(existing.get('output_dirs', []))
        if output_file_path:
            d = str(Path(output_file_path).parent.resolve())
            if Path(d).exists():
                dirs.add(d)
        if input_file_path:
            d = str((Path(input_file_path).parent.parent / 'output').resolve())
            if Path(d).exists():
                dirs.add(d)

        # Only keep dirs that actually exist
        dirs = [d for d in dirs if Path(d).exists()]

        with open(state_file, 'w', encoding='utf-8') as f:
            json.dump({'output_dirs': dirs}, f, indent=2)
    except Exception as e:
        print(f"[WARN] Could not persist output dirs: {e}", flush=True)


def _load_persisted_output_dirs():
    """Load previously saved output directories from disk (survives restarts)."""
    try:
        state_file = RESOURCES_PATH / 'config' / 'last_output.json'
        if state_file.exists():
            with open(state_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [d for d in data.get('output_dirs', []) if Path(d).exists()]
    except Exception:
        pass
    return []


def _get_report_scan_dirs():
    """Return all directories where output/report .xlsx files may live.

    Sources (in order):
      1. electron-app/output — dev mode CWD writes
      2. RESOURCES_PATH/output — production EXE mode
      3. Directory of the current processing output file (from [OUTPUT_FILE] marker)
      4. Derived from the input Excel path — same logic as ExcelProcessor
      5. Persisted output dirs from last_output.json (survives server restarts)
      6. User home / output — common fallback for ExcelProcessor outputs"""
    scan_dirs = set()

    # 1. electron-app/output — where subprocess CWD writes in dev mode
    electron_output = Path(__file__).parent.parent / 'output'
    electron_output.mkdir(exist_ok=True)
    scan_dirs.add(str(electron_output.resolve()))

    # 2. RESOURCES_PATH/output — where files go in production EXE mode
    project_output = RESOURCES_PATH / 'output'
    if project_output.exists():
        scan_dirs.add(str(project_output.resolve()))

    # 3. Directory of the current/latest output file (set by subprocess via [OUTPUT_FILE])
    output_file = processing_state.get('output_file_path', '')
    if output_file:
        output_dir = Path(output_file).parent
        if output_dir.exists():
            scan_dirs.add(str(output_dir.resolve()))

    # 4. Derived from input Excel path — mirrors ExcelProcessor logic:
    #    base_dir = Path(excel).parent.parent  →  output_dir = base_dir / "output"
    input_file = processing_state.get('file_path', '')
    if input_file:
        derived_output = Path(input_file).parent.parent / 'output'
        if derived_output.exists():
            scan_dirs.add(str(derived_output.resolve()))

    # 5. Persisted output dirs from last_output.json (survive server restarts)
    for d in _load_persisted_output_dirs():
        scan_dirs.add(d)

    # 6. User home / output — common fallback (ExcelProcessor often writes here)
    user_home_output = Path.home() / 'output'
    if user_home_output.exists():
        scan_dirs.add(str(user_home_output.resolve()))

    # 7. Profile Manager reports — use resolved path (handles empty storage_path config)
    try:
        # _get_storage_path() always returns the real resolved path regardless of config
        pm_reports = profile_manager._get_storage_path() / 'reports'
        pm_reports.mkdir(parents=True, exist_ok=True)
        scan_dirs.add(str(pm_reports.resolve()))
    except Exception:
        pass

    # 7b. Also honour any custom storage_path set in config
    try:
        pm_storage = profile_manager.get_config().get('storage_path', '')
        if pm_storage:
            custom_reports = Path(pm_storage) / 'reports'
            custom_reports.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(custom_reports.resolve()))
    except Exception:
        pass

    # 8. AppData/Roaming/MailNexusPro fallback (default on Windows)
    try:
        appdata_roaming = os.environ.get('APPDATA', '')
        if appdata_roaming:
            mnp = Path(appdata_roaming) / 'MailNexusPro' / 'profiles' / 'reports'
            mnp.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(mnp.resolve()))
    except Exception:
        pass

    # 9. AppData/Local/GmailBotPro legacy fallback
    try:
        appdata_local = Path(os.environ.get('LOCALAPPDATA', '')) / 'GmailBotPro' / 'browser_profiles' / 'reports'
        if appdata_local.parent.exists():
            appdata_local.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(appdata_local.resolve()))
    except Exception:
        pass

    return scan_dirs


def _find_latest_output_file():
    """Find the most recent *_output_*.xlsx file across all known directories.
    Used when processing_state has no output_file_path (e.g. after server restart)."""
    scan_dirs = _get_report_scan_dirs()
    best = None
    best_mtime = 0
    for d in scan_dirs:
        dp = Path(d)
        if not dp.exists():
            continue
        for f in dp.rglob('*_output_*.xlsx'):
            # Skip MailNexus reports (they are derived, not source)
            if 'mailnexus' in f.name.lower() or 'Mailnexus' in f.name:
                continue
            mt = f.stat().st_mtime
            if mt > best_mtime:
                best_mtime = mt
                best = f
    return str(best) if best else None


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'success': True, 'message': 'Server is running'})


@app.route('/api/app/version', methods=['GET'])
def app_version():
    """Return current app version."""
    return jsonify({'version': APP_VERSION})


@app.route('/api/app/check-update', methods=['GET'])
def app_check_update():
    """Check for updates by fetching the version manifest from GitHub Gist."""
    import urllib.request

    manifest_url = _get_version_manifest_url()
    if not manifest_url:
        return jsonify({'update_available': False, 'current_version': APP_VERSION,
                        'message': 'Update check not configured'})

    # Cache for 1 hour
    now = time.time()
    if now - _version_cache['last_check'] < 3600 and _version_cache['data'] is not None:
        manifest = _version_cache['data']
    else:
        try:
            req = urllib.request.Request(manifest_url, headers={
                'User-Agent': 'MailNexus-Pro/' + APP_VERSION,
                'Cache-Control': 'no-cache',
            })
            with urllib.request.urlopen(req, timeout=5) as resp:
                manifest = json.loads(resp.read().decode('utf-8'))
            _version_cache['data'] = manifest
            _version_cache['last_check'] = now
        except Exception:
            return jsonify({'update_available': False, 'current_version': APP_VERSION,
                            'message': 'Could not reach update server'})

    latest = manifest.get('latest_version', APP_VERSION)

    # Simple semver compare: split by '.', compare each part as int
    def ver_tuple(v):
        try:
            return tuple(int(x) for x in v.split('.'))
        except Exception:
            return (0, 0, 0)

    update_available = ver_tuple(latest) > ver_tuple(APP_VERSION)

    result = {
        'update_available': update_available,
        'current_version': APP_VERSION,
        'latest_version': latest,
        'download_url': manifest.get('download_url', ''),
        'release_date': manifest.get('release_date', ''),
        'changelog': manifest.get('changelog', []),
    }
    return jsonify(result)


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    """Shutdown the Flask server from Electron UI"""
    def wait_and_kill():
        time.sleep(1)
        # sys.exit triggers SystemExit which Flask can handle for cleanup
        # os._exit is used as final fallback only
        try:
            sys.exit(0)
        except SystemExit:
            os._exit(0)
    threading.Thread(target=wait_and_kill, daemon=True).start()
    return jsonify({'success': True, 'message': 'Server shutting down...'})


@app.route('/api/file-info', methods=['POST'])
def get_file_info():
    """Get Excel file statistics"""
    try:
        data = request.json or {}
        file_path = data.get('file_path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'})

        df = pd.read_excel(file_path, engine='openpyxl')

        # Add Status column if not present
        if 'Status' not in df.columns:
            df['Status'] = ''

        # Mark empty as PENDING
        df.loc[df['Status'].isna() | (df['Status'] == ''), 'Status'] = 'PENDING'

        total = len(df)
        success = len(df[df['Status'].str.upper() == 'SUCCESS'])
        failed = len(df[df['Status'].str.upper() == 'FAILED'])
        pending = len(df[df['Status'].str.upper() == 'PENDING'])

        return jsonify({
            'success': True,
            'data': {
                'total': total,
                'success': success,
                'failed': failed,
                'pending': pending
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/start-processing', methods=['POST'])
def start_processing():
    """Start processing accounts"""
    global processing_thread, stop_flag

    try:
        data = request.json or {}
        file_path = data.get('file_path')
        operations = data.get('operations', '1,4,5')
        new_password = data.get('new_password', '')
        recovery_email = data.get('recovery_email', '')
        recovery_phone = data.get('recovery_phone', '')
        num_workers = data.get('num_workers', 5)
        bot_step = int(data.get('bot_step', 2))  # UI explicit payload

        # Multi-step support
        bot_steps = data.get('bot_steps', [bot_step])  # array of step numbers
        linked = data.get('linked', False)              # linked mode toggle
        ops_per_step = data.get('ops_per_step', {})     # per-step operations map

        # Validation
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'})

        # Check if already processing (thread-safe)
        if not _processing_lock.acquire(blocking=False):
            return jsonify({'success': False, 'message': 'Already processing'})
        try:
            if processing_state['status'] == 'processing':
                return jsonify({'success': False, 'message': 'Already processing'})

            # Reset state — all inside lock to prevent race conditions
            stop_flag.clear()
            _last_progress.clear()  # Clear stale cache from previous run
            processing_state['status'] = 'processing'
            processing_state['current'] = 0
            processing_state['total'] = 0
            processing_state['current_account'] = ''
            processing_state['recent_logs'] = []
            processing_state['file_path'] = file_path
            processing_state['operations'] = operations
        finally:
            _processing_lock.release()

        # Reset in-memory counters
        with _progress_lock:
            _progress_counters['total'] = 0
            _progress_counters['success'] = 0
            _progress_counters['failed'] = 0

        is_multi = len(bot_steps) > 1
        step_label = '+'.join(str(s) for s in bot_steps)

        # Track step info for reports and UI
        _step_labels = {1: 'Language/Activity', 2: 'Security', 3: 'Maps Reviews', 4: 'Appeals'}
        if is_multi:
            processing_state['step_name'] = '+'.join(f'step{s}' for s in bot_steps)
            processing_state['step_label'] = 'Steps ' + '+'.join(f'{s}' for s in bot_steps)
        else:
            s = bot_steps[0]
            processing_state['step_name'] = f'step{s}'
            processing_state['step_label'] = f'Step {s} - {_step_labels.get(s, "")}'

        add_log('Preparing Excel file with common settings...', 'info')
        add_log(f'[DEBUG] File path: {file_path}', 'info')
        add_log(f'[DEBUG] Operations: {operations}', 'info')
        add_log(f'[DEBUG] Bot steps: {bot_steps} | Linked: {linked}', 'info')

        # Step 1: Prepare Excel
        try:
            add_log('[DEBUG] Calling prepare_excel_with_common_settings...', 'info')
            prepare_excel_with_common_settings(
                file_path,
                operations,
                new_password,
                recovery_email,
                recovery_phone
            )
            add_log(f'[DEBUG] prepare_excel_with_common_settings returned successfully', 'info')
            add_log(f'Excel prepared successfully', 'success')
        except Exception as prep_error:
            import traceback
            traceback.print_exc()
            add_log(f'ERROR preparing Excel: {str(prep_error)}', 'error')
            processing_state['status'] = 'idle'
            return jsonify({'success': False, 'message': f'Excel preparation failed: {str(prep_error)}'})

        add_log(f'[DEBUG] About to start background thread...', 'info')
        if is_multi and linked:
            add_log(f'Running LINKED Steps {step_label} matrix', 'info')
        elif is_multi:
            add_log(f'Running Steps {step_label} SEQUENTIALLY (unlinked)', 'info')
        else:
            add_log(f'Running STEP {bot_steps[0]} matrix', 'info')

        # Step 2: Start processing in background thread
        processing_thread = threading.Thread(
            target=run_processing_worker,
            args=(file_path, num_workers, bot_steps[0] if not is_multi else bot_steps[0]),
            kwargs={
                'bot_steps': bot_steps,
                'linked': linked,
                'ops_per_step': ops_per_step,
            },
            daemon=True,
            name=f'ProcessingWorker-{step_label}'
        )
        processing_thread.start()

        add_log('Background worker thread started successfully', 'success')
        add_log(f'Thread is alive: {processing_thread.is_alive()}', 'info')
        return jsonify({'success': True, 'message': 'Processing started'})

    except Exception as e:
        processing_state['status'] = 'idle'
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/stop-processing', methods=['POST'])
def stop_processing():
    """Stop processing"""
    try:
        stop_flag.set()
        with _processing_lock:
            processing_state['status'] = 'stopped'
        add_log('Stop requested - waiting for current account to finish...', 'warning')
        return jsonify({'success': True, 'message': 'Stopping...'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


_last_progress = {}  # Cache last successful progress read to avoid 0-glitch


@app.route('/api/progress', methods=['GET'])
def get_progress():
    """Get current progress - reads real-time from Excel"""
    global _last_progress
    try:
        # Snapshot processing_state under lock to prevent torn reads
        with _processing_lock:
            snap_status = processing_state['status']
            snap_output = processing_state.get('output_file_path', '')
            snap_file = processing_state.get('file_path', '')
            snap_ops = processing_state.get('operations', '')
            snap_account = processing_state.get('current_account', '')
            snap_step = processing_state.get('step_label', '')
            snap_logs = list(processing_state.get('recent_logs', []))

        # ── Use in-memory counters during processing ──────────────────
        # Only falls back to Excel on completed/stopped for final tally.
        if snap_status in ('processing', 'completed', 'stopped'):

            # During active processing, use in-memory counters (fast, no I/O)
            if snap_status == 'processing':
                with _progress_lock:
                    total = _progress_counters['total']
                    success = _progress_counters['success']
                    failed = _progress_counters['failed']

                # If total is still 0, do ONE quick read for row count at startup
                if total == 0:
                    excel_file = snap_output or snap_file
                    if excel_file:
                        try:
                            df = pd.read_excel(excel_file, engine='openpyxl')
                            total = len(df)
                            with _progress_lock:
                                _progress_counters['total'] = total
                        except Exception:
                            pass

                current = success + failed
                pending = max(0, total - current)

                # A3 Live Check mode: live_mode stays False during processing
                # (Live Check Status column is written per-row but reading it
                # per poll would cause file contention — final tally on completion
                # gives the accurate Live/Missing split via Excel read)
                is_a3_only = snap_ops.strip() == 'A3'

                prog = {
                    'status': snap_status,
                    'current': current,
                    'total': total,
                    'success': success,
                    'failed': failed,
                    'pending': pending,
                    'current_account': snap_account,
                    'step_label': snap_step,
                    'logs': snap_logs,
                    'live_mode': False,  # Full live_mode tally done on completion
                }
                _last_progress = prog
                return jsonify({'success': True, 'progress': prog})

            # On completed/stopped — do a final Excel read for accurate tally
            excel_file = snap_output or snap_file
            if excel_file:
                try:
                    df = pd.read_excel(excel_file, engine='openpyxl')

                    if 'Status' not in df.columns:
                        df['Status'] = ''

                    total = len(df)
                    success = len(df[df['Status'].str.upper() == 'SUCCESS'])
                    failed = len(df[df['Status'].str.upper() == 'FAILED'])
                    pending = len(df[(df['Status'].isna()) | (df['Status'] == '') | (df['Status'].str.upper() == 'PENDING')])
                    current = success + failed

                    with _processing_lock:
                        processing_state['total'] = total
                        processing_state['current'] = current

                    # Live Check mode: only when Step 4 A3 is the active operation
                    live_mode = False
                    live_count = 0
                    missing_count = 0
                    is_a3_only = snap_ops.strip() == 'A3'
                    if is_a3_only and 'Live Check Status' in df.columns:
                        lcs = df['Live Check Status'].fillna('').astype(str).str.strip().str.lower()
                        live_count = int((lcs == 'live').sum())
                        missing_count = int(current - live_count)
                        live_mode = True

                    prog = {
                            'status': snap_status,
                            'current': current,
                            'total': total,
                            'success': success if not live_mode else live_count,
                            'failed': failed if not live_mode else missing_count,
                            'pending': pending,
                            'current_account': snap_account,
                            'step_label': snap_step,
                            'logs': snap_logs,
                            'live_mode': live_mode,
                    }
                    _last_progress = prog
                    return jsonify({'success': True, 'progress': prog})
                except Exception as read_error:
                    # File locked — return cached values instead of 0s
                    if _last_progress:
                        _last_progress['status'] = snap_status
                        _last_progress['current_account'] = snap_account
                        _last_progress['logs'] = snap_logs
                        return jsonify({'success': True, 'progress': _last_progress})

        # Check if bulk re-login is running — show its progress on dashboard
        try:
            rl = profile_manager.get_bulk_relogin_status()
            if rl.get('running') or rl.get('status') in ('processing', 'completed'):
                total = rl.get('total', 0)
                done = rl.get('done', 0)
                pending = max(0, total - done)
                return jsonify({
                    'success': True,
                    'progress': {
                        'status': rl['status'],
                        'current': done,
                        'total': total,
                        'success': rl.get('success', 0),
                        'failed': rl.get('failed', 0),
                        'pending': pending,
                        'current_account': rl.get('current_account', ''),
                        'step_label': 'Bulk Re-Login',
                        'logs': processing_state.get('recent_logs', [])[-20:],
                        'job_type': 'bulk_relogin',
                        'report_path': rl.get('report_path', ''),
                        'success': rl.get('success', 0),
                        'failed': rl.get('failed', 0),
                    }
                })
        except Exception:
            pass

        # Check if batch login is running — show its progress on dashboard
        try:
            bl = profile_manager.get_batch_login_progress()
            if bl.get('running') or bl.get('status') in ('processing', 'completed'):
                done = bl.get('success', 0) + bl.get('failed', 0)
                return jsonify({
                    'success': True,
                    'progress': {
                        'status': bl['status'],
                        'current': done,
                        'total': bl.get('total', 0),
                        'success': bl.get('success', 0),
                        'failed': bl.get('failed', 0),
                        'pending': bl.get('pending', 0),
                        'current_account': bl.get('current_account', ''),
                        'step_label': 'Batch Login',
                        'logs': processing_state.get('recent_logs', [])[-20:],
                        'job_type': 'batch_login',
                    }
                })
        except Exception:
            pass

        # Default return (idle state)
        return jsonify({
            'success': True,
            'progress': {
                'status': processing_state['status'],
                'current': processing_state.get('current', 0),
                'total': processing_state.get('total', 0),
                'success': 0,
                'failed': 0,
                'pending': 0,
                'current_account': processing_state.get('current_account', ''),
                'step_label': processing_state.get('step_label', ''),
                'logs': processing_state.get('recent_logs', [])[-20:]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports', methods=['GET'])
def get_reports():
    """Get list of generated reports (MailNexus Pro + legacy).
    Only scans the project /output directory."""
    try:
        scan_dirs = _get_report_scan_dirs()

        if not scan_dirs:
            return jsonify({'success': True, 'reports': []})

        # Scan all directories, de-duplicate by file path
        seen_paths = set()
        reports = []
        for dir_path in scan_dirs:
            d = Path(dir_path)
            if not d.exists():
                continue
            for file in d.rglob('*.xlsx'):
                fpath = str(file.resolve())
                if fpath in seen_paths:
                    continue
                seen_paths.add(fpath)
                is_mailnexus = 'Mailnexus' in file.name or 'mailnexus' in file.name.lower()
                reports.append({
                    'name': file.name,
                    'path': str(file),
                    'size': file.stat().st_size,
                    'modified': file.stat().st_mtime,
                    'type': 'mailnexus' if is_mailnexus else 'legacy',
                })

        # Sort by modified time (newest first)
        reports.sort(key=lambda x: x['modified'], reverse=True)

        return jsonify({'success': True, 'reports': reports})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/generate', methods=['POST'])
def generate_mailnexus_report():
    """Generate MailNexus Pro report.
    - source='profiles': all profiles with appeal + health status
    - source='appeal':   only appeal tracking data
    - source='health':   only health activity tracking data
    - source='':         legacy — from step-processing output Excel file"""
    try:
        data = request.json or {}
        source = data.get('source', '')

        from shared.report_generator import generate_report, generate_from_excel
        from datetime import datetime as _dt

        def _fmt_dt(iso):
            try:
                return _dt.fromisoformat(iso.replace('Z', '+00:00')).strftime('%d %b %Y %H:%M')
            except Exception:
                return iso

        output_dir = profile_manager._get_storage_path() / 'reports'
        output_dir.mkdir(parents=True, exist_ok=True)

        if source in ('profiles', 'appeal', 'health'):
            profiles = profile_manager.list_profiles()
            if not profiles:
                return jsonify({'success': False, 'message': 'No profiles found.'})

            accounts_data = []

            if source == 'appeal':
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    ok_str   = '✓' if p.get('last_appeal_ok') else ('✗' if p.get('last_appeal_at') else '—')
                    last_run = _fmt_dt(p['last_appeal_at']) if p.get('last_appeal_at') else 'Never'
                    summary  = p.get('last_appeal_summary', '') or ''
                    history  = p.get('appeal_history', [])
                    history_str = ' | '.join(
                        f"{_fmt_dt(h['date'])} {'✓' if h.get('ok') else '✗'}" for h in history
                    ) if history else 'No history'
                    accounts_data.append({
                        'Email':        p.get('email', ''),
                        'Profile Name': p.get('name', ''),
                        'Login Status': login_status.replace('_', ' ').title(),
                        'Result':       ok_str,
                        'Last Appeal':  last_run,
                        'Summary':      summary,
                        'Run History':  history_str,
                    })
                label = 'appeal'

            elif source == 'health':
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    ok_str   = '✓' if p.get('last_health_ok') else ('✗' if p.get('last_health_at') else '—')
                    last_run = _fmt_dt(p['last_health_at']) if p.get('last_health_at') else 'Never'
                    done     = p.get('last_health_done', 0)
                    total    = p.get('last_health_total', 0)
                    history  = p.get('health_history', [])
                    history_str = ' | '.join(
                        f"{_fmt_dt(h['date'])} {h.get('done',0)}/{h.get('total',0)}" for h in history
                    ) if history else 'No history'
                    accounts_data.append({
                        'Email':           p.get('email', ''),
                        'Profile Name':    p.get('name', ''),
                        'Login Status':    login_status.replace('_', ' ').title(),
                        'Result':          ok_str,
                        'Last Health Run': last_run,
                        'Activities Done': f"{done}/{total}",
                        'Run History':     history_str,
                    })
                label = 'health'

            else:  # 'profiles' combined
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    status = 'SUCCESS' if login_status == 'logged_in' else (
                        'FAILED' if login_status == 'login_failed' else 'PENDING')
                    appeal_info = (
                        f"{'✓' if p.get('last_appeal_ok') else '✗'} {_fmt_dt(p['last_appeal_at'])}"
                        + (f" — {p.get('last_appeal_summary','')}" if p.get('last_appeal_summary') else '')
                        if p.get('last_appeal_at') else 'Never'
                    )
                    health_info = (
                        f"{'✓' if p.get('last_health_ok') else '✗'} {_fmt_dt(p['last_health_at'])} — {p.get('last_health_done',0)}/{p.get('last_health_total',0)} activities"
                        if p.get('last_health_at') else 'Never'
                    )
                    accounts_data.append({
                        'Email':                p.get('email', ''),
                        'Profile Name':         p.get('name', ''),
                        'Status':               status,
                        'Login':                login_status.replace('_', ' ').title(),
                        'Last Appeal':          appeal_info,
                        'Last Health Activity': health_info,
                        'Operations Done':      f"Appeal: {appeal_info} | Health: {health_info}",
                    })
                label = 'profiles'

            report_path = generate_report(
                output_dir=str(output_dir),
                accounts_data=accounts_data,
                step_name=label,
            )
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'{label.title()} report generated: {Path(report_path).name} ({len(accounts_data)} accounts)'
            })

        elif source == 'file':
            # ── Generate from a specific raw output file ──
            file_path = data.get('file_path', '')
            if not file_path or not Path(file_path).exists():
                return jsonify({'success': False, 'message': 'File not found.'})
            report_path = generate_from_excel(file_path)
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'Pro report generated: {Path(report_path).name}'
            })

        else:
            # ── Legacy: generate from latest output Excel ──
            output_file = (
                data.get('output_file')
                or processing_state.get('output_file_path')
                or _find_latest_output_file()
            )
            step_name = data.get('step', '')

            if not output_file or not Path(output_file).exists():
                return jsonify({'success': False, 'message': 'No output file found.'})

            report_path = generate_from_excel(output_file, step_name=step_name)
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'Pro report generated: {Path(report_path).name}'
            })

    except PermissionError:
        return jsonify({'success': False, 'message': 'Report file is open in Excel. Close it and try again.'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Report generation failed: {str(e)}'})


@app.route('/api/template/generate', methods=['POST'])
def generate_step_template():
    """Generate a blank XLS template for a step showing required input columns."""
    try:
        data = request.json or {}
        step = data.get('step')
        if not step:
            return jsonify({'success': False, 'message': 'Missing "step" parameter (1-4).'})

        step_name = f'step{int(step)}'
        output_dir = str(RESOURCES_PATH / 'templates')

        from shared.report_generator import generate_template
        template_path = generate_template(step_name, output_dir)

        return jsonify({
            'success': True,
            'template_path': str(template_path),
            'message': f'Template generated: {Path(template_path).name}'
        })
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Template generation failed: {str(e)}'})


def _auto_generate_report():
    """Auto-generate MailNexus Pro report from the output file after processing."""
    output_file = processing_state.get('output_file_path')
    if not output_file or not Path(output_file).exists():
        add_log('[REPORT] No output file available for report generation.', 'warning')
        return
    try:
        from shared.report_generator import generate_from_excel
        step_name = processing_state.get('step_name', '')
        # For multi-step (e.g. 'step1+step2'), use first step for column filtering
        if '+' in step_name:
            step_name = step_name.split('+')[0]
        report_path = generate_from_excel(output_file, step_name=step_name)
        add_log(f'[REPORT] MailNexus Pro report: {Path(report_path).name}', 'success')
        add_log(f'[REPORT] Saved to: {report_path}', 'info')

    except Exception as e:
        add_log(f'[REPORT] Auto-report generation failed: {e}', 'error')


def _build_step_cmd(bot_step, file_path, num_workers):
    """Build subprocess command for a single step."""
    import sys
    step_scripts = {
        1: 'gmail_bot_step1.py',
        2: 'gmail_bot_step2.py',
        3: 'gmail_bot_step3.py',
        4: 'gmail_bot_step4.py',
    }
    if getattr(sys, 'frozen', False):
        cmd = [sys.executable, f'--step{int(bot_step)}', file_path, str(num_workers)]
        cmd_display = f'backend.exe --step{int(bot_step)} ...'
    else:
        script_name = step_scripts.get(int(bot_step), 'gmail_bot_step2.py')
        script_path = RESOURCES_PATH / script_name
        if not script_path.exists():
            return None, None, f'Script not found: {script_path}'
        cmd = [sys.executable, str(script_path), file_path, str(num_workers)]
        cmd_display = f'python {script_name} ...'
    return cmd, cmd_display, None


def _build_linked_cmd(file_path, num_workers, steps_json):
    """Build subprocess command for linked multi-step mode."""
    import sys
    if getattr(sys, 'frozen', False):
        cmd = [sys.executable, '--linked', file_path, str(num_workers), steps_json]
        cmd_display = f'backend.exe --linked ...'
    else:
        script_path = RESOURCES_PATH / 'gmail_bot_linked.py'
        if not script_path.exists():
            return None, None, f'Script not found: {script_path}'
        cmd = [sys.executable, str(script_path), file_path, str(num_workers), steps_json]
        cmd_display = f'python gmail_bot_linked.py ...'
    return cmd, cmd_display, None


def _spawn_and_stream(cmd, cmd_display, label=''):
    """Spawn a subprocess and stream its output to the UI log. Returns exit code."""
    import subprocess

    add_log(f'[DEBUG] Command: {cmd_display}', 'info')
    add_log(f'[DEBUG] About to spawn process...', 'info')

    try:
        unbuffered_env = {**os.environ, 'PYTHONUNBUFFERED': '1'}
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            encoding='utf-8',
            errors='replace',
            env=unbuffered_env,
        )
        add_log(f'[DEBUG] Process spawned with PID: {process.pid}', 'info')
    except Exception as spawn_error:
        add_log(f'[ERROR] Failed to spawn process: {spawn_error}', 'error')
        return -1

    line_count = 0
    for line in iter(process.stdout.readline, ''):
        if stop_flag.is_set():
            process.terminate()
            add_log(f'Processing stopped by user{" (" + label + ")" if label else ""}', 'warning')
            break

        if line:
            line_count += 1
            line = line.strip()

            if '[OUTPUT_FILE]' in line:
                output_path = line.split('[OUTPUT_FILE]')[1].strip()
                # Resolve to absolute path (subprocess may print relative)
                output_path = str(Path(output_path).resolve())
                processing_state['output_file_path'] = output_path
                # Persist to disk so the path survives server restarts
                _persist_output_dirs(
                    output_file_path=output_path,
                    input_file_path=processing_state.get('file_path', ''),
                )
                add_log(f'Output file: {output_path}', 'info')
                continue

            # ── In-memory progress tracking ────────────────────────
            # Parse [TOTAL] marker (printed by ExcelProcessor at startup)
            if '[TOTAL]' in line:
                try:
                    t = int(line.split('[TOTAL]')[1].strip())
                    with _progress_lock:
                        _progress_counters['total'] = t
                except Exception:
                    pass

            # Parse [EXCEL] Row N -> SUCCESS/FAILED markers from base_runner
            if '[EXCEL] Row' in line:
                if '-> SUCCESS' in line:
                    with _progress_lock:
                        _progress_counters['success'] += 1
                elif '-> FAILED' in line:
                    with _progress_lock:
                        _progress_counters['failed'] += 1

            # Also count from generic stdout markers
            if 'ACCOUNT DONE:' in line and '= SUCCESS' in line:
                # Deduplicate: base_runner prints both [EXCEL] and ACCOUNT DONE
                # Only count if [EXCEL] wasn't already in this line
                pass  # Already counted via [EXCEL] marker above

            if 'ERROR' in line or 'FAILED' in line:
                add_log(line, 'error')
            elif 'SUCCESS' in line:
                add_log(line, 'success')
            elif 'WARNING' in line:
                add_log(line, 'warning')
            else:
                add_log(line, 'info')

            if '[WORKER' in line and 'Processing:' in line:
                parts = line.split('Processing:')
                if len(parts) > 1:
                    email = parts[1].strip().split()[0]
                    processing_state['current_account'] = email

    add_log(f'[DEBUG] Finished reading output. Total lines: {line_count}', 'info')
    return_code = process.wait()
    add_log(f'[DEBUG] Process exited with code: {return_code}', 'info')
    return return_code


def run_processing_worker(file_path, num_workers, bot_step=2, bot_steps=None,
                          linked=False, ops_per_step=None):
    """Background worker to run production script(s).

    Modes:
      1. Single step        — spawn gmail_bot_stepN.py
      2. Multi-step linked  — spawn gmail_bot_linked.py (one session per account)
      3. Multi-step unlinked — spawn each step's script sequentially
    """
    import json

    if bot_steps is None:
        bot_steps = [bot_step]
    if ops_per_step is None:
        ops_per_step = {}

    is_multi = len(bot_steps) > 1

    try:
        if not Path(file_path).exists():
            add_log(f'ERROR: Excel file not found at {file_path}', 'error')
            processing_state['status'] = 'idle'
            return

        add_log(f'Starting {num_workers} workers...', 'info')
        add_log(f'[DEBUG] Excel file: {file_path}', 'info')

        if is_multi and linked:
            # ── LINKED MODE: single process, all steps in one session ──────
            steps_json = json.dumps({
                'steps': bot_steps,
                'ops_per_step': ops_per_step,
            })
            cmd, cmd_display, err = _build_linked_cmd(file_path, num_workers, steps_json)
            if err:
                add_log(f'ERROR: {err}', 'error')
                processing_state['status'] = 'idle'
                return

            return_code = _spawn_and_stream(cmd, cmd_display, label='linked')

            if return_code == 0:
                add_log('Processing completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                add_log(f'Processing failed with exit code: {return_code}', 'error')
                processing_state['status'] = 'idle'

        elif is_multi and not linked:
            # ── UNLINKED MODE: run each step sequentially ──────────────────
            all_ok = True
            for step_num in bot_steps:
                if stop_flag.is_set():
                    add_log('Processing stopped by user before next step', 'warning')
                    all_ok = False
                    break

                add_log(f'--- Starting Step {step_num} ---', 'info')
                cmd, cmd_display, err = _build_step_cmd(step_num, file_path, num_workers)
                if err:
                    add_log(f'ERROR: {err}', 'error')
                    all_ok = False
                    continue

                return_code = _spawn_and_stream(cmd, cmd_display, label=f'Step {step_num}')
                if return_code != 0:
                    add_log(f'Step {step_num} failed with exit code: {return_code}', 'error')
                    all_ok = False
                else:
                    add_log(f'Step {step_num} completed successfully', 'success')

                # Reset statuses for next step so accounts are re-processed
                # (Each step script picks up PENDING rows)

            if all_ok:
                add_log('All steps completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                processing_state['status'] = 'idle'

        else:
            # ── SINGLE STEP MODE (original behavior) ──────────────────────
            cmd, cmd_display, err = _build_step_cmd(bot_steps[0], file_path, num_workers)
            if err:
                add_log(f'ERROR: {err}', 'error')
                processing_state['status'] = 'idle'
                return

            return_code = _spawn_and_stream(cmd, cmd_display)

            if return_code == 0:
                add_log('Processing completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                add_log(f'Processing failed with exit code: {return_code}', 'error')
                processing_state['status'] = 'idle'

        # ── Auto-generate MailNexus Pro report ────────────────────────────
        _auto_generate_report()

    except Exception as e:
        import traceback
        traceback.print_exc()
        processing_state['status'] = 'idle'
        add_log(f'Processing error: {str(e)}', 'error')
        _auto_generate_report()


@app.route('/api/log-stream')
def log_stream():
    """
    Server-Sent Events endpoint for real-time log streaming.
    The browser connects once; the server pushes each new log entry
    within ~100 ms of it being generated — no polling delay.
    """
    # Only replay logs that arrived AFTER the last clear operation so that
    # a page reload never brings back entries the user already cleared.
    clear_id = processing_state.get('log_clear_id', 0)
    snapshot = [lg for lg in processing_state['recent_logs'] if lg['id'] > clear_id]
    # Use clear_id as the floor so the live-stream loop also ignores old ids
    start_id = snapshot[-1]['id'] if snapshot else clear_id

    def generate():
        # ── Replay logs the client missed before connecting ───────────────
        for log in snapshot:
            yield f"data: {json.dumps(log)}\n\n"

        last_id = start_id
        idle_cycles = 0  # track consecutive empty polls
        max_cycles = 18000  # ~30 min at 100ms intervals → auto-close stale connections

        # ── Stream live logs as they arrive ──────────────────────────────
        while idle_cycles < max_cycles:
            try:
                new_logs = [
                    lg for lg in processing_state['recent_logs']
                    if lg['id'] > last_id
                ]
                for lg in new_logs:
                    yield f"data: {json.dumps(lg)}\n\n"
                    last_id = lg['id']
                    idle_cycles = 0

                if not new_logs:
                    idle_cycles += 1
                    # Send keepalive every ~3 seconds (not every 100ms)
                    if idle_cycles % 30 == 0:
                        yield ": keepalive\n\n"

                time.sleep(0.1)
            except GeneratorExit:
                # Client disconnected — clean exit
                return

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        }
    )


@app.route('/api/config', methods=['GET'])
def get_config():
    """Read config/settings.json and config/urls.json"""
    try:
        cfg_path = RESOURCES_PATH / 'config' / 'settings.json'
        url_path = RESOURCES_PATH / 'config' / 'urls.json'

        settings = {}
        urls = {}

        if cfg_path.exists():
            try:
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except (json.JSONDecodeError, ValueError):
                settings = {}

        if url_path.exists():
            try:
                with open(url_path, 'r', encoding='utf-8') as f:
                    urls = json.load(f)
            except (json.JSONDecodeError, ValueError):
                urls = {}

        return jsonify({'success': True, 'settings': settings, 'urls': urls})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/config', methods=['POST'])
def save_config():
    """Write config/settings.json and config/urls.json"""
    try:
        data = request.json or {}
        cfg_path = RESOURCES_PATH / 'config' / 'settings.json'
        url_path = RESOURCES_PATH / 'config' / 'urls.json'

        if 'settings' in data:
            with open(cfg_path, 'w', encoding='utf-8') as f:
                json.dump(data['settings'], f, indent=2)

        if 'urls' in data:
            with open(url_path, 'w', encoding='utf-8') as f:
                json.dump(data['urls'], f, indent=2)

        return jsonify({'success': True, 'message': 'Configuration saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/all', methods=['DELETE'])
def delete_all_reports():
    """Delete all .xlsx files inside the app output folders ONLY.
    Never touches user's Downloads or any other directory."""
    try:
        scan_dirs = _get_report_scan_dirs()
        if not scan_dirs:
            return jsonify({'success': True, 'deleted': 0})

        count = 0
        errors = []
        for dir_path in scan_dirs:
            d = Path(dir_path)
            if not d.exists():
                continue
            for f in d.rglob('*.xlsx'):
                # Extra safety: ensure file is actually inside the output dir
                try:
                    f.resolve().relative_to(d.resolve())
                except ValueError:
                    continue
                try:
                    f.unlink()
                    count += 1
                except Exception as e:
                    errors.append(f'{f.name}: {e}')

        if errors:
            return jsonify({'success': False, 'deleted': count,
                            'message': f'Deleted {count}, but {len(errors)} error(s): {"; ".join(errors[:3])}'})
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} report(s)'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/single', methods=['DELETE'])
def delete_single_report():
    """Delete a single report file by its absolute path.
    Only allows deletion from the project /output folder."""
    try:
        data = request.json or {}
        file_path = data.get('path', '')

        if not file_path:
            return jsonify({'success': False, 'message': 'No file path provided.'})

        fp = Path(file_path)
        if not fp.exists():
            return jsonify({'success': False, 'message': 'File not found.'})

        if not fp.suffix.lower() == '.xlsx':
            return jsonify({'success': False, 'message': 'Only .xlsx files can be deleted.'})

        # Security: only allow deletion from known output folders
        scan_dirs = _get_report_scan_dirs()
        resolved = fp.resolve()
        in_known_dir = any(
            _is_safe_child(resolved, Path(sd))
            for sd in scan_dirs
        )
        if not in_known_dir:
            return jsonify({'success': False, 'message': 'File is not in a known output directory.'})

        fp.unlink()
        return jsonify({'success': True, 'message': f'Deleted: {fp.name}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/clear', methods=['POST'])
def clear_logs():
    """Clear all in-memory logs and set a clear threshold so SSE never
    replays old entries on page reload."""
    logs = processing_state['recent_logs']
    if logs:
        processing_state['log_clear_id'] = logs[-1]['id']
    processing_state['recent_logs'] = []
    return jsonify({'success': True, 'message': 'Logs cleared',
                    'clear_id': processing_state['log_clear_id']})


@app.route('/api/sms-code', methods=['POST'])
def receive_sms_code():
    """Receive an SMS verification code from the phone relay app."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        code = str(data.get('code', '')).strip()
        if not code:
            return jsonify({'success': False, 'error': 'No code provided'}), 400

        sms_codes.append({
            'code': code,
            'sender': data.get('sender', ''),
            'full_message': data.get('full_message', ''),
            'timestamp': time.time(),
            'used': False
        })
        # Keep only last 20 codes
        if len(sms_codes) > 20:
            sms_codes[:] = sms_codes[-20:]

        add_log(f'[SMS] Code received: {code} from {data.get("sender", "unknown")}', 'info')
        return jsonify({'success': True, 'message': f'Code {code} stored'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sms-code', methods=['GET'])
def get_sms_code():
    """Get the latest unused SMS verification code.
    ?peek=1 returns without marking as used (for UI display).
    ?max_age=120 limits how old a code can be (seconds)."""
    try:
        max_age = float(request.args.get('max_age', 120))
        peek = request.args.get('peek', '0') == '1'
        now = time.time()
        for entry in reversed(sms_codes):
            if not entry['used'] and (now - entry['timestamp']) < max_age:
                if not peek:
                    entry['used'] = True
                return jsonify({
                    'success': True,
                    'code': entry['code'],
                    'sender': entry.get('sender', ''),
                    'timestamp': entry['timestamp']
                })
        return jsonify({'success': False, 'code': None})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fingerprint', methods=['GET'])
def get_fingerprint():
    """Read config/fingerprint.json and return fingerprint settings."""
    try:
        fp_path = RESOURCES_PATH / 'config' / 'fingerprint.json'
        if not fp_path.exists():
            return jsonify({'success': True, 'fingerprint': {'os_type': 'random', 'auto_timezone': True}})
        with open(fp_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify({'success': True, 'fingerprint': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/fingerprint', methods=['POST'])
def save_fingerprint():
    """Write fingerprint settings to config/fingerprint.json."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        _valid_os = ('random', 'windows', 'macos', 'linux', 'android', 'ios')
        os_type      = str(data.get('os_type', 'random')).lower()
        os_type      = os_type if os_type in _valid_os else 'random'
        auto_timezone = bool(data.get('auto_timezone', True))

        fp_path = RESOURCES_PATH / 'config' / 'fingerprint.json'
        fp_path.parent.mkdir(parents=True, exist_ok=True)
        with open(fp_path, 'w', encoding='utf-8') as f:
            json.dump({'os_type': os_type, 'auto_timezone': auto_timezone}, f, indent=2)

        tz_label = 'Auto from IP (geo-lookup)' if auto_timezone else 'Random pool'
        return jsonify({
            'success': True,
            'message': f'Fingerprint saved — OS: {os_type} | Timezone: {tz_label}'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/launch', methods=['POST'])
def debug_launch():
    """Launch debug browser(s) with current proxy + fingerprint for manual inspection."""
    try:
        add_log('[DEBUG] Importing debug_launcher...', 'info')
        from shared import debug_launcher
        # Connect debug_launcher logs → UI log panel so user sees everything live
        debug_launcher.set_ui_logger(add_log)

        data         = request.json or {}
        num_browsers = max(1, min(int(data.get('num_browsers', 1)), 10))
        test_url     = str(data.get('test_url', 'https://ipinfo.io')).strip() or 'https://ipinfo.io'

        add_log(f'[DEBUG] Calling launch({num_browsers}, {test_url})...', 'info')
        debug_launcher.launch(num_browsers=num_browsers, test_url=test_url)
        add_log(f'[DEBUG] launch() returned — threads are starting...', 'info')

        return jsonify({
            'success': True,
            'message': f'Launched {num_browsers} debug browser(s). Check logs below for IP/proxy/fingerprint details.',
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        add_log(f'[DEBUG-LAUNCH ERROR] {e}', 'error')
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/close', methods=['POST'])
def debug_close():
    """Close all open debug browsers."""
    try:
        from shared import debug_launcher
        debug_launcher.close_all()
        return jsonify({'success': True, 'message': 'Close signal sent — debug browsers shutting down.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/status', methods=['GET'])
def debug_status():
    """Return current debug browser status."""
    try:
        from shared import debug_launcher
        s = debug_launcher.status()
        return jsonify({'success': True, **s})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'open': 0, 'total': 0, 'running': False})


def add_log(message, log_type='info'):
    """Add a log entry — visible immediately via SSE stream."""
    log_entry = {
        'id': next(_log_id),          # collision-free incrementing ID
        'message': message,
        'type': log_type,
        'timestamp': time.time()
    }

    processing_state['recent_logs'].append(log_entry)

    # Keep only last 500 logs in memory
    if len(processing_state['recent_logs']) > 500:
        processing_state['recent_logs'] = processing_state['recent_logs'][-500:]

    try:
        print(f"[{log_type.upper()}] {message}")
    except UnicodeEncodeError:
        safe_msg = message.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8')
        print(f"[{log_type.upper()}] {safe_msg}")


# ── Auth Endpoints (licensing enforced) ──────────────────────────────────────
from shared import license_manager
license_manager.init(RESOURCES_PATH)


@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    info = license_manager.get_license_info()
    return jsonify({
        'success': True,
        'auth_enabled': True,
        'license_activated': info.get('valid', False),
        'reason': info.get('reason'),
        'tier': info.get('tier'),
        'days_remaining': info.get('days_remaining'),
    })


@app.route('/api/license/info', methods=['GET'])
def license_info():
    return jsonify(license_manager.get_license_info())


@app.route('/api/license/activate', methods=['POST'])
def license_activate():
    body = request.get_json(silent=True) or {}
    key = (body.get('license_key') or '').strip()
    if not key:
        return jsonify({'success': False, 'message': 'license_key required'}), 400
    return jsonify(license_manager.activate(key))


@app.route('/api/license/deactivate', methods=['POST'])
def license_deactivate():
    return jsonify(license_manager.deactivate())


@app.route('/api/license/reseal', methods=['POST'])
def license_reseal():
    """Re-sign existing license.json after a secret rotation. Admin-only."""
    ok = license_manager.reseal_existing()
    return jsonify({'success': ok})


# Whitelist of paths that should always be reachable, even when license invalid.
_LICENSE_OPEN_PATHS = {
    '/api/health', '/api/app/version', '/api/auth/status',
    '/api/license/info', '/api/license/activate',
    '/api/license/deactivate', '/api/license/reseal',
    '/api/shutdown', '/api/log-stream',
}


@app.before_request
def _gate_license():
    """Block API endpoints when license is invalid (except whitelisted ones)."""
    p = request.path or ''
    if not p.startswith('/api/'):
        return None
    if p in _LICENSE_OPEN_PATHS:
        return None
    if license_manager.is_licensed():
        return None
    return jsonify({
        'success': False,
        'license_required': True,
        'message': 'License required. Please activate first.',
    }), 403


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TOOLS API — routes/tools.py Blueprint (registered below)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Routes are in electron-app/backend/routes/tools.py.
# Registered at bottom of file alongside other Blueprints.


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROFILE MANAGER API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from shared import nexus_profile_manager as profile_manager, recovery_tracker
profile_manager.init(RESOURCES_PATH)
from shared import review_stats_scraper as review_stats
review_stats.init(RESOURCES_PATH)
profile_manager.set_ui_logger(add_log)
recovery_tracker.init(RESOURCES_PATH)

# Safety net: close all managed browsers when backend exits
# This ensures Chrome processes we launched don't become orphans.
# Does NOT touch the user's own Chrome browser.
def _cleanup_browsers_on_exit():
    try:
        profile_manager.close_all_profiles()
    except Exception:
        pass
atexit.register(_cleanup_browsers_on_exit)



@app.route('/api/profiles/vault-status', methods=['GET'])
def profiles_vault_status():
    """Return credential vault encryption status.

    Response:
        {
          "vault_available": true,      # credential_vault module loaded OK
          "encryption_active": true,    # DPAPI (or XOR fallback) is encrypting
          "backend": "dpapi"|"xor"|"none",
          "encrypted_count": 12,        # profiles with ≥1 encrypted field
          "total_count": 15
        }
    """
    try:
        from shared.credential_vault import is_encrypted, _is_windows
        import sys
        backend = 'dpapi' if _is_windows() else 'xor'
        profiles = profile_manager.list_profiles()
        # list_profiles() returns decrypted profiles — read raw file to count
        from shared.profile_manager import _profiles_file
        import json as _json
        raw_profiles: list[dict] = []
        pf = _profiles_file()
        if pf.exists():
            try:
                raw_profiles = _json.loads(pf.read_text(encoding='utf-8'))
            except Exception:
                pass
        enc_count = sum(1 for p in raw_profiles if is_encrypted(p))
        return jsonify({
            'success': True,
            'vault_available': True,
            'encryption_active': True,
            'backend': backend,
            'encrypted_count': enc_count,
            'total_count': len(raw_profiles),
        })
    except ImportError:
        return jsonify({
            'success': True,
            'vault_available': False,
            'encryption_active': False,
            'backend': 'none',
            'encrypted_count': 0,
            'total_count': 0,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/profiles/vault-migrate', methods=['POST'])
def profiles_vault_migrate():
    """Force-encrypt all profiles that still have plain-text credentials.

    Safe to call multiple times — already-encrypted fields are skipped.
    Returns counts of migrated vs already-encrypted profiles.
    """
    try:
        from shared.credential_vault import is_encrypted, encrypt_profile_fields
        from shared.profile_manager import _read_profiles, _write_profiles, _file_lock
        with _file_lock:
            profiles = _read_profiles()   # returns decrypted (plain) values
            _write_profiles(profiles)     # re-writes with encryption applied
        # Count result
        from shared.profile_manager import _profiles_file
        import json as _json
        from shared.credential_vault import is_encrypted as _ie
        raw = _json.loads(_profiles_file().read_text(encoding='utf-8'))
        enc_count = sum(1 for p in raw if _ie(p))
        return jsonify({
            'success': True,
            'message': f'Migration complete: {enc_count}/{len(raw)} profiles encrypted',
            'encrypted_count': enc_count,
            'total_count': len(raw),
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/profiles/groups', methods=['GET'])
def profiles_list_groups():
    """Return all unique profile group names with counts."""
    profiles = profile_manager.list_profiles()
    from collections import Counter
    counts = Counter()
    for p in profiles:
        for g in profile_manager._get_groups(p):
            counts[g] += 1
    groups = sorted(counts.keys())
    return jsonify({'success': True, 'groups': groups, 'counts': dict(counts)})


@app.route('/api/profiles/bulk-assign-group', methods=['POST'])
def profiles_bulk_assign_group():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    group = (data.get('group') or 'default').strip()
    mode = data.get('mode', 'add')  # 'add' or 'set'
    note = (data.get('note') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    updated = profile_manager.bulk_assign_group(ids, group, mode=mode)
    notes_updated = 0
    if note:
        for pid in ids:
            try:
                profile_manager.update_profile(pid, notes=note)
                notes_updated += 1
            except Exception:
                pass
    return jsonify({'success': True, 'updated': updated, 'notes_updated': notes_updated})


@app.route('/api/profiles/bulk-remove-group', methods=['POST'])
def profiles_bulk_remove_group():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    group = (data.get('group') or '').strip()
    note = (data.get('note') or '').strip()
    if not ids or not group:
        return jsonify({'success': False, 'message': 'ids and group required'})
    updated = profile_manager.remove_profile_from_group(ids, group)
    notes_updated = 0
    if note:
        for pid in ids:
            try:
                profile_manager.update_profile(pid, notes=note)
                notes_updated += 1
            except Exception:
                pass
    return jsonify({'success': True, 'updated': updated, 'notes_updated': notes_updated})


@app.route('/api/profiles/bulk-update-notes', methods=['POST'])
def profiles_bulk_update_notes():
    """Update notes on multiple profiles at once without changing groups."""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    note = (data.get('note') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    if not note:
        return jsonify({'success': False, 'message': 'Note is empty'})
    updated = 0
    for pid in ids:
        try:
            profile_manager.update_profile(pid, notes=note)
            updated += 1
        except Exception:
            pass
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/bulk-update-proxy', methods=['POST'])
def profiles_bulk_update_proxy():
    """Update proxy username and/or password for multiple profiles at once."""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    proxy_user = (data.get('proxy_user') or '').strip()
    proxy_pass = (data.get('proxy_pass') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    if not proxy_user and not proxy_pass:
        return jsonify({'success': False, 'message': 'Provide at least proxy user or password'})
    updated = 0
    for pid in ids:
        try:
            p = profile_manager.get_profile(pid)
            if not p:
                continue
            proxy = dict(p.get('proxy') or {})
            if proxy_user:
                proxy['username'] = proxy_user
            if proxy_pass:
                proxy['password'] = proxy_pass
            profile_manager.update_profile(pid, proxy=proxy)
            updated += 1
        except Exception:
            pass
    return jsonify({'success': True, 'updated': updated})


# ─────────────────────────────────────────────────────────────────────────────
# Bulk proxy: assign a list of proxies round-robin to many profiles, OR
# toggle the proxy on/off across a scope (all profiles / single group /
# explicit ids list). Cache & cookies wiping. Single-profile export.
# ─────────────────────────────────────────────────────────────────────────────

# Shared background-job status for bulk-set-proxies and bulk-toggle-proxy.
# Frontend polls /api/profiles/bulk-proxy/status while a job is running.
import threading as _bp_threading
_bulk_proxy_lock = _bp_threading.Lock()
_bulk_proxy_status: dict = {
    'running': False, 'op': '',         # op = 'set' | 'toggle'
    'total': 0, 'done': 0, 'ok': 0, 'failed': 0,
    'current_email': '', 'started_at': '', 'finished_at': '',
    'extra': {},                         # op-specific summary (e.g. proxies)
}


def _bp_reset(op: str, total: int, extra: dict | None = None):
    from datetime import datetime as _dt
    with _bulk_proxy_lock:
        _bulk_proxy_status.update({
            'running': True, 'op': op,
            'total': total, 'done': 0, 'ok': 0, 'failed': 0,
            'current_email': '',
            'started_at': _dt.utcnow().isoformat() + 'Z',
            'finished_at': '',
            'extra': extra or {},
        })


def _bp_tick(email: str = '', ok: bool = True):
    with _bulk_proxy_lock:
        _bulk_proxy_status['done'] += 1
        if ok:
            _bulk_proxy_status['ok'] += 1
        else:
            _bulk_proxy_status['failed'] += 1
        if email:
            _bulk_proxy_status['current_email'] = email


def _bp_finish():
    from datetime import datetime as _dt
    with _bulk_proxy_lock:
        _bulk_proxy_status['running'] = False
        _bulk_proxy_status['finished_at'] = _dt.utcnow().isoformat() + 'Z'


def _resolve_proxy_scope(data: dict) -> list[str]:
    """Resolve a scope dict {ids?, all?, group?} → list of profile IDs.
    `ids` (explicit list) takes precedence, then `all` (every profile),
    then `group` (every profile whose groups[] contains the name).
    """
    ids = data.get('ids') or []
    if ids:
        return list(ids)
    everyone = profile_manager.list_profiles()
    if data.get('all'):
        return [p['id'] for p in everyone]
    group = (data.get('group') or '').strip()
    if group:
        return [p['id'] for p in everyone
                if group in profile_manager._get_groups(p)]
    return []


def _bp_worker_set(target_ids: list, parsed: list):
    """Background worker for bulk-set-proxies. Updates _bulk_proxy_status
    as it goes so the UI progress bar can poll."""
    try:
        for i, pid in enumerate(target_ids):
            proxy = dict(parsed[i % len(parsed)])
            try:
                p = profile_manager.get_profile(pid)
                em = (p or {}).get('email', '') if p else ''
                res = profile_manager.update_profile(pid, proxy=proxy)
                _bp_tick(em, ok=bool(res))
            except Exception:
                _bp_tick('', ok=False)
    finally:
        _bp_finish()


def _bp_worker_toggle(target_ids: list, enabled: bool):
    """Background worker for bulk-toggle-proxy."""
    try:
        for pid in target_ids:
            try:
                p = profile_manager.get_profile(pid)
                if not p:
                    _bp_tick('', ok=False)
                    continue
                em = p.get('email', '')
                if enabled:
                    saved = p.get('proxy_saved') or {}
                    if saved:
                        profile_manager.update_profile(pid, proxy=saved)
                        profile_manager._update_profile_field(pid, 'proxy_saved', None)
                        _bp_tick(em, ok=True)
                    else:
                        # No archive to restore — count as skipped (not failed)
                        _bp_tick(em, ok=False)
                else:
                    cur = p.get('proxy') or {}
                    if cur:
                        profile_manager._update_profile_field(pid, 'proxy_saved', cur)
                        profile_manager.update_profile(pid, proxy={})
                        _bp_tick(em, ok=True)
                    else:
                        # Nothing to turn off
                        _bp_tick(em, ok=False)
            except Exception:
                _bp_tick('', ok=False)
    finally:
        _bp_finish()


@app.route('/api/profiles/bulk-set-proxies', methods=['POST'])
def profiles_bulk_set_proxies():
    """Distribute a list of proxy strings round-robin across profiles.
    Runs in a background thread — caller polls /bulk-proxy/status."""
    from shared.profile_manager import _parse_proxy_string
    with _bulk_proxy_lock:
        if _bulk_proxy_status['running']:
            return jsonify({'success': False,
                            'message': 'Another bulk-proxy job is running'}), 409
    data = request.get_json(force=True, silent=True) or {}
    lines = data.get('proxy_lines') or []
    if isinstance(lines, str):
        lines = lines.splitlines()
    parsed = []
    for raw in lines:
        pr = _parse_proxy_string((raw or '').strip())
        if pr:
            parsed.append(pr)
    if not parsed:
        return jsonify({'success': False,
                        'message': 'No valid proxy lines parsed'}), 400
    target_ids = _resolve_proxy_scope(data)
    if not target_ids:
        return jsonify({'success': False, 'message': 'No matching profiles'}), 400
    _bp_reset('set', len(target_ids), {'proxies': len(parsed)})
    t = _bp_threading.Thread(
        target=_bp_worker_set,
        args=(target_ids, parsed),
        daemon=True, name='bulk-set-proxies',
    )
    t.start()
    return jsonify({'success': True, 'started': True,
                    'targets': len(target_ids), 'proxies': len(parsed)})


@app.route('/api/profiles/bulk-toggle-proxy', methods=['POST'])
def profiles_bulk_toggle_proxy():
    """Turn proxy on or off for many profiles in one call.

    `enabled=False`  → archive current proxy under `proxy_saved` and clear
                       `proxy` so the next launch uses no proxy.
    `enabled=True`   → restore `proxy_saved` back to `proxy` if archived.
    Runs in a background thread — caller polls /bulk-proxy/status."""
    with _bulk_proxy_lock:
        if _bulk_proxy_status['running']:
            return jsonify({'success': False,
                            'message': 'Another bulk-proxy job is running'}), 409
    data = request.get_json(force=True, silent=True) or {}
    enabled = bool(data.get('enabled'))
    target_ids = _resolve_proxy_scope(data)
    if not target_ids:
        return jsonify({'success': False, 'message': 'No matching profiles'}), 400
    _bp_reset('toggle', len(target_ids), {'enabled': enabled})
    t = _bp_threading.Thread(
        target=_bp_worker_toggle,
        args=(target_ids, enabled),
        daemon=True, name='bulk-toggle-proxy',
    )
    t.start()
    return jsonify({'success': True, 'started': True,
                    'targets': len(target_ids), 'enabled': enabled})


@app.route('/api/profiles/bulk-proxy/status', methods=['GET'])
def profiles_bulk_proxy_status():
    """Snapshot of the current bulk-proxy job for UI polling."""
    with _bulk_proxy_lock:
        return jsonify(dict(_bulk_proxy_status))


@app.route('/api/profiles/<profile_id>/clear-cache', methods=['POST'])
def profiles_clear_cache(profile_id):
    """Wipe Chrome cache folders from this profile's user-data-dir."""
    import shutil
    p = profile_manager.get_profile(profile_id)
    if not p:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    if p.get('browser_open') == 'running':
        return jsonify({'success': False,
                        'message': 'Close the browser first'}), 409
    pdir = (p.get('profile_dir') or '').strip()
    if not pdir or not os.path.isdir(pdir):
        return jsonify({'success': False,
                        'message': 'Profile directory not found'}), 404
    targets = [
        os.path.join(pdir, 'Default', 'Cache'),
        os.path.join(pdir, 'Default', 'Code Cache'),
        os.path.join(pdir, 'Default', 'GPUCache'),
        os.path.join(pdir, 'Default', 'Service Worker', 'CacheStorage'),
        os.path.join(pdir, 'Default', 'Service Worker', 'ScriptCache'),
        os.path.join(pdir, 'ShaderCache'),
        os.path.join(pdir, 'GrShaderCache'),
    ]
    cleared = 0
    bytes_freed = 0
    for t in targets:
        if not os.path.isdir(t):
            continue
        try:
            # measure first so we can report
            for dp, dn, fn in os.walk(t):
                for f in fn:
                    try:
                        bytes_freed += os.path.getsize(os.path.join(dp, f))
                    except Exception:
                        pass
            shutil.rmtree(t, ignore_errors=True)
            cleared += 1
        except Exception:
            pass
    return jsonify({'success': True, 'cleared': cleared,
                    'freed_mb': round(bytes_freed / 1024 / 1024, 1)})


@app.route('/api/profiles/<profile_id>/clear-cookies', methods=['POST'])
def profiles_clear_cookies(profile_id):
    """Wipe Chrome cookie store from this profile's user-data-dir."""
    p = profile_manager.get_profile(profile_id)
    if not p:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    if p.get('browser_open') == 'running':
        return jsonify({'success': False,
                        'message': 'Close the browser first'}), 409
    pdir = (p.get('profile_dir') or '').strip()
    if not pdir or not os.path.isdir(pdir):
        return jsonify({'success': False,
                        'message': 'Profile directory not found'}), 404
    # Chrome stores cookies in a few possible locations depending on version.
    # Wipe every candidate; profile re-logs in next launch.
    targets = [
        os.path.join(pdir, 'Default', 'Network', 'Cookies'),
        os.path.join(pdir, 'Default', 'Network', 'Cookies-journal'),
        os.path.join(pdir, 'Default', 'Cookies'),
        os.path.join(pdir, 'Default', 'Cookies-journal'),
    ]
    cleared = 0
    for t in targets:
        if os.path.isfile(t):
            try:
                os.remove(t)
                cleared += 1
            except Exception:
                pass
    # Mark logged-out so the UI reflects the cleared state immediately.
    try:
        profile_manager.update_profile(profile_id, status='not_logged_in')
    except Exception:
        pass
    return jsonify({'success': True, 'cleared': cleared})


@app.route('/api/profiles/groups/rename', methods=['POST'])
def profiles_rename_group():
    data = request.get_json() or {}
    old_name = (data.get('old_name') or '').strip()
    new_name = (data.get('new_name') or '').strip()
    if not old_name or not new_name:
        return jsonify({'success': False, 'message': 'old_name and new_name are required'})
    updated = profile_manager.rename_group(old_name, new_name)
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/groups/<path:group_name>', methods=['DELETE'])
def profiles_delete_group(group_name):
    data = request.get_json() or {}
    reassign_to = (data.get('reassign_to') or 'default').strip()
    updated = profile_manager.delete_group(group_name, reassign_to)
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/export-excel', methods=['POST'])
def profiles_export_excel():
    """Export selected profiles to Excel — full credentials + Write Review columns."""
    import io
    import json as _json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True, silent=True) or {}
    ids = data.get('ids', [])
    export_all = bool(data.get('all'))
    if not ids and not export_all:
        return jsonify({'success': False, 'message': 'No profile IDs provided'})

    # Use _read_profiles() directly so password/totp/backup_codes are included
    all_profiles = profile_manager._read_profiles()
    if export_all:
        # Mirrors the "Bookmark All" pattern: pagination prevents
        # bulk-select-all in the UI, so this flag lets the user dump
        # every profile in one shot regardless of current filter/page.
        selected = list(all_profiles)
    else:
        id_set = set(ids)
        selected = [p for p in all_profiles if p.get('id') in id_set]
    if not selected:
        return jsonify({'success': False, 'message': 'No matching profiles found'})

    wb = openpyxl.Workbook()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 1 — Credentials
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = 'Credentials'

    thin       = Side(style='thin', color='CBD5E1')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', fgColor='1E3A5F')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='left', vertical='center')

    # Divider style — separates credential block from write-review block
    div_fill   = PatternFill('solid', fgColor='1E3A5F')
    div_font   = Font(name='Calibri', bold=True, color='FCD34D', size=10)

    status_colors = {
        'logged_in':    'D1FAE5',
        'login_failed': 'FEE2E2',
        'pending':      'FEF9C3',
    }

    # Column layout: (header, width)
    cred_cols = [
        ('Name',          20),
        ('Email',         30),
        ('Password',      22),
        ('TOTP Secret',   28),
        ('Backup Codes',  45),
        ('Login Status',  15),
        ('Groups',        18),
        ('Engine',        10),
        ('Proxy',         38),
        ('Address',       35),
        ('Notes',         30),
        ('Created At',    18),
    ]
    wr_cols = [
        ('GMB URL',      40),
        ('Review Text',  50),
        ('Review Stars', 14),
    ]
    all_cols = cred_cols + [('', 3)] + wr_cols  # empty column as visual divider

    for col_idx, (h, w) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
        if h == '':
            # Divider column — dark fill, no text
            cell.fill = PatternFill('solid', fgColor='334155')
            continue
        if col_idx <= len(cred_cols):
            cell.font = hdr_font
            cell.fill = hdr_fill
        else:
            # Write Review columns — amber header
            cell.font = Font(name='Calibri', bold=True, color='1C1917', size=11)
            cell.fill = PatternFill('solid', fgColor='FCD34D')
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, p in enumerate(selected, 2):
        status = p.get('status', '')
        groups = ', '.join(profile_manager._get_groups(p))

        # Proxy → readable string  host:port (user:pass)
        proxy = p.get('proxy') or {}
        if proxy and proxy.get('host'):
            host = proxy.get('host', '')
            port = proxy.get('port', '')
            pu   = proxy.get('username', '')
            pp   = proxy.get('password', '')
            ptype = proxy.get('type', 'http')
            # Standard URL format: scheme://user:pass@host:port
            if pu or pp:
                proxy_str = f"{ptype}://{pu}:{pp}@{host}:{port}"
            else:
                proxy_str = f"{ptype}://{host}:{port}"
        else:
            proxy_str = ''

        # Backup codes → one per line
        backup_codes = p.get('backup_codes') or []
        if isinstance(backup_codes, list):
            codes_str = '\n'.join(str(c) for c in backup_codes if c)
        else:
            codes_str = str(backup_codes)

        row_data = [
            p.get('name', ''),
            p.get('email', ''),
            p.get('password', ''),
            p.get('totp_secret', ''),
            codes_str,
            status,
            groups,
            (p.get('engine', 'nexus') or 'nexus').upper(),
            proxy_str,
            p.get('address', ''),
            p.get('notes', ''),
            p.get('created_at', ''),
            # divider column
            '',
            # Write Review (blank — user fills)
            '',
            '',
            '',
        ]

        row_fill_color = status_colors.get(status, 'FFFFFF')
        row_fill = PatternFill('solid', fgColor=row_fill_color)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # Divider column
            if col_idx == len(cred_cols) + 1:
                cell.fill = PatternFill('solid', fgColor='334155')
                continue
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if col_idx <= len(cred_cols):
                cell.fill = row_fill
            # Backup codes column — slightly larger row height handled below

        # Row height — taller if there are backup codes
        n_codes = len(backup_codes) if isinstance(backup_codes, list) else 1
        ws.row_dimensions[row_idx].height = max(18, min(14 * n_codes, 90))

    ws.freeze_panes = 'B2'  # freeze Name column + header

    # ── Section label above Write Review columns ──────────────────────────
    wr_label_col = len(cred_cols) + 2  # first WR column
    label_cell = ws.cell(row=1, column=wr_label_col)
    # Already set above, add a small note below header in row 0 — skip, header is enough

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2 — Write Review (email + WR cols only, ready to use directly)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wr = wb.create_sheet('Write Review')
    wr_headers = [('Email', 30), ('GMB URL', 40), ('Review Text', 50), ('Review Stars', 14)]
    for col_idx, (h, w) in enumerate(wr_headers, 1):
        cell = wr.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        wr.column_dimensions[get_column_letter(col_idx)].width = w
    wr.row_dimensions[1].height = 24

    for row_idx, p in enumerate(selected, 2):
        for col_idx, value in enumerate([p.get('email', ''), '', '', ''], 1):
            cell = wr.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = border
        wr.row_dimensions[row_idx].height = 18

    wr.freeze_panes = 'A2'

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as _dt
    date_str = _dt.now().strftime('%Y%m%d_%H%M')
    prefix = 'profiles_export_all' if export_all else 'profiles_export'
    filename = f'{prefix}_{len(selected)}accs_{date_str}.xlsx'

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── Profile list search index ────────────────────────────────────────────────
# Pre-lowercased lightweight "row" per profile, keyed on profiles.json mtime.
# Used by /api/profiles to filter in pure Python lists of strings rather than
# rebuilding lowercase strings for 1800+ full dicts on every keystroke.
_profile_index_cache = {'mtime': None, 'rows': None, 'sorted_ids': None}
_profile_index_lock = threading.Lock()


def _get_profile_index():
    """Return (rows_by_id, sorted_ids_newest_first). Rebuilt only when
    profiles.json mtime changes — every search call after that is a dict lookup.

    NOTE: `browser_open` is NOT stored in the cached rows or the cached
    profile dict. That field is a LIVE state read from _active_browsers and
    changes outside the file (Close All / browser crash / launch) without
    bumping profiles.json mtime. Caller must annotate it per-request from
    the live tracking dict — see _annotate_live_browser_state() below.
    """
    try:
        from shared.nexus_profile_manager import _profiles_file
        pf = _profiles_file()
        mtime_key = pf.stat().st_mtime_ns if pf.exists() else 0
    except Exception:
        mtime_key = 0
    cur = _profile_index_cache
    if cur['rows'] is not None and cur['mtime'] == mtime_key:
        return cur['rows'], cur['sorted_ids']
    with _profile_index_lock:
        if cur['rows'] is not None and cur['mtime'] == mtime_key:
            return cur['rows'], cur['sorted_ids']
        # Read directly from disk (bypasses list_profiles' browser_open
        # annotation — we never want stale browser_open baked into the cache).
        from shared.nexus_profile_manager import _read_profiles
        profiles = _read_profiles()
        profiles.sort(key=lambda p: p.get('created_at') or '', reverse=True)
        rows = {}
        sorted_ids = []
        for p in profiles:
            pid = p.get('id') or ''
            if not pid:
                continue
            # Strip any accidental browser_open value from the cached dict;
            # the live annotation step will set it on every request.
            p.pop('browser_open', None)
            sorted_ids.append(pid)
            groups = p.get('groups') if isinstance(p.get('groups'), list) and p.get('groups') else [p.get('group') or 'default']
            proxy = p.get('proxy') or {}
            rows[pid] = {
                'profile': p,  # full dict reference for the page slice
                'name_lc':     (p.get('name') or '').lower(),
                'email_lc':    (p.get('email') or '').lower(),
                'notes_lc':    (p.get('notes') or p.get('note') or '').lower(),
                'groups_lc':   [str(g).lower() for g in groups],
                'proxy_host_lc':   (proxy.get('host') or '').lower(),
                'proxy_server_lc': (proxy.get('server') or '').lower(),
                'status':       p.get('status', ''),
                'engine':       p.get('engine', 'nexus'),
                # browser_open intentionally left out — annotated live per request
            }
        cur['rows'] = rows
        cur['sorted_ids'] = sorted_ids
        cur['mtime'] = mtime_key
        return rows, sorted_ids


def _annotate_live_browser_state(rows):
    """Read _active_browsers under its lock and stamp browser_open onto each row
    AND the row's profile dict. Called per-request so Close All / launches / crashes
    are reflected immediately, even though the index cache itself is mtime-keyed."""
    try:
        from shared.nexus_profile_manager import _active_browsers, _lock
        with _lock:
            snap = {pid: (info or {}).get('status', 'stopped') for pid, info in _active_browsers.items()}
    except Exception:
        snap = {}
    for pid, row in rows.items():
        bo = snap.get(pid, 'stopped')
        row['browser_open'] = bo
        # Stamp on the shared profile dict too — page_slice returns these
        # references straight to the JSON response.
        try:
            row['profile']['browser_open'] = bo
        except Exception:
            pass


@app.route('/api/profiles', methods=['GET'])
def profiles_list():
    """List profiles with pagination support.

    Searches/filters against an mtime-cached lowercase index instead of
    rebuilding lowercase strings on every request. Under 4 concurrent ops
    hammering profiles.json this drops search latency from ~500ms-1s to ~5ms.
    """
    search = request.args.get('search', '').lower()
    status_filter = request.args.get('filter', 'all').lower()
    group_filter = request.args.get('group', '').lower()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    per_page = max(10, min(per_page, 10000))
    search_by = request.args.get('search_by', 'name').lower()

    rows, sorted_ids = _get_profile_index()
    # Refresh the live browser_open field on every request — the cached index
    # only stores file-derived fields. Without this, Close All / launches
    # don't reflect until profiles.json itself changes.
    _annotate_live_browser_state(rows)

    # ── Build a predicate that runs on the lightweight row dict ──────────
    def keep(row):
        if search:
            if search_by == 'email':
                if search not in row['email_lc']: return False
            elif search_by == 'notes':
                if search not in row['notes_lc']: return False
            elif search_by == 'proxy':
                if search not in row['proxy_host_lc'] and search not in row['proxy_server_lc']:
                    return False
            elif search_by == 'group':
                if not any(search in g for g in row['groups_lc']): return False
            else:  # 'name' (default) matches name OR email
                if search not in row['name_lc'] and search not in row['email_lc']:
                    return False
        if group_filter and group_filter not in row['groups_lc']:
            return False
        if status_filter == 'running':
            if row['browser_open'] != 'running': return False
        elif status_filter == 'logged_in':
            if row['status'] != 'logged_in': return False
        elif status_filter == 'not_logged_in':
            if row['status'] in ('logged_in', 'login_failed'): return False
        elif status_filter == 'login_failed':
            if row['status'] != 'login_failed': return False
        elif status_filter == 'nst':
            if row['engine'] != 'nst': return False
        elif status_filter == 'nexus':
            if row['engine'] != 'nexus': return False
        return True

    matching_ids = [pid for pid in sorted_ids if keep(rows[pid])]
    profiles = [rows[pid]['profile'] for pid in matching_ids]

    # ── Pagination (was inadvertently dropped in the index refactor) ─────
    total = len(profiles)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    page_slice = profiles[start:end]

    # `slim=1` returns only the fields picker modals (appeal/health/etc.) need.
    # The full profile dict is ~5KB each — at 1800 profiles that's ~10MB on
    # every modal open, taking 30-60s. Slim mode drops it to ~200 bytes each.
    if request.args.get('slim') in ('1', 'true', 'yes'):
        page_slice = [{
            'id': p.get('id'),
            'name': p.get('name', ''),
            'email': p.get('email', ''),
            'group': p.get('group', ''),
            'groups': p.get('groups', []),
            'status': p.get('status', ''),
            'engine': p.get('engine', 'nexus'),
            'browser_open': p.get('browser_open', ''),
            'created_at': p.get('created_at', ''),
            'last_used': p.get('last_used'),
        } for p in page_slice]

    return jsonify({
        'success': True,
        'profiles': page_slice,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@app.route('/api/profiles/counts', methods=['GET'])
def profiles_counts():
    """Lightweight summary endpoint for filter chips + group dropdown.

    Returns ONLY counts and the list of group names — no per-profile data.
    The /api/profiles endpoint returns a paginated slice (50 rows by default),
    so the frontend can't compute global filter counts from it. Sending the
    full profile array just for counts (the old `per_page=9999` pattern)
    becomes a 10-50 MB JSON payload at 100k profiles. This endpoint reads
    the same profile list ONCE on the server and returns the small summary.
    """
    profiles = profile_manager.list_profiles()
    total = len(profiles)
    by_filter = {
        'all': total,
        'logged_in': sum(1 for p in profiles if p.get('status') == 'logged_in'),
        'not_logged_in': sum(1 for p in profiles
                              if p.get('status') not in ('logged_in', 'login_failed')),
        'login_failed': sum(1 for p in profiles if p.get('status') == 'login_failed'),
        'running': sum(1 for p in profiles if p.get('browser_open') == 'running'),
        'nst': sum(1 for p in profiles if p.get('engine', 'nexus') == 'nst'),
        'nexus': sum(1 for p in profiles if p.get('engine', 'nexus') == 'nexus'),
    }
    # Collect unique group names (preserves what the dropdown needs to render).
    groups: set = set()
    for p in profiles:
        for g in profile_manager._get_groups(p):
            if g:
                groups.add(g)
    return jsonify({
        'success': True,
        'total': total,
        'by_filter': by_filter,
        'groups': sorted(groups, key=str.lower),
    })


@app.route('/api/profiles/generate-fingerprint', methods=['POST'])
def profiles_generate_fingerprint():
    """Generate a fingerprint for preview (without creating a profile).

    In NST mode, fingerprints are managed by NST Browser — we return
    placeholder values indicating NST will handle it.
    """
    data = request.get_json(force=True, silent=True) or {}
    os_type = data.get('os', 'windows')

    # Check if NST mode is active
    use_nst = False
    try:
        import json as _json
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            bcfg = _json.loads(bj.read_text('utf-8'))
            use_nst = bcfg.get('use_nst', False)
    except Exception:
        pass

    if use_nst:
        # NST Browser handles fingerprints — return defaults with NST marker
        os_map = {'windows': 'Windows 11', 'macos': 'macOS 14', 'linux': 'Linux'}
        fp = {
            'overview': {
                'os': os_type,
                'os_version': os_map.get(os_type, 'Windows 11'),
                'device_type': 'desktop',
                'browser_kernel': 'nstbrowser',
                'kernel_version': 133,
                'user_agent': '(managed by NST Browser)',
                'platform': 'nst',
                'startup_urls': [],
            },
            'hardware': {
                'webgl': 'noise', 'webgl_metadata': 'masked',
                'webgl_vendor': '(managed by NST Browser)',
                'webgl_renderer': '(managed by NST Browser)',
                'canvas': 'noise', 'canvas_seed': 0,
                'audio_context': 'noise', 'audio_seed': 0,
                'client_rects': 'real', 'speech_voice': 'masked',
                'media_devices': {'mode': 'custom', 'video_inputs': 0,
                                  'audio_inputs': 1, 'audio_outputs': 1},
                'battery': 'masked',
                'hardware_concurrency': 4, 'device_memory': 8,
                'device_name': '', 'mac_address': '',
                'hardware_acceleration': True,
            },
            'advanced': {
                'language': 'based_on_ip', 'language_value': '',
                'timezone': 'based_on_ip', 'timezone_value': '',
                'geolocation_prompt': 'prompt', 'geolocation_source': 'based_on_ip',
                'webrtc': 'masked',
                'screen_resolution': 'custom',
                'screen_width': 1920, 'screen_height': 1080,
                'fonts': 'masked', 'do_not_track': False,
                'port_scan_protection': 'disabled',
                'disable_image_loading': False, 'save_tabs': True,
                'launch_args': '',
            },
        }
        return jsonify({'success': True, 'fingerprint': fp, 'nst_mode': True})

    # NST mode not active — return empty placeholder
    return jsonify({'success': False, 'message': 'NST Browser required for fingerprint generation'}), 400


@app.route('/api/profiles', methods=['POST'])
def profiles_create():
    """Create a new profile."""
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Profile name is required'}), 400
    proxy = data.get('proxy')
    notes = data.get('notes', '')
    fingerprint_prefs = data.get('fingerprint_prefs', {})
    password = data.get('password', '')
    totp_secret = data.get('totp_secret', '')
    backup_codes = data.get('backup_codes', [])
    recovery_email = data.get('recovery_email', '').strip()
    recovery_phone = data.get('recovery_phone', '').strip()
    # Pass full overview/hardware/advanced from frontend (if user edited them)
    frontend_sections = {}
    if data.get('overview'):
        frontend_sections['overview'] = data['overview']
    if data.get('hardware'):
        frontend_sections['hardware'] = data['hardware']
    if data.get('advanced'):
        frontend_sections['advanced'] = data['advanced']
    engine = data.get('engine', 'nexus')
    if engine not in ('nst', 'nexus'):
        engine = 'nexus'
    group = (data.get('group', '') or 'default').strip()
    fingerprint_config = data.get('fingerprint_config')
    profile = profile_manager.create_profile(
        name, email, proxy=proxy, notes=notes,
        fingerprint_prefs=fingerprint_prefs,
        password=password, totp_secret=totp_secret, backup_codes=backup_codes,
        recovery_email=recovery_email, recovery_phone=recovery_phone,
        frontend_sections=frontend_sections,
        engine=engine,
    )
    # Set group (not accepted by create_profile directly)
    if group:
        profile_manager.update_profile(profile['id'], group=group)
        profile['group'] = group
    # Persist fingerprint_config if provided
    if fingerprint_config and profile:
        profile_manager.update_profile(profile['id'], fingerprint_config=fingerprint_config)
        profile['fingerprint_config'] = fingerprint_config
    # Persist bookmarks_text if provided — applied to Chrome data dir on launch
    bookmarks_text = (data.get('bookmarks_text') or '').strip()
    if bookmarks_text and profile:
        profile_manager.update_profile(profile['id'], bookmarks_text=bookmarks_text)
        profile['bookmarks_text'] = bookmarks_text
    # Background: fetch real country + outbound IP through the proxy and persist.
    if profile and profile.get('id') and (proxy or {}).get('host'):
        _refresh_proxy_geo_async(profile['id'])
    nst_err = profile.pop('_nst_create_error', None)
    if nst_err:
        # Profile was created locally — return success with warning
        return jsonify({'success': True, 'profile': profile, 'warning': nst_err})
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/<profile_id>', methods=['GET'])
def profiles_get(profile_id):
    """Get a single profile."""
    profile = profile_manager.get_profile(profile_id)
    if not profile:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/<profile_id>', methods=['PUT'])
def profiles_update(profile_id):
    """Update a profile."""
    data = request.get_json(force=True, silent=True) or {}
    profile = profile_manager.update_profile(profile_id, **data)
    if not profile:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/<profile_id>/proxy-country', methods=['GET'])
def profiles_proxy_country(profile_id):
    """Look up the ACTUAL session country + outbound IP by tunneling through the profile's proxy.

    Uses shared.nexus_proxy_manager.check_proxy() which makes a real HTTP request via the
    proxy to ip-api.com — so the IP returned is what websites would actually see, not the
    geolocation of the proxy's gateway host. Session-based proxies rotate IPs, so the
    current_ip field reflects the IP at the time of the call.

    Cache: country/country_code/current_ip are cached on profile.proxy. Pass ?refresh=1
    to bust the cache and re-check live.
    """
    from shared.nexus_proxy_manager import check_proxy

    profile = profile_manager.get_profile(profile_id)
    if not profile:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404

    proxy = profile.get('proxy') or {}
    if not proxy.get('host') and not proxy.get('server'):
        return jsonify({'success': True, 'country': 'No proxy', 'country_code': '', 'current_ip': ''})

    refresh = request.args.get('refresh') in ('1', 'true', 'yes')
    # Return cached value only if it's a real positive hit (not the "Unknown" fallback
    # left behind by an earlier failed SOCKS5 lookup).
    cached_country = (proxy.get('country') or '').strip()
    cached_cc = (proxy.get('country_code') or '').strip()
    cache_is_real = cached_country and cached_cc and cached_country.lower() != 'unknown'
    if not refresh and cache_is_real:
        return jsonify({
            'success': True,
            'country': cached_country,
            'country_code': cached_cc,
            'current_ip': proxy.get('current_ip', ''),
            'cached': True,
        })

    info = check_proxy(proxy, timeout=12)
    if info.get('success'):
        country = info.get('country', 'Unknown')
        cc = info.get('country_code', '')
        ip = info.get('ip', '')
        fresh = profile_manager.get_profile(profile_id)
        if fresh is not None:
            merged_proxy = dict(fresh.get('proxy') or {})
            merged_proxy['country'] = country
            merged_proxy['country_code'] = cc
            merged_proxy['current_ip'] = ip
            profile_manager.update_profile(profile_id, proxy=merged_proxy)
        return jsonify({
            'success': True,
            'country': country,
            'country_code': cc,
            'current_ip': ip,
            'cached': False,
        })

    err = info.get('error', 'lookup failed')
    add_log(f"[proxy-country] {profile_id} via proxy {proxy.get('host')}:{proxy.get('port')} → {err}", 'warning')
    return jsonify({'success': True, 'country': 'Unknown', 'country_code': '', 'current_ip': '', 'error': err})


@app.route('/api/profiles/delete-all', methods=['DELETE'])
def profiles_delete_all():
    """Delete ALL profiles and their data directories."""
    try:
        profiles = profile_manager.list_profiles()
        count = len(profiles)
        for p in profiles:
            try:
                profile_manager.delete_profile(p['id'])
            except Exception:
                pass
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} profiles'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/profiles/delete-by-engine/<engine>', methods=['DELETE'])
def profiles_delete_by_engine(engine):
    """Delete all profiles matching a specific engine (nst or nexus)."""
    if engine not in ('nst', 'nexus'):
        return jsonify({'success': False, 'message': 'Engine must be nst or nexus'}), 400
    try:
        count = profile_manager.delete_all_by_engine(engine)
        label = 'NST' if engine == 'nst' else 'Local'
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} {label} profiles'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


_bulk_delete_progress = {
    'running': False, 'status': 'idle',
    'total': 0, 'deleted': 0, 'failed': 0, 'pending': 0,
    'current_profile': '',
}

@app.route('/api/profiles/delete-bulk', methods=['DELETE'])
def profiles_delete_bulk():
    """Delete multiple profiles by IDs with progress tracking."""
    global _bulk_delete_progress
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No profile IDs provided'}), 400

    if _bulk_delete_progress.get('running'):
        return jsonify({'success': False, 'message': 'Bulk delete already running'}), 409

    # Set progress BEFORE spawning thread
    _bulk_delete_progress.update({
        'running': True, 'status': 'processing',
        'total': len(ids), 'deleted': 0, 'failed': 0, 'pending': len(ids),
        'current_profile': '',
    })

    def _worker():
        global _bulk_delete_progress
        deleted = 0
        failed = 0
        for pid in ids:
            # Get profile name for progress display
            try:
                p = profile_manager.get_profile(pid)
                name = p.get('name', p.get('email', pid)) if p else pid
            except Exception:
                name = pid
            _bulk_delete_progress['current_profile'] = name
            try:
                if profile_manager.delete_profile(pid):
                    deleted += 1
                else:
                    failed += 1
            except Exception:
                failed += 1
            _bulk_delete_progress.update({
                'deleted': deleted, 'failed': failed,
                'pending': max(0, len(ids) - deleted - failed),
            })
        _bulk_delete_progress.update({
            'running': False, 'status': 'completed',
            'deleted': deleted, 'failed': failed, 'pending': 0,
            'current_profile': '',
        })

    import threading
    threading.Thread(target=_worker, daemon=True, name='bulk-delete').start()
    return jsonify({'success': True, 'total': len(ids), 'message': 'Bulk delete started'})


@app.route('/api/profiles/delete-bulk-status', methods=['GET'])
def profiles_delete_bulk_status():
    """Return current bulk delete progress."""
    return jsonify({'success': True, 'progress': dict(_bulk_delete_progress)})


@app.route('/api/profiles/duplicates', methods=['GET'])
def profiles_duplicates():
    """Find all profiles that share the same email address."""
    try:
        from collections import defaultdict
        profiles = profile_manager.list_profiles()
        groups = defaultdict(list)
        for p in profiles:
            email = (p.get('email') or '').strip().lower()
            if email:
                groups[email].append(p)
        duplicates = []
        for email, profs in groups.items():
            if len(profs) >= 2:
                profs_sorted = sorted(profs, key=lambda x: x.get('created_at', ''), reverse=True)
                duplicates.append({
                    'email': email,
                    'count': len(profs_sorted),
                    'profiles': profs_sorted,
                })
        duplicates.sort(key=lambda x: x['count'], reverse=True)
        return jsonify({'success': True, 'duplicates': duplicates, 'total_groups': len(duplicates)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/profiles/<profile_id>', methods=['DELETE'])
def profiles_delete(profile_id):
    """Delete a profile and its data directory."""
    ok = profile_manager.delete_profile(profile_id)
    if not ok:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'message': 'Profile deleted'})


def _refresh_proxy_geo_async(profile_id: str):
    """Run check_proxy() in a background thread and persist country/cc/current_ip.

    Fired after profile-create or profile-launch so the row shows fresh data without
    forcing the user to wait for a 5–10s proxy round-trip in the API response.
    """
    import threading
    from shared.nexus_proxy_manager import check_proxy

    def _job():
        try:
            p = profile_manager.get_profile(profile_id)
            if not p:
                return
            proxy = p.get('proxy') or {}
            if not (proxy.get('host') or proxy.get('server')):
                return
            info = check_proxy(proxy, timeout=12)
            if not info.get('success'):
                add_log(f"[proxy-geo bg] {profile_id} → {info.get('error','failed')}", 'debug')
                return
            fresh = profile_manager.get_profile(profile_id) or {}
            merged = dict(fresh.get('proxy') or {})
            merged['country'] = info.get('country', merged.get('country', ''))
            merged['country_code'] = info.get('country_code', merged.get('country_code', ''))
            merged['current_ip'] = info.get('ip', '')
            profile_manager.update_profile(profile_id, proxy=merged)
        except Exception as e:
            add_log(f"[proxy-geo bg] {profile_id} exception: {e}", 'warning')

    threading.Thread(target=_job, daemon=True, name=f'proxy-geo-{profile_id[:8]}').start()


@app.route('/api/profiles/<profile_id>/launch', methods=['POST'])
def profiles_launch(profile_id):
    """Launch a profile browser."""
    result = profile_manager.launch_profile(profile_id)
    if not result['success']:
        return jsonify(result), 400
    # Fire-and-forget IP/country refresh — user wants it updated only at meaningful moments.
    _refresh_proxy_geo_async(profile_id)
    return jsonify(result)


@app.route('/api/profiles/<profile_id>/close', methods=['POST'])
def profiles_close(profile_id):
    """Close a profile browser — INSTANT response.

    Both close_profile and stop_profile_browser are now fast (direct kill,
    no graceful asyncio waits). We still call both so a tracked profile
    gets its launcher killed even if psutil somehow misses it.
    The wait for OS process-table updates happens in a background thread
    so the HTTP response returns immediately — the user sees the button
    flip back to Open without waiting on Windows to settle.
    """
    import threading as _t

    managed_ok = False
    try:
        managed_ok = bool(profile_manager.close_profile(profile_id))
    except Exception:
        pass

    # Fire the broad psutil sweep in a background thread — it's all force-kills,
    # so the work is essentially synchronous OS calls; we just don't block on
    # the post-kill wait_procs poll.
    def _bg_kill():
        try:
            profile_manager.stop_profile_browser(profile_id)
        except Exception:
            pass
    _t.Thread(target=_bg_kill, daemon=True, name=f'close-{profile_id[:8]}').start()

    return jsonify({'success': True, 'managed': managed_ok})


@app.route('/api/profiles/close-all', methods=['POST'])
def profiles_close_all():
    """Close all open profile browsers."""
    profile_manager.close_all_profiles()
    return jsonify({'success': True})


# ── Performance / Fast Mode ──────────────────────────────────────────────────
# Allowed perf keys mirror DEFAULT_PERF in nexus_profile_manager. Anything else
# gets silently dropped so the frontend can't inject arbitrary settings.
_ALLOWED_PERF_KEYS = {
    'block_images', 'block_autoplay', 'disable_hw_accel',
    'block_notifications', 'block_popups',
}


def _sanitize_perf(perf_in):
    if not isinstance(perf_in, dict):
        return {}
    return {k: bool(v) for k, v in perf_in.items() if k in _ALLOWED_PERF_KEYS}


@app.route('/api/profiles/<profile_id>/perf', methods=['GET'])
def profiles_get_perf(profile_id):
    """Return a profile's performance/fast-mode settings."""
    p = profile_manager.get_profile(profile_id)
    if not p:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    from shared.nexus_profile_manager import DEFAULT_PERF
    perf = {**DEFAULT_PERF, **(p.get('perf') or {})}
    return jsonify({'success': True, 'perf': perf})


@app.route('/api/profiles/<profile_id>/perf', methods=['PATCH', 'POST'])
def profiles_set_perf(profile_id):
    """Update a profile's performance/fast-mode settings. Body: { perf: {...} }."""
    data = request.get_json(force=True, silent=True) or {}
    incoming = _sanitize_perf(data.get('perf'))
    p = profile_manager.get_profile(profile_id)
    if not p:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    merged = {**(p.get('perf') or {}), **incoming}
    profile_manager.update_profile(profile_id, perf=merged)
    return jsonify({'success': True, 'perf': merged})


@app.route('/api/profiles/bulk-perf', methods=['POST'])
def profiles_bulk_perf():
    """Apply perf settings to multiple profiles.

    Body:
      { perf: {block_images: true, ...},
        scope: 'selected' | 'all' | 'group',
        ids:   [...],          # required if scope=selected
        group: 'group-name' }  # required if scope=group

    Only the keys present in `perf` are written — others stay unchanged on
    each target. So toggling only "images" doesn't clobber a profile's
    autoplay setting.
    """
    data = request.get_json(force=True, silent=True) or {}
    incoming = _sanitize_perf(data.get('perf'))
    if not incoming:
        return jsonify({'success': False, 'message': 'No valid perf keys provided'}), 400

    scope = (data.get('scope') or 'selected').lower()
    target_ids = []
    all_profiles = profile_manager.list_profiles()
    if scope == 'all':
        target_ids = [p['id'] for p in all_profiles]
    elif scope == 'group':
        group_name = (data.get('group') or '').strip()
        if not group_name:
            return jsonify({'success': False, 'message': 'group name required for scope=group'}), 400
        for p in all_profiles:
            groups = p.get('groups') or ([p.get('group')] if p.get('group') else [])
            if group_name in groups:
                target_ids.append(p['id'])
    else:  # selected
        target_ids = [i for i in (data.get('ids') or []) if i]

    if not target_ids:
        return jsonify({'success': False, 'message': 'No target profiles'}), 400

    # Run as a background worker so the multi-card progress popup can show
    # live progress (especially for "All profiles" scope with thousands of writes).
    num_workers = max(1, min(int(data.get('workers') or data.get('num_workers') or 5), 20))
    result = profile_manager.bulk_apply_perf_async(target_ids, incoming, num_workers=num_workers)
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify({
        'success': True,
        'started': True,
        'total': result.get('total', len(target_ids)),
        'requested': len(target_ids),
        'perf': incoming,
    })


@app.route('/api/profiles/bulk-perf-status', methods=['GET'])
def profiles_bulk_perf_status():
    """Live progress for bulk Fast Mode apply."""
    try:
        return jsonify(profile_manager.get_bulk_perf_status())
    except Exception:
        return jsonify({'running': False, 'status': 'idle'})


@app.route('/api/profiles/bulk-bookmark-status', methods=['GET'])
def profiles_bulk_bookmark_status():
    """Live progress for bulk bookmark apply."""
    try:
        return jsonify(profile_manager.get_bulk_bookmark_status())
    except Exception:
        return jsonify({'running': False, 'status': 'idle'})


@app.route('/api/profiles/cleanup', methods=['POST'])
def profiles_cleanup():
    """Delete orphan profile folders not in profiles.json."""
    result = profile_manager.cleanup_orphans()
    return jsonify({'success': True, **result})


@app.route('/api/profiles/<profile_id>/change-device-type', methods=['POST'])
def profiles_change_device_type(profile_id):
    """Change device type of an existing profile (mobile ↔ desktop).
    Body: {os_type: 'windows'|'macos'|'linux'|'android'|'ios'}
    Updates fingerprint (screen, UA, hardware, etc.) accordingly."""
    data = request.get_json(force=True, silent=True) or {}
    os_type = data.get('os_type', '').strip().lower()
    if not os_type:
        return jsonify({'success': False, 'error': 'os_type required'}), 400
    result = profile_manager.change_device_type(profile_id, os_type)
    status = 200 if result.get('success') else 400
    return jsonify(result), status


# ── Live Status Check (GMB review URLs) ──────────────────────────────────────
from shared import live_status_check as _live_check


def _is_url_like(s: str) -> bool:
    """True only if the cell value is a real http(s) URL — skips notes,
    empty strings, formula error markers, etc. so the link count matches
    what the user actually sees in the column."""
    if not s:
        return False
    s = s.strip().lower()
    return s.startswith('http://') or s.startswith('https://')


@app.route('/api/profiles/live-check/preview', methods=['POST'])
def profiles_live_check_preview():
    """Read the Excel file and report how many rows have a Review Live Link."""
    body = request.get_json(silent=True) or {}
    file_path = (body.get('file_path') or '').strip()
    if not file_path:
        return jsonify({'success': False, 'message': 'file_path is required'}), 400
    try:
        import openpyxl
        from pathlib import Path as _P
        if not _P(file_path).exists():
            return jsonify({'success': False, 'message': f'File not found: {file_path}'})
        # Use default (non-read-only) mode for accurate row counting; read-only
        # can report inflated max_row when the sheet has trailing styling.
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in ws[1]]
        link_idx = next(
            (i + 1 for i, h in enumerate(headers)
             if h.strip().lower() == 'review live link'),
            None,
        )
        if link_idx is None:
            wb.close()
            return jsonify({'success': False,
                            'message': "Header 'Review Live Link' not found in the file."})
        seen = set()
        non_empty = 0
        duplicates = 0
        first_url_row = None
        last_url_row = None
        sample_first = ''
        sample_last = ''
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=link_idx).value
            if v is None:
                continue
            url = str(v).strip()
            if not url:
                continue
            # Only count actual URLs (skip stray text / whitespace junk)
            if not _is_url_like(url):
                continue
            non_empty += 1
            if first_url_row is None:
                first_url_row = r
                sample_first = url
            last_url_row = r
            sample_last = url
            key = url.lower()
            if key in seen:
                duplicates += 1
            else:
                seen.add(key)
        # Convert column index (1-based) to Excel letter for the user's
        # cross-reference (so they can verify it's the right column).
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(link_idx)
        wb.close()
        return jsonify({
            'success': True,
            'file_name': _P(file_path).name,
            'total_links': non_empty,
            'unique_links': len(seen),
            'duplicates': duplicates,
            'header_column': headers[link_idx - 1],
            'column_letter': col_letter,
            'first_row': first_url_row,
            'last_row': last_url_row,
            'sample_first': sample_first[:80],
            'sample_last': sample_last[:80],
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Read error: {e}'})


@app.route('/api/profiles/live-check/start', methods=['POST'])
def profiles_live_check_start():
    body = request.get_json(silent=True) or {}
    workers = int(body.get('workers') or 5)
    timeout = int(body.get('timeout_sec') or 20)
    show_browser = bool(body.get('show_browser', False))

    # Status filter — list of canonical Status values to limit the run to.
    # Empty list (the default and the "All" case) means "check every row".
    # Specific values like 'live' / 'appealed' / 'done' / 'missing' /
    # 'disabled' restrict the run to rows whose Status column already
    # holds that value.
    sf_raw = body.get('status_filter') or []
    status_filter: list[str] = []
    if isinstance(sf_raw, list):
        for s in sf_raw:
            sv = str(s or '').strip().lower()
            if sv and sv != 'all':
                status_filter.append(sv)

    sheet_id = (body.get('sheet_id') or '').strip()
    tab_name = (body.get('tab_name') or '').strip()
    # Multi-tab support: accept a 'tabs' list of tab names. Falls back to
    # the legacy 'tab_name' field when only one tab is selected, so older
    # callers keep working untouched.
    tabs_raw = body.get('tabs')
    tabs: list[str] = []
    if isinstance(tabs_raw, list):
        tabs = [str(t).strip() for t in tabs_raw if str(t or '').strip()]
    elif tab_name:
        tabs = [tab_name]
    file_path = (body.get('file_path') or '').strip()

    if sheet_id and tabs:
        if len(tabs) == 1:
            return jsonify(_live_check.start_from_sheet(
                sheet_id, tabs[0], workers, timeout,
                RESOURCES_PATH, show_browser,
                status_filter=status_filter,
            ))
        return jsonify(_live_check.start_from_sheet_tabs(
            sheet_id, tabs, workers, timeout,
            RESOURCES_PATH, show_browser,
            status_filter=status_filter,
        ))
    if not file_path:
        return jsonify({'success': False,
                        'message': 'Provide either file_path or sheet_id+tabs'}), 400
    return jsonify(_live_check.start(file_path, workers, timeout,
                                     RESOURCES_PATH, show_browser,
                                     status_filter=status_filter))


# ── Sheets integration endpoints ─────────────────────────────────────────────
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GOOGLE SHEETS API — routes/sheets.py Blueprint (registered below)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Routes are in electron-app/backend/routes/sheets.py.
# Registered at bottom of file alongside other Blueprints.


@app.route('/api/profiles/live-check/status', methods=['GET'])
def profiles_live_check_status():
    return jsonify(_live_check.get_status())


@app.route('/api/profiles/live-check/cancel', methods=['POST'])
def profiles_live_check_cancel():
    _live_check.cancel()
    return jsonify({'success': True})


@app.route('/api/profiles/review-stats', methods=['GET'])
def api_review_stats_all():
    """Bulk cached fetch (counts only, no `reviews` array)."""
    return jsonify({
        'success': True,
        'stats':   review_stats.get_all_stats(),
    })


@app.route('/api/profiles/<profile_id>/review-stats', methods=['GET'])
def api_review_stats_one(profile_id):
    """Full per-profile record incl. reviews array."""
    record = review_stats.get_profile_stats(profile_id)
    if record is None:
        return jsonify({'success': False, 'message': 'Never scanned'}), 404
    return jsonify({'success': True, 'stats': record})


@app.route('/api/profiles/review-stats/scan', methods=['POST'])
def api_review_stats_scan():
    """Start a bulk scan."""
    body = request.get_json(silent=True) or {}
    profile_ids = body.get('profile_ids')   # None means scan all
    num_workers = int(body.get('num_workers') or 3)
    result = review_stats.start_scan(profile_ids=profile_ids,
                                     num_workers=num_workers)
    if not result.get('success'):
        return jsonify(result), 409
    return jsonify(result)


@app.route('/api/profiles/review-stats/status', methods=['GET'])
def api_review_stats_status():
    return jsonify(review_stats.get_status())


@app.route('/api/profiles/review-stats/cancel', methods=['POST'])
def api_review_stats_cancel():
    review_stats.cancel_scan()
    return jsonify({'success': True})


# ── Profile Drive Backup / Restore ───────────────────────────────────────────
from shared import drive_backup as _drive_backup
try:
    _drive_backup.start_auto_backup_loop(RESOURCES_PATH)
except Exception:
    pass


@app.route('/api/profiles/drive/status', methods=['GET'])
def profiles_drive_status():
    return jsonify(_drive_backup.status(RESOURCES_PATH))


@app.route('/api/profiles/drive/backup', methods=['POST'])
def profiles_drive_backup():
    return jsonify(_drive_backup.backup_now(RESOURCES_PATH))


@app.route('/api/profiles/drive/backups', methods=['GET'])
def profiles_drive_backups():
    return jsonify(_drive_backup.list_backups(RESOURCES_PATH))


@app.route('/api/profiles/drive/restore', methods=['POST'])
def profiles_drive_restore():
    body = request.get_json(silent=True) or {}
    file_id = (body.get('file_id') or '').strip()
    if not file_id:
        return jsonify({'success': False, 'message': 'file_id is required'}), 400
    return jsonify(_drive_backup.restore(RESOURCES_PATH, file_id))


@app.route('/api/profiles/drive/reauthorize', methods=['POST'])
def profiles_drive_reauthorize():
    """Run OAuth2 in a background thread so the request doesn't time out
    while we wait for the user to log in via their browser."""
    import threading
    state = {'done': False, 'result': None}

    def _worker():
        try:
            state['result'] = _drive_backup.reauthorize(RESOURCES_PATH)
        except Exception as e:
            state['result'] = {'success': False, 'message': str(e)}
        state['done'] = True

    t = threading.Thread(target=_worker, daemon=True, name='drive-reauth')
    t.start()
    # Wait up to 3 minutes for the user to complete the flow
    t.join(timeout=180)
    if not state['done']:
        return jsonify({'success': False,
                        'message': 'Timed out waiting for OAuth consent. Try again.'})
    return jsonify(state['result'])


@app.route('/api/profiles/drive/auto-backup', methods=['POST'])
def profiles_drive_auto_backup():
    body = request.get_json(silent=True) or {}
    enabled = bool(body.get('enabled', False))
    interval = int(body.get('interval_hours') or 24)
    res = _drive_backup.set_auto_backup(RESOURCES_PATH, enabled, interval)
    if enabled:
        try:
            _drive_backup.start_auto_backup_loop(RESOURCES_PATH)
        except Exception:
            pass
    return jsonify(res)


# ── Chrome Extension Manager ─────────────────────────────────────────────────
from shared import extension_manager as _ext_mgr


@app.route('/api/extensions', methods=['GET'])
def extensions_list():
    return jsonify({'success': True, 'extensions': _ext_mgr.list_extensions(RESOURCES_PATH)})


@app.route('/api/extensions/install-url', methods=['POST'])
def extensions_install_url():
    body = request.get_json(silent=True) or {}
    url = (body.get('url') or '').strip()
    if not url:
        return jsonify({'success': False, 'message': 'url is required'}), 400
    return jsonify(_ext_mgr.install_from_url(RESOURCES_PATH, url))


@app.route('/api/extensions/install-zip', methods=['POST'])
def extensions_install_zip():
    f = request.files.get('file')
    if not f:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    name = (request.form.get('name') or '').strip()
    data = f.read()
    return jsonify(_ext_mgr.install_from_bytes(RESOURCES_PATH, data, f.filename or '', name))


@app.route('/api/extensions/<ext_id>', methods=['DELETE'])
def extensions_delete(ext_id):
    return jsonify(_ext_mgr.remove_extension(RESOURCES_PATH, ext_id))


@app.route('/api/extensions/<ext_id>', methods=['PATCH'])
def extensions_update(ext_id):
    body = request.get_json(silent=True) or {}
    kwargs = {}
    if 'apply_to_all' in body:
        kwargs['apply_to_all'] = bool(body['apply_to_all'])
    if 'pinned' in body:
        kwargs['pinned'] = bool(body['pinned'])
    return jsonify(_ext_mgr.update_extension(RESOURCES_PATH, ext_id, **kwargs))


@app.route('/api/profiles/restore-from-nst', methods=['POST'])
def profiles_restore_from_nst():
    """Recover profiles missing from local profiles.json by pulling them from NST.

    Body: {"group": "Sanjid" (optional), "dry_run": true|false (default false)}
    """
    body = request.get_json(silent=True) or {}
    group = body.get('group') or None
    dry_run = bool(body.get('dry_run', False))
    try:
        result = profile_manager.restore_missing_from_nst(group=group, dry_run=dry_run)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/<profile_id>/status', methods=['GET'])
def profiles_status(profile_id):
    """Get browser status for a profile."""
    return jsonify(profile_manager.profile_status(profile_id))


@app.route('/api/profiles/status', methods=['GET'])
def profiles_all_status():
    """Get counts of open/total profile browsers."""
    return jsonify(profile_manager.all_status())


@app.route('/api/profiles/batch-login-preview', methods=['POST'])
def profiles_batch_login_preview():
    """Read Excel and return count of valid accounts without running login."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path)
        total = len(df)
        existing_emails = {p.get('email', '').lower() for p in profile_manager.list_profiles()}
        valid = 0
        skipped = 0
        for _, row in df.iterrows():
            email = str(row.get('Email', '')).strip()
            password = str(row.get('Password', '')).strip()
            if email and password and email.lower() != 'nan' and password.lower() != 'nan':
                valid += 1
                if email.lower() in existing_emails:
                    skipped += 1
        cols = list(df.columns)
        return jsonify({'success': True, 'total': total, 'valid': valid, 'skipped': skipped, 'columns': cols})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/profiles/batch-login', methods=['POST'])
def profiles_batch_login():
    """Start batch login from Excel file.

    Optional `perf` body field carries Fast Mode toggles that get applied to
    every profile this batch touches (newly created or matched-existing) —
    blocking images during the login itself massively cuts MB usage.
    """
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    num_workers = int(data.get('workers', 3))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'random')
    group = data.get('group', 'default') or 'default'
    stagger_delay = max(0, min(int(data.get('stagger_delay', 0)), 120))
    raw_perf = data.get('perf') or {}
    perf = _sanitize_perf(raw_perf) if isinstance(raw_perf, dict) else {}
    if not file_path:
        return jsonify({'success': False, 'message': 'File path is required'}), 400
    result = profile_manager.batch_login(
        file_path, num_workers, engine=engine, os_type=os_type, group=group,
        stagger_delay=stagger_delay, perf=perf,
    )
    return jsonify(result)


@app.route('/api/profiles/config', methods=['GET'])
def profiles_config_get():
    """Get profile storage config."""
    return jsonify({'success': True, 'config': profile_manager.get_config()})


@app.route('/api/profiles/config', methods=['POST'])
def profiles_config_set():
    """Set profile storage path."""
    data = request.get_json(force=True, silent=True) or {}
    storage_path = data.get('storage_path', '')
    try:
        config = profile_manager.set_storage_path(storage_path)
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


# ── Profile Manager: Do All Appeal ────────────────────────────────────────────
from shared import sheets_integration as _sheets_int

@app.route('/api/profiles/do-all-appeal', methods=['POST'])
def profiles_do_all_appeal():
    """Start Do All Appeal for selected profiles."""
    data = request.get_json(force=True, silent=True) or {}
    num_workers = int(data.get('num_workers', 5))
    profile_ids = data.get('profile_ids', [])
    result = profile_manager.do_all_appeal_profiles(num_workers=num_workers, profile_ids=profile_ids)
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/appeal-status', methods=['GET'])
def profiles_appeal_status():
    """Get status of running Do All Appeal operation."""
    return jsonify(profile_manager.get_appeal_status())


@app.route('/api/profiles/appeal/match-sheet', methods=['POST'])
def profiles_appeal_match_sheet():
    """Preview: how many rows have Status='Missing' across the selected
    tab(s), and how many of their emails match existing profiles.

    Body accepts either:
      - tabs: ['Tab A', 'Tab B', ...]   ← preferred (multi-tab aggregate)
      - tab_name: 'Tab A'               ← legacy single-tab fallback

    Mirrors the scan that do_all_appeal_from_sheet_tabs runs at start —
    so the preview number matches what Start Appeal will actually do.
    """
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    tab_name = (body.get('tab_name') or '').strip()
    target = (body.get('target_status') or 'Missing').strip()
    tabs_raw = body.get('tabs')
    tabs: list[str] = []
    if isinstance(tabs_raw, list):
        tabs = [str(t).strip() for t in tabs_raw if str(t or '').strip()]
    elif tab_name:
        tabs = [tab_name]
    if not sheet_id or not tabs:
        return jsonify({'success': False, 'message': 'sheet_id + tabs[] (or tab_name) required'}), 400

    all_profiles = profile_manager.list_profiles()
    by_email = {(p.get('email') or '').strip().lower(): p for p in all_profiles}

    # ONE batchGet for every selected tab — beats N sequential reads
    # that previously froze the modal on 5+ tab selections.
    batch_res = _sheets_int.batch_read_rows_by_status(
        RESOURCES_PATH, sheet_id, tabs, target,
    )
    if not batch_res.get('success'):
        return jsonify({'success': False,
                        'message': batch_res.get('message') or 'Batch read failed'}), 400

    total_missing_rows = 0
    matched: set = set()
    not_found: set = set()
    value_counts: dict = {}
    status_header_used: str = ''
    tab_errors: list = []
    per_tab: list = []

    for t in (batch_res.get('tabs') or []):
        tab = t.get('tab')
        if not t.get('success'):
            tab_errors.append({'tab': tab, 'message': t.get('message') or 'read failed'})
            per_tab.append({'tab': tab, 'success': False,
                            'missing_rows': 0, 'matched': 0, 'not_found': 0,
                            'message': t.get('message') or 'read failed'})
            continue
        if not status_header_used:
            status_header_used = t.get('status_header_used') or ''
        for v in (t.get('unique_status_values') or []):
            k = v.get('value', '')
            value_counts[k] = value_counts.get(k, 0) + int(v.get('count') or 0)
        rows = t.get('rows') or []
        tab_matched: set = set()
        tab_not_found: set = set()
        for r in rows:
            em = (r.get('email') or '').strip().lower()
            if not em:
                continue
            if em in by_email:
                tab_matched.add(em)
                matched.add(em)
            else:
                tab_not_found.add(em)
                not_found.add(em)
        total_missing_rows += len(rows)
        per_tab.append({
            'tab': tab,
            'success': True,
            'missing_rows': len(rows),
            'matched': len(tab_matched),
            'not_found': len(tab_not_found),
        })

    aggregated_status_values = sorted(
        ({'value': k, 'count': v} for k, v in value_counts.items()),
        key=lambda x: -x['count'],
    )[:12]

    return jsonify({
        'success': True,
        'total_missing_rows': total_missing_rows,
        'matched_count': len(matched),
        'not_found_count': len(not_found),
        'not_found_sample': list(not_found)[:5],
        'tabs_scanned': len(tabs) - len(tab_errors),
        'tabs_failed': tab_errors,
        # Per-tab breakdown so the UI can show "X missing" next to each tab
        'per_tab': per_tab,
        # Diagnostic — surface which column was used + what values are
        # actually present so a 0-match preview is debuggable from the UI.
        'status_header_used': status_header_used,
        'unique_status_values': aggregated_status_values,
    })


@app.route('/api/profiles/appeal/match-sheet-batch', methods=['POST'])
def profiles_appeal_match_sheet_batch():
    """Per-tab match preview for the Appeal modal's checkbox list.
    Body: {sheet_id, tabs: [name, ...], target_status?}.
    Uses a single batchGet across all tabs so the picker stays
    responsive even with 10+ tabs selected.
    Returns per-tab matched / not-found counts so the picker can show
    inline numbers next to each tab."""
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    target = (body.get('target_status') or 'Missing').strip()
    tabs_raw = body.get('tabs') or []
    if not isinstance(tabs_raw, list):
        return jsonify({'success': False, 'message': 'tabs must be a list'}), 400
    tabs = [str(t).strip() for t in tabs_raw if str(t or '').strip()]
    if not sheet_id or not tabs:
        return jsonify({'success': False, 'message': 'sheet_id + tabs[] required'}), 400

    all_profiles = profile_manager.list_profiles()
    by_email = {(p.get('email') or '').strip().lower(): p for p in all_profiles}

    batch_res = _sheets_int.batch_read_rows_by_status(
        RESOURCES_PATH, sheet_id, tabs, target,
    )
    if not batch_res.get('success'):
        return jsonify({'success': False,
                        'message': batch_res.get('message') or 'Batch read failed'}), 400

    out = []
    for t in (batch_res.get('tabs') or []):
        tab = t.get('tab')
        if not t.get('success'):
            out.append({
                'tab': tab, 'success': False,
                'message': t.get('message') or 'read failed',
            })
            continue
        rows = t.get('rows') or []
        matched: set = set()
        not_found: set = set()
        for r in rows:
            em = (r.get('email') or '').strip().lower()
            if not em:
                continue
            if em in by_email:
                matched.add(em)
            else:
                not_found.add(em)
        out.append({
            'tab': tab,
            'success': True,
            'total_missing_rows': len(rows),
            'matched_count': len(matched),
            'not_found_count': len(not_found),
        })
    return jsonify({'success': True, 'tabs': out})


@app.route('/api/profiles/appeal/start-from-sheet', methods=['POST'])
def profiles_appeal_start_from_sheet():
    """Run Do All Appeal on profiles whose email appears in rows of one
    or more Google Sheet tabs where Status='Missing'. After each appeal
    finishes the corresponding row(s) on every tab are updated with
    'Appealed' (success) or 'Failed' (error)."""
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    tab_name = (body.get('tab_name') or '').strip()
    tabs_raw = body.get('tabs')
    workers = int(body.get('workers') or 3)
    # Multi-tab: accept either tabs[] (preferred) or a single tab_name.
    tabs: list = []
    if isinstance(tabs_raw, list):
        tabs = [str(t).strip() for t in tabs_raw if str(t or '').strip()]
    elif tab_name:
        tabs = [tab_name]
    if not sheet_id or not tabs:
        return jsonify({'success': False,
                        'message': 'sheet_id + tabs[] (or tab_name) required'}), 400
    if len(tabs) == 1:
        result = profile_manager.do_all_appeal_from_sheet(
            sheet_id=sheet_id, tab_name=tabs[0],
            num_workers=workers, resources_path=RESOURCES_PATH,
        )
    else:
        result = profile_manager.do_all_appeal_from_sheet_tabs(
            sheet_id=sheet_id, tabs=tabs,
            num_workers=workers, resources_path=RESOURCES_PATH,
        )
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/appeal-match-excel', methods=['POST'])
def profiles_appeal_match_excel():
    """Read an Excel file, extract emails, and return matched profile IDs."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '')
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        # Find the email column (case-insensitive)
        email_col = None
        for c in df.columns:
            if str(c).strip().lower() == 'email':
                email_col = c
                break
        if email_col is None:
            return jsonify({'success': False, 'message': 'No "Email" column found in the Excel file'})
        # Collect unique emails from Excel
        emails_in_excel = set()
        for _, row in df.iterrows():
            e = str(row.get(email_col, '')).strip().lower()
            if e and e != 'nan':
                emails_in_excel.add(e)
        if not emails_in_excel:
            return jsonify({'success': False, 'message': 'No emails found in the file'})
        # Match against existing profiles
        all_profiles = profile_manager.list_profiles()
        matched = []
        not_found = []
        for email in emails_in_excel:
            found = False
            for p in all_profiles:
                if (p.get('email') or '').strip().lower() == email:
                    matched.append({'id': p['id'], 'email': p.get('email', '')})
                    found = True
                    break
            if not found:
                not_found.append(email)
        return jsonify({
            'success': True,
            'total_emails': len(emails_in_excel),
            'matched': matched,
            'matched_count': len(matched),
            'not_found': not_found,
            'not_found_count': len(not_found),
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/profiles/<profile_id>/relogin', methods=['POST'])
def profiles_relogin(profile_id):
    """Re-login a single profile using its saved credentials."""
    try:
        result = profile_manager.relogin_profile(profile_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/bulk-relogin', methods=['POST'])
def profiles_bulk_relogin():
    """Re-login multiple selected profiles with worker control."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = data.get('ids', [])
        num_workers = max(1, min(int(data.get('workers', 2)), 10))
        stagger_delay = max(0, min(int(data.get('stagger_delay', 0)), 120))
        if not ids:
            return jsonify({'success': False, 'error': 'No profiles selected'})
        result = profile_manager.bulk_relogin_profiles(ids, num_workers, stagger_delay)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/bulk-relogin-status', methods=['GET'])
def profiles_bulk_relogin_status():
    try:
        return jsonify(profile_manager.get_bulk_relogin_status())
    except Exception:
        return jsonify({'running': False, 'status': 'idle'})


@app.route('/api/profiles/batch-login-status', methods=['GET'])
def profiles_batch_login_status():
    """Dedicated status endpoint for batch login.

    Previously batch-login + bulk-relogin both polled the shared /api/progress
    endpoint. When the user kicked off Live Check and a Re-Login back-to-back,
    the two cards' polls fought over the same processing_state and returned
    wrong totals on whichever popup the user wasn't watching. Each operation
    now has its own isolated endpoint backed by its own state dict.
    """
    try:
        return jsonify(profile_manager.get_batch_login_progress())
    except Exception:
        return jsonify({'running': False, 'status': 'idle'})


@app.route('/api/profiles/switch-to-local', methods=['POST'])
def profiles_switch_to_local():
    """Switch all (or selected) NST-engine profiles to local nexus engine."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = data.get('ids') or []
        result = profile_manager.switch_profiles_to_local(ids if ids else None)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/add-bookmarks', methods=['POST'])
def profiles_add_bookmarks():
    """Add/replace bookmarks in selected (or all) profile Chrome data directories.

    Runs in a background worker pool so the multi-card progress popup shows
    live progress. Pass `workers` (1-20) to tune parallelism.
    """
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = data.get('ids', [])
        bookmarks = data.get('bookmarks', [])          # [{path, name, url}, ...]
        bookmarks_text = data.get('bookmarks_text', '') # raw bookmark:: text
        replace = data.get('replace', True)
        num_workers = max(1, min(int(data.get('workers') or data.get('num_workers') or 5), 20))
        if not bookmarks and not bookmarks_text:
            return jsonify({'success': False, 'error': 'No bookmarks provided'})
        result = profile_manager.add_bookmarks_to_profiles_async(
            ids, bookmarks=bookmarks, bookmarks_text=bookmarks_text,
            replace=replace, num_workers=num_workers,
        )
        if not result.get('success'):
            return jsonify(result), 400
        return jsonify({'success': True, 'started': True, 'total': result.get('total', 0)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/stop-appeal', methods=['POST'])
def profiles_stop_appeal():
    """Stop running appeal operation."""
    return jsonify(profile_manager.stop_appeal())


@app.route('/api/profiles/stop-health', methods=['POST'])
def profiles_stop_health():
    """Stop running health activity."""
    return jsonify(profile_manager.stop_health())


# ── Profile Manager: Run Operations (Step 1/2) ──────────────────────────────

@app.route('/api/profiles/run-operations', methods=['POST'])
def profiles_run_operations():
    """Run Step 1/2 operations on all logged-in profiles."""
    data = request.get_json(force=True, silent=True) or {}
    operations = data.get('operations', '')
    num_workers = int(data.get('num_workers', 5))
    params = {
        'new_password': data.get('new_password', ''),
        'recovery_email': data.get('recovery_email', ''),
        'recovery_phone': data.get('recovery_phone', ''),
        'name_country': data.get('name_country', 'US'),
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
    }
    result = profile_manager.run_operations_on_profiles(
        operations=operations, num_workers=num_workers, params=params
    )
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/ops-status', methods=['GET'])
def profiles_ops_status():
    """Get status of running operations."""
    return jsonify(profile_manager.get_ops_status())


@app.route('/api/profiles/run-ops', methods=['POST'])
def profiles_run_ops():
    """Run operations on selected profiles."""
    data = request.get_json(force=True, silent=True) or {}
    profile_ids = data.get('profile_ids', [])
    operations = data.get('operations', '')
    params = data.get('params', {})
    num_workers = max(1, min(int(data.get('num_workers', 5)), 20))

    if not profile_ids:
        return jsonify({'success': False, 'error': 'No profiles selected'})
    if not operations:
        return jsonify({'success': False, 'error': 'No operations selected'})

    try:
        result = profile_manager.run_operations_on_profiles(
            operations=operations,
            num_workers=num_workers,
            params=params,
            profile_ids=profile_ids,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/run-health', methods=['POST'])
def profiles_run_health():
    """Start health activity on selected profiles with specific activities."""
    data = request.get_json(force=True, silent=True) or {}
    num_workers = int(data.get('num_workers', 3))
    activities = data.get('activities', [])
    profile_ids = data.get('profile_ids', [])
    country = data.get('country', 'US')
    rounds = int(data.get('rounds', 1))
    duration_minutes = int(data.get('duration_minutes', 0))
    gmb_name = data.get('gmb_name', '')
    gmb_address = data.get('gmb_address', '')
    result = profile_manager.run_health_activity(
        num_workers=num_workers,
        activities=activities,
        profile_ids=profile_ids,
        country=country,
        rounds=rounds,
        duration_minutes=duration_minutes,
        gmb_name=gmb_name,
        gmb_address=gmb_address,
    )
    return jsonify(result)


@app.route('/api/profiles/health-status', methods=['GET'])
def profiles_health_status():
    """Get status of running health activity."""
    return jsonify(profile_manager.get_health_status())


# ── Write Review — Google Sheet mode ─────────────────────────────────────────
from shared import sheet_review_orchestrator as _sheet_review


@app.route('/api/profiles/write-review/sheet/preview', methods=['POST'])
def profiles_write_review_sheet_preview():
    """For a chosen list of tabs, return how many rows are eligible
    to post (Status blank, has Review Text + Direct Review Link)
    versus already-posted rows."""
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    tabs = body.get('tabs') or []
    if not sheet_id or not tabs:
        return jsonify({'success': False,
                        'message': 'sheet_id + tabs[] required'}), 400
    return jsonify(_sheet_review.list_tabs_summary(
        RESOURCES_PATH, sheet_id, [str(t) for t in tabs],
    ))


@app.route('/api/profiles/write-review/sheet/start', methods=['POST'])
def profiles_write_review_sheet_start():
    """Kick off Write Review from a Google Sheet workbook with one tab
    per business. Body:
        sheet_id     : str
        tabs_config  : [{tab_name, count}]
        workers      : int
        profile_ids  : [str]  — which profiles will post (round-robin)"""
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    tabs_config = body.get('tabs_config') or []
    workers = int(body.get('workers') or 3)
    profile_ids = body.get('profile_ids') or []
    if not sheet_id or not tabs_config:
        return jsonify({'success': False,
                        'message': 'sheet_id + tabs_config required'}), 400
    if not profile_ids:
        return jsonify({'success': False,
                        'message': 'profile_ids[] required — pick at least one profile'}), 400
    result = profile_manager.do_write_review_from_sheet(
        sheet_id=sheet_id, tabs_config=tabs_config,
        num_workers=workers, resources_path=RESOURCES_PATH,
        profile_ids=profile_ids,
    )
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/write-review/sheet/parse', methods=['POST'])
def profiles_write_review_sheet_parse():
    """Full parse of a single tab — returns eligible rows in order so
    the start endpoint or a debug UI can show exactly what would run."""
    body = request.get_json(silent=True) or {}
    sheet_id = (body.get('sheet_id') or '').strip()
    tab = (body.get('tab_name') or '').strip()
    if not sheet_id or not tab:
        return jsonify({'success': False,
                        'message': 'sheet_id + tab_name required'}), 400
    return jsonify(_sheet_review.parse_business_tab(RESOURCES_PATH, sheet_id, tab))


@app.route('/api/profiles/do-write-review', methods=['POST'])
def profiles_do_write_review():
    """Start Write Review operation from Excel file for matched profiles."""
    data = request.get_json(force=True, silent=True) or {}
    excel_file = data.get('excel_file', '').strip()
    num_workers = int(data.get('num_workers', 3))
    profile_ids = data.get('profile_ids') or None
    if not excel_file:
        return jsonify({'success': False, 'message': 'excel_file path is required'}), 400
    if not os.path.isfile(excel_file):
        return jsonify({'success': False, 'message': f'File not found: {excel_file}'}), 400
    result = profile_manager.do_write_review_profiles(
        excel_file=excel_file,
        num_workers=num_workers,
        profile_ids=profile_ids,
    )
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/review-status', methods=['GET'])
def profiles_review_status():
    """Get status of running Write Review operation."""
    return jsonify(profile_manager.get_review_status())


@app.route('/api/profiles/write-review-template', methods=['GET'])
def profiles_write_review_template():
    """Return a pre-filled Excel template showing how to prepare the Write Review sheet."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 1 — Reviews (the actual data sheet)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = 'Reviews'

    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header style
    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Column definitions: (header, width, note)
    columns = [
        ('Email',        30),
        ('Review URL',   55),
        ('GMB URL',      45),
        ('Review Text',  55),
        ('Review Stars', 14),
    ]
    for col_idx, (h, w) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 24

    # Example rows  (Review URL, GMB URL, Review Text, Stars)
    examples = [
        ('john123@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0x1234:0x5678!12e1', 'https://maps.google.com/maps?cid=123456789012345', 'Amazing place! Highly recommend to everyone. Very professional and friendly staff.', 5),
        ('mary456@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xAAAA:0xBBBB!12e1', 'https://maps.google.com/maps?cid=987654321098765', 'Great experience overall. Clean, organised and the team is very helpful.', 4),
        ('alex789@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xCCCC:0xDDDD!12e1', 'https://maps.google.com/maps?cid=111222333444555', 'Excellent service and quality. Will definitely visit again. Truly outstanding!', 5),
        ('sara001@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xEEEE:0xFFFF!12e1', 'https://maps.google.com/maps?cid=666777888999000', '',                                                                                  5),
        ('test002@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0x1111:0x2222!12e1', 'https://maps.google.com/maps?cid=222333444555666', 'Very good. Satisfied with the service provided by the team here.',                 4),
    ]

    row_fills = ['EFF6FF', 'F0FDF4', 'EFF6FF', 'FEF9C3', 'F0FDF4']
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row_idx, (email, gmb, text, stars) in enumerate(examples, 2):
        fill = PatternFill('solid', fgColor=row_fills[row_idx - 2])
        for col_idx, value in enumerate([email, gmb, text, stars], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = border
            cell.fill = fill
        ws.row_dimensions[row_idx].height = 32

    ws.freeze_panes = 'A2'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2 — Instructions
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wi = wb.create_sheet('Instructions')
    wi.sheet_view.showGridLines = False
    wi.column_dimensions['A'].width = 2
    wi.column_dimensions['B'].width = 22
    wi.column_dimensions['C'].width = 60

    title_font  = Font(name='Calibri', bold=True, size=16, color='1E3A5F')
    h2_font     = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    h2_fill     = PatternFill('solid', fgColor='1E3A5F')
    label_font  = Font(name='Calibri', bold=True, size=11, color='1E40AF')
    body_font   = Font(name='Calibri', size=11, color='374151')
    note_font   = Font(name='Calibri', italic=True, size=10, color='6B7280')
    ok_font     = Font(name='Calibri', bold=True, size=11, color='065F46')
    warn_font   = Font(name='Calibri', bold=True, size=11, color='991B1B')
    green_fill  = PatternFill('solid', fgColor='D1FAE5')
    red_fill    = PatternFill('solid', fgColor='FEE2E2')

    def _irow(row, col_b='', col_c='', bfont=None, cfont=None, bfill=None, cfill=None, height=20, merge_bc=False):
        wi.row_dimensions[row].height = height
        if col_b:
            cb = wi.cell(row=row, column=2, value=col_b)
            if bfont: cb.font = bfont
            if bfill: cb.fill = bfill
            cb.alignment = Alignment(vertical='center', horizontal='left', indent=1)
        if col_c:
            cc = wi.cell(row=row, column=3, value=col_c)
            if cfont: cc.font = cfont
            if cfill: cc.fill = cfill
            cc.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        if merge_bc:
            wi.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    r = 1
    _irow(r, 'WRITE REVIEW — Excel Sheet Guide', '', bfont=title_font, height=36, merge_bc=True); r += 1
    _irow(r, height=10); r += 1

    # Required columns section
    _irow(r, '  REQUIRED COLUMNS', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    _irow(r, 'Email', 'The Gmail account email address. Must match exactly what is saved in the profile.', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, 'Review URL', 'Direct review link that opens the review popup instantly. Use "GMB → Review URL" tool to generate these. RECOMMENDED — much faster than GMB URL.', bfont=label_font, cfont=body_font, height=40); r += 1
    _irow(r, height=8); r += 1

    # Optional columns section
    _irow(r, '  OPTIONAL COLUMNS', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    _irow(r, 'GMB URL', 'Fallback: Google Maps business page URL. Only used if Review URL is empty for that row. Slower — requires clicking "Write a review" button.', bfont=label_font, cfont=body_font, height=40); r += 1
    _irow(r, 'Review Text', 'The review text to post. Leave blank to post stars only (no text review).', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, 'Review Stars', 'Number from 1 to 5. If left blank or missing, defaults to 5 stars.', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, height=8); r += 1

    # Rules section
    _irow(r, '  RULES', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    rules = [
        '1.  Column headers must be EXACTLY: Email, Review URL, GMB URL, Review Text, Review Stars',
        '2.  Spelling and capitalisation matters — "email" or "review url" will NOT work',
        '3.  One account per row — do not merge or duplicate rows',
        '4.  Rows with empty Email or no URL (both Review URL and GMB URL empty) are skipped',
        '5.  Stars must be a plain number: 1, 2, 3, 4 or 5 — not "5 stars" or "five"',
        '6.  The system matches each row to a profile by email — no manual selection needed',
        '7.  If an email is in the sheet but not saved as a profile, that row is skipped',
        '8.  Review URL is preferred — if both Review URL and GMB URL are filled, Review URL is used',
    ]
    for rule in rules:
        _irow(r, '', rule, cfont=body_font, height=22); r += 1
    _irow(r, height=8); r += 1

    # GMB URL section
    _irow(r, '  HOW TO GET GMB URL', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    gmb_steps = [
        'Step 1 →  Open Google Maps (maps.google.com)',
        'Step 2 →  Search for the business name',
        'Step 3 →  Click on the business listing',
        'Step 4 →  Copy the full URL from the browser address bar',
        'Step 5 →  Paste it into the GMB URL column',
        '',
        'Example URL:  https://www.google.com/maps/place/Business+Name/@lat,lng,zoom/...',
        'Also works:   https://maps.google.com/maps?cid=123456789012345',
    ]
    for step in gmb_steps:
        _irow(r, '', step, cfont=body_font, height=22); r += 1
    _irow(r, height=8); r += 1

    # Do / Don't
    _irow(r, '  ✅  CORRECT EXAMPLES', '', bfont=Font(name='Calibri', bold=True, size=12, color='065F46'), bfill=green_fill, height=24, merge_bc=True); r += 1
    goods = [
        'john@gmail.com  |  Review URL: https://google.com/maps/place//data=...  |  Great service!  |  5  (fastest)',
        'mary@gmail.com  |  Review URL: (blank)  |  GMB URL: https://maps.google.com/maps?cid=123  |  4  (fallback)',
    ]
    for g in goods:
        _irow(r, '', g, cfont=ok_font, height=22); r += 1
    _irow(r, height=8); r += 1

    _irow(r, '  ❌  WRONG EXAMPLES (will be skipped or cause errors)', '', bfont=Font(name='Calibri', bold=True, size=12, color='991B1B'), bfill=red_fill, height=24, merge_bc=True); r += 1
    bads = [
        'john123  |  (missing @gmail.com — won\'t match any profile)',
        '(blank)  |  https://google.com/maps/place/...  |  Review  |  5  — row skipped: no email',
        'mary@gmail.com  |  (both Review URL and GMB URL blank)  — row skipped: no URL',
        'john@gmail.com  |  Review URL  |  text  |  5 stars  — stars must be a number',
    ]
    for b in bads:
        _irow(r, '', b, cfont=warn_font, height=22); r += 1

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='WriteReview_Template.xlsx'
    )


@app.route('/api/profiles/write-review-preview', methods=['POST'])
def profiles_write_review_preview():
    """Preview Excel file for Write Review — returns matched profile count."""
    data = request.get_json(force=True, silent=True) or {}
    excel_file = data.get('excel_file', '').strip()
    if not excel_file or not os.path.isfile(excel_file):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(excel_file)
        cols = list(df.columns)
        has_review_url = 'Review URL' in cols
        has_gmb_url = 'GMB URL' in cols
        has_required = 'Email' in cols and (has_review_url or has_gmb_url)
        valid = 0
        if has_required:
            for _, row in df.iterrows():
                email = str(row.get('Email', '')).strip()
                review_url = str(row.get('Review URL', '')).strip() if has_review_url else ''
                gmb = str(row.get('GMB URL', '')).strip() if has_gmb_url else ''
                if review_url.lower() == 'nan': review_url = ''
                if gmb.lower() == 'nan': gmb = ''
                if email and email.lower() != 'nan' and (review_url or gmb):
                    valid += 1
        # Count how many profiles match the emails
        emails_in_excel = set()
        if has_required:
            for _, row in df.iterrows():
                e = str(row.get('Email', '')).strip().lower()
                if e and e != 'nan': emails_in_excel.add(e)
        all_profiles = profile_manager.list_profiles()
        matched = sum(1 for p in all_profiles if (p.get('email') or '').strip().lower() in emails_in_excel)
        return jsonify({
            'success': True, 'total_rows': len(df), 'valid_rows': valid,
            'matched_profiles': matched, 'columns': cols,
            'has_review_text': 'Review Text' in cols,
            'has_stars': 'Review Stars' in cols,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/profiles/write-review/import-emails', methods=['POST'])
def profiles_write_review_import_emails():
    """Read an Excel file and return profile IDs whose email matches any row."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'message': 'Only .xlsx / .xls files are accepted'})
    try:
        df = pd.read_excel(file_path)
        if df.empty:
            return jsonify({'success': True, 'matched_ids': [], 'matched_count': 0, 'not_found': []})
        # Find email column: prefer one named "Email" (case-insensitive), else use first column
        email_col = next(
            (c for c in df.columns if str(c).strip().lower() == 'email'),
            df.columns[0]
        )
        emails_in_xl = set()
        for val in df[email_col]:
            e = str(val).strip().lower()
            if e and e != 'nan':
                emails_in_xl.add(e)
        all_profiles = profile_manager.list_profiles()
        matched_ids = []
        matched_emails = set()
        for p in all_profiles:
            e = (p.get('email') or '').strip().lower()
            if e in emails_in_xl:
                matched_ids.append(p['id'])
                matched_emails.add(e)
        not_found = sorted(emails_in_xl - matched_emails)
        return jsonify({
            'success': True,
            'matched_ids': matched_ids,
            'matched_count': len(matched_ids),
            'not_found': not_found,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/gmb-to-review/preview', methods=['POST'])
def gmb_to_review_preview():
    """Preview Excel file for GMB → Review URL — returns row count and GMB URL count."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path)
        cols = list(df.columns)
        if 'GMB URL' not in cols:
            return jsonify({'success': False, 'message': 'Missing required column: "GMB URL"'})
        gmb_count = int(df['GMB URL'].dropna().astype(str).str.strip().loc[lambda s: (s != '') & (s.str.lower() != 'nan')].shape[0])
        return jsonify({
            'success': True,
            'total_rows': len(df),
            'gmb_url_count': gmb_count,
            'columns': cols,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


_gmb_review_progress = {
    'running': False, 'total': 0, 'done': 0, 'success': 0, 'failed': 0,
    'current_url': '', 'results': [], 'report_path': None,
}

def _gmb_review_worker(file_path):
    """Background thread: resolve GMB URLs and build Review URLs."""
    import re
    import requests as req_lib
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    global _gmb_review_progress
    try:
        df = pd.read_excel(file_path)
        rows = []
        for _, row in df.iterrows():
            gmb_url = str(row.get('GMB URL', '')).strip()
            if gmb_url and gmb_url.lower() != 'nan':
                rows.append((row, gmb_url))
            else:
                rows.append((row, ''))

        _gmb_review_progress['total'] = len([r for r in rows if r[1]])
        _gmb_review_progress['done'] = 0
        _gmb_review_progress['success'] = 0
        _gmb_review_progress['failed'] = 0
        _gmb_review_progress['results'] = []

        review_urls = []
        for row_data, gmb_url in rows:
            if not gmb_url:
                review_urls.append('')
                continue
            _gmb_review_progress['current_url'] = gmb_url
            try:
                r = req_lib.head(gmb_url, allow_redirects=True, timeout=15)
                match = re.search(r'(?:!1s|ftid=)(0x[0-9a-fA-F]+:0x[0-9a-fA-F]+)', r.url)
                if match:
                    hex_cid = match.group(1)
                    review_url = f"https://www.google.com/maps/place//data=!4m3!3m2!1s{hex_cid}!12e1"
                    review_urls.append(review_url)
                    _gmb_review_progress['success'] += 1
                    _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'success', 'review_url': review_url})
                else:
                    review_urls.append('ERROR: CID not found')
                    _gmb_review_progress['failed'] += 1
                    _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'failed', 'error': 'CID not found'})
            except Exception as e:
                review_urls.append(f'ERROR: {e}')
                _gmb_review_progress['failed'] += 1
                _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'failed', 'error': str(e)})
            _gmb_review_progress['done'] += 1

        df['Review URL'] = review_urls

        # Build styled Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Review URLs'

        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill('solid', fgColor='1E3A5F')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        headers = list(df.columns)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, col in enumerate(headers, 1):
                val = row[col]
                if pd.isna(val):
                    val = ''
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if col == 'Review URL':
                    if str(val).startswith('ERROR'):
                        cell.font = Font(color='DC2626')
                    elif val:
                        cell.font = Font(color='059669')

        for c_idx, col in enumerate(headers, 1):
            max_len = len(str(col))
            for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max_len + 4

        ws.freeze_panes = 'A2'

        # Save to output directory so it appears in Reports tab
        output_dir = str((Path(__file__).parent.parent / 'output').resolve())
        os.makedirs(output_dir, exist_ok=True)
        src_name = os.path.splitext(os.path.basename(file_path))[0]
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{src_name}_ReviewURLs_{ts}.xlsx'
        out_path = os.path.join(output_dir, filename)
        wb.save(out_path)
        _gmb_review_progress['report_path'] = out_path

    except Exception as e:
        _gmb_review_progress['results'].append({'url': '', 'status': 'failed', 'error': str(e)})
    finally:
        _gmb_review_progress['running'] = False
        _gmb_review_progress['current_url'] = ''


@app.route('/api/gmb-to-review/process', methods=['POST'])
def gmb_to_review_process():
    """Start background thread to resolve GMB URLs and build Review URLs."""
    import threading

    global _gmb_review_progress
    if _gmb_review_progress.get('running'):
        return jsonify({'success': False, 'message': 'Already processing'}), 400

    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'}), 400

    try:
        df = pd.read_excel(file_path)
        if 'GMB URL' not in df.columns:
            return jsonify({'success': False, 'message': 'Missing required column: "GMB URL"'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    _gmb_review_progress = {
        'running': True, 'total': 0, 'done': 0, 'success': 0, 'failed': 0,
        'current_url': '', 'results': [], 'report_path': None,
    }

    t = threading.Thread(target=_gmb_review_worker, args=(file_path,), daemon=True)
    t.start()

    return jsonify({'success': True, 'message': 'Processing started'})


@app.route('/api/gmb-to-review/status', methods=['GET'])
def gmb_to_review_status():
    """Return current progress of GMB → Review URL processing."""
    return jsonify({
        'running': _gmb_review_progress.get('running', False),
        'total': _gmb_review_progress.get('total', 0),
        'done': _gmb_review_progress.get('done', 0),
        'success': _gmb_review_progress.get('success', 0),
        'failed': _gmb_review_progress.get('failed', 0),
        'current_url': _gmb_review_progress.get('current_url', ''),
        'report_path': _gmb_review_progress.get('report_path'),
    })


@app.route('/api/gmb-to-review/download', methods=['GET'])
def gmb_to_review_download():
    """Download the generated Review URLs Excel file."""
    report_path = _gmb_review_progress.get('report_path')
    if not report_path or not os.path.isfile(report_path):
        return jsonify({'success': False, 'message': 'No report file available'}), 404
    return send_file(
        report_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=os.path.basename(report_path)
    )


@app.route('/api/nst/status', methods=['GET'])
def nst_status():
    """Check NST Browser API connectivity."""
    try:
        from shared.nexus_profile_manager import _nst_check, _nst_api_base
        connected = _nst_check()
        return jsonify({
            'success': True,
            'connected': connected,
            'api_base': _nst_api_base,
        })
    except Exception as e:
        return jsonify({'success': False, 'connected': False, 'message': str(e)})


@app.route('/api/nst/config', methods=['GET'])
def nst_config_get():
    """Get NST Browser config (browser.json)."""
    try:
        import json as _json
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = _json.loads(bj.read_text('utf-8'))
            # Mask key for display — only show last 8 chars
            key = cfg.get('nst_api_key', '')
            masked = ('*' * max(0, len(key) - 8) + key[-8:]) if len(key) > 8 else key
            return jsonify({
                'success': True,
                'nst_api_key': key,
                'nst_api_key_masked': masked,
                'nst_api_base': cfg.get('nst_api_base', 'http://localhost:8848/api/v2'),
            })
        return jsonify({'success': True, 'nst_api_key': '', 'nst_api_base': 'http://localhost:8848/api/v2'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nst/config', methods=['POST'])
def nst_config_save():
    """Save NST Browser config (API key + base URL) to browser.json and reload."""
    try:
        import json as _json
        data = request.get_json(force=True)
        bj = RESOURCES_PATH / 'config' / 'browser.json'

        # Load existing config
        cfg = {}
        if bj.exists():
            cfg = _json.loads(bj.read_text('utf-8'))

        # Update NST fields
        new_key = data.get('nst_api_key', '').strip()
        new_base = data.get('nst_api_base', '').strip()
        if new_key:
            cfg['nst_api_key'] = new_key
        if new_base:
            cfg['nst_api_base'] = new_base

        # Ensure NST mode is on
        cfg['use_nst'] = True

        # Save
        bj.write_text(_json.dumps(cfg, indent=4), 'utf-8')

        # Reload in profile manager
        try:
            import shared.nexus_profile_manager as npm
            npm._nst_api_key = cfg.get('nst_api_key', '')
            npm._nst_api_base = cfg.get('nst_api_base', 'http://localhost:8848/api/v2')
            connected = npm._nst_check()
        except Exception:
            connected = False

        return jsonify({'success': True, 'connected': connected, 'message': 'NST config saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nexus-api-key', methods=['GET'])
def nexus_api_key_get():
    """Get the Nexus API key from browser.json."""
    try:
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
            return jsonify({'success': True, 'nexus_api_key': cfg.get('nexus_api_key', '')})
        return jsonify({'success': True, 'nexus_api_key': ''})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nexus-api-key', methods=['POST'])
def nexus_api_key_save():
    """Save the Nexus API key to browser.json."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        cfg = {}
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
        cfg['nexus_api_key'] = data.get('nexus_api_key', '').strip()
        bj.write_text(json.dumps(cfg, indent=4), 'utf-8')
        return jsonify({'success': True, 'message': 'Nexus API key saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/name-countries', methods=['GET'])
def get_name_countries():
    """Get available countries for random name generation."""
    from shared.random_names import get_available_countries
    return jsonify(get_available_countries())


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VPN API — routes/vpn.py Blueprint (registered below)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Routes are in electron-app/backend/routes/vpn.py.
# Registered at bottom of file alongside other Blueprints.


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# NEXUS API v2 — External REST API (NST-compatible response format)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

nexus_api = Blueprint('nexus_api', __name__, url_prefix='/api/nexus/v2')


def _napi(data=None, msg='success', err=False, status=200):
    """NST-compatible JSON response wrapper."""
    return jsonify({'err': err, 'msg': msg, 'data': data}), status


def _load_nexus_api_key():
    """Load nexus_api_key from browser.json."""
    try:
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
            return cfg.get('nexus_api_key', '')
    except Exception:
        pass
    return ''


@nexus_api.before_request
def _nexus_api_check_key():
    """Validate x-api-key header if nexus_api_key is configured."""
    expected = _load_nexus_api_key()
    if not expected:
        return None  # No key set — allow all requests
    provided = request.headers.get('x-api-key', '')
    if provided != expected:
        return _napi(msg='Invalid or missing API key', err=True, status=401)


# ── Profile CRUD ─────────────────────────────────────────────────────────────

@nexus_api.route('/profiles/groups', methods=['GET'])
def napi_list_groups():
    """Return all unique profile group names."""
    profiles = profile_manager.list_profiles()
    all_groups = set()
    for p in profiles:
        for g in profile_manager._get_groups(p):
            all_groups.add(g)
    return _napi({'groups': sorted(all_groups)})


@nexus_api.route('/profiles', methods=['GET'])
def napi_list_profiles():
    """List profiles with optional filtering and pagination."""
    search = request.args.get('search', '').lower()
    filt = request.args.get('filter', 'all').lower()
    group_filter = request.args.get('group', '').lower()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    per_page = max(1, min(per_page, 10000))

    profiles = profile_manager.list_profiles()
    if search:
        profiles = [p for p in profiles if search in p.get('name', '').lower()
                    or search in p.get('email', '').lower()]
    if group_filter:
        profiles = [p for p in profiles if group_filter in [g.lower() for g in profile_manager._get_groups(p)]]
    if filt == 'running':
        profiles = [p for p in profiles if p.get('browser_open') == 'running']
    elif filt == 'logged_in':
        profiles = [p for p in profiles if p.get('status') == 'logged_in']
    elif filt == 'not_logged_in':
        profiles = [p for p in profiles if p.get('status') not in ('logged_in', 'login_failed')]
    elif filt == 'login_failed':
        profiles = [p for p in profiles if p.get('status') == 'login_failed']

    total = len(profiles)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page

    return _napi({
        'profiles': profiles[start:start + per_page],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@nexus_api.route('/profiles/<profile_id>', methods=['GET'])
def napi_get_profile(profile_id):
    """Get a single profile."""
    p = profile_manager.get_profile(profile_id)
    if not p:
        return _napi(msg='Profile not found', err=True, status=404)
    return _napi(p)


@nexus_api.route('/profiles', methods=['POST'])
def napi_create_profile():
    """Create a new profile."""
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', f'Profile {secrets.token_hex(3)}')
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'windows')

    # Parse proxy from string if given as string
    proxy = data.get('proxy')
    if isinstance(proxy, str) and proxy:
        from shared.nexus_proxy_manager import parse_proxy
        proxy = parse_proxy(proxy)

    try:
        profile = profile_manager.create_profile(
            name=name,
            email=data.get('email', ''),
            proxy=proxy,
            notes=data.get('notes', ''),
            fingerprint_prefs={'os_type': os_type},
            password=data.get('password', ''),
            totp_secret=data.get('totp_secret', ''),
            backup_codes=data.get('backup_codes', []),
            engine=engine,
            frontend_sections={
                'overview': data.get('overview', {}),
                'advanced': data.get('advanced', {}),
            },
        )
        nst_err = profile.pop('_nst_create_error', None)
        if nst_err:
            # Profile created locally — return success with warning
            return _napi(profile, msg=f'Profile created locally ({nst_err})', status=201)
        return _napi(profile, msg='Profile created', status=201)
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/<profile_id>', methods=['PUT'])
def napi_update_profile(profile_id):
    """Update profile fields."""
    data = request.get_json(force=True, silent=True) or {}
    # Parse proxy from string if given as string
    if 'proxy' in data and isinstance(data['proxy'], str) and data['proxy']:
        from shared.nexus_proxy_manager import parse_proxy
        data['proxy'] = parse_proxy(data['proxy'])
    try:
        result = profile_manager.update_profile(profile_id, **data)
        if result:
            return _napi(result, msg='Profile updated')
        return _napi(msg='Profile not found', err=True, status=404)
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/<profile_id>', methods=['DELETE'])
def napi_delete_profile(profile_id):
    """Delete a single profile."""
    ok = profile_manager.delete_profile(profile_id)
    if ok:
        return _napi(msg='Profile deleted')
    return _napi(msg='Profile not found', err=True, status=404)


@nexus_api.route('/profiles', methods=['DELETE'])
def napi_delete_all_profiles():
    """Delete ALL profiles."""
    profile_manager.delete_all_profiles()
    return _napi(msg='All profiles deleted')


# ── Browser Control ──────────────────────────────────────────────────────────

@nexus_api.route('/browsers/<profile_id>', methods=['POST'])
def napi_launch_browser(profile_id):
    """Launch browser and return CDP WebSocket URL."""
    try:
        ws = profile_manager.launch_and_connect(profile_id)
        return _napi({'webSocketDebuggerUrl': ws}, msg='Browser launched')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/browsers/<profile_id>', methods=['DELETE'])
def napi_stop_browser(profile_id):
    """Stop browser for a profile."""
    profile_manager.stop_nst_browser(profile_id)
    return _napi(msg='Browser stopped')


@nexus_api.route('/browsers/<profile_id>/launch', methods=['POST'])
def napi_launch_profile(profile_id):
    """Launch browser for UI viewing (Play button equivalent)."""
    result = profile_manager.launch_profile(profile_id)
    if result.get('success'):
        return _napi(result, msg='Browser launched')
    return _napi(msg=result.get('error', 'Launch failed'), err=True, status=500)


@nexus_api.route('/browsers/<profile_id>/close', methods=['POST'])
def napi_close_profile(profile_id):
    """Close browser for a profile."""
    ok = profile_manager.close_profile(profile_id)
    if ok:
        return _napi(msg='Browser closed')
    return _napi(msg='Browser not running', err=True, status=404)


@nexus_api.route('/browsers/close-all', methods=['POST'])
def napi_close_all():
    """Close all running browsers."""
    profile_manager.close_all_profiles()
    return _napi(msg='All browsers closed')


# ── Status ───────────────────────────────────────────────────────────────────

@nexus_api.route('/browsers/<profile_id>/status', methods=['GET'])
def napi_profile_status(profile_id):
    """Get browser status for a single profile."""
    return _napi(profile_manager.profile_status(profile_id))


@nexus_api.route('/browsers/status', methods=['GET'])
def napi_all_status():
    """Get status of all profiles."""
    profiles = profile_manager.list_profiles()
    statuses = {}
    for p in profiles:
        statuses[p['id']] = {
            'browser_open': p.get('browser_open', 'stopped'),
            'engine': p.get('engine', 'nexus'),
        }
    running = sum(1 for s in statuses.values() if s['browser_open'] == 'running')
    return _napi({
        'profiles': statuses,
        'running_count': running,
        'total_count': len(statuses),
    })


# ── Batch Operations ─────────────────────────────────────────────────────────

@nexus_api.route('/profiles/batch-create', methods=['POST'])
def napi_batch_create():
    """Create multiple profiles at once."""
    data = request.get_json(force=True, silent=True) or {}
    count = int(data.get('count', 1))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'windows')
    proxy_list = data.get('proxy_list', [])

    created = []
    for i in range(count):
        proxy = None
        if proxy_list:
            raw = proxy_list[i % len(proxy_list)]
            if isinstance(raw, str):
                from shared.nexus_proxy_manager import parse_proxy
                proxy = parse_proxy(raw)
            else:
                proxy = raw
        try:
            p = profile_manager.create_profile(
                name=f'Profile {i + 1}',
                fingerprint_prefs={'os_type': os_type},
                proxy=proxy,
                engine=engine,
            )
            created.append(p)
        except Exception:
            pass

    return _napi({'created': len(created), 'profiles': created}, msg=f'{len(created)} profiles created')


@nexus_api.route('/profiles/batch-login', methods=['POST'])
def napi_batch_login():
    """Batch login from Excel file."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '')
    workers = int(data.get('num_workers', 3))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'random')
    group = data.get('group', 'default') or 'default'
    if not file_path:
        return _napi(msg='file_path required', err=True, status=400)
    try:
        result = profile_manager.batch_login(file_path, workers, engine=engine, os_type=os_type, group=group)
        return _napi(result, msg='Batch login started')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/run-operations', methods=['POST'])
def napi_run_operations():
    """Run bot operations on profiles."""
    data = request.get_json(force=True, silent=True) or {}
    operations = data.get('operations', '')
    workers = int(data.get('num_workers', 5))
    params = data.get('params', {})
    try:
        result = profile_manager.run_operations_on_profiles(operations, workers, params)
        return _napi(result, msg='Operations started')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/ops-status', methods=['GET'])
def napi_ops_status():
    """Get current operations progress."""
    return _napi(profile_manager.get_ops_status())


# ── Config ───────────────────────────────────────────────────────────────────

@nexus_api.route('/config', methods=['GET'])
def napi_config_get():
    """Get profile manager config."""
    cfg = profile_manager.get_config()
    return _napi(cfg)


@nexus_api.route('/config', methods=['POST'])
def napi_config_set():
    """Update profile manager config."""
    data = request.get_json(force=True, silent=True) or {}
    if 'storage_path' in data:
        result = profile_manager.set_storage_path(data['storage_path'])
        return _napi(result, msg='Config updated')
    return _napi(msg='No config fields to update', err=True, status=400)


@nexus_api.route('/profiles/export', methods=['POST'])
def napi_export():
    """Export profiles to JSON."""
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get('profile_ids', [])
    result = profile_manager.export_profiles(ids)
    return _napi(result, msg='Exported')


# Register the Nexus API blueprint
app.register_blueprint(nexus_api)

# Register extracted route Blueprints
from routes.tools  import create_tools_blueprint
from routes.sheets import create_sheets_blueprint
from routes.vpn    import create_vpn_blueprint

app.register_blueprint(create_tools_blueprint(SCREENSHOTS_PATH))
app.register_blueprint(create_sheets_blueprint(RESOURCES_PATH))
app.register_blueprint(create_vpn_blueprint())


# ══════════════════════════════════════════════════════════════════════════════
# GMAIL CREATION CAMPAIGN API
# ══════════════════════════════════════════════════════════════════════════════

_active_campaigns: dict = {}  # name -> GmailCampaign


@app.route('/api/gmail-campaign/create', methods=['POST'])
def gmail_campaign_create():
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', f'Campaign_{int(time.time())}')
    config = data.get('config', {})
    try:
        from shared.gmail_campaign import GmailCampaign
        from shared.username_generator import generate_batch
        campaign = GmailCampaign(name, config)
        count = int(data.get('count', 1))
        identities = data.get('identities') or generate_batch(count)
        sms_country = config.get('sms_country', 'india')
        for ident in identities:
            ident['sms_country'] = sms_country
        campaign.add_accounts(identities)
        _active_campaigns[name] = campaign
        return jsonify({'ok': True, 'name': name, 'count': len(identities)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/gmail-campaign/start', methods=['POST'])
def gmail_campaign_start():
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', '')
    campaign = _active_campaigns.get(name)
    if not campaign:
        return jsonify({'ok': False, 'error': 'Campaign not found'}), 404
    import threading
    def _run():
        import asyncio
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(campaign.start())
        finally:
            loop.close()
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({'ok': True, 'message': f'Campaign started'})


@app.route('/api/gmail-campaign/stop', methods=['POST'])
def gmail_campaign_stop():
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', '')
    campaign = _active_campaigns.get(name)
    if not campaign:
        return jsonify({'ok': False, 'error': 'Campaign not found'}), 404
    campaign.stop()
    return jsonify({'ok': True})


@app.route('/api/gmail-campaign/status', methods=['GET'])
def gmail_campaign_status():
    name = request.args.get('name', '')
    if name:
        campaign = _active_campaigns.get(name)
        if not campaign:
            return jsonify({'ok': False, 'error': 'Campaign not found'}), 404
        return jsonify({'ok': True, 'campaign': campaign.get_status()})
    return jsonify({'ok': True, 'campaigns': {n: c.get_status() for n, c in _active_campaigns.items()}})


@app.route('/api/gmail-campaign/delete', methods=['DELETE'])
def gmail_campaign_delete():
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', '')
    if name in _active_campaigns:
        _active_campaigns[name].stop()
        del _active_campaigns[name]
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Not found'}), 404


@app.route('/api/gmail-campaign/sms-balance', methods=['POST'])
def gmail_sms_balance():
    data = request.get_json(force=True, silent=True) or {}
    try:
        from shared.sms_service import SMSService
        svc = SMSService(data.get('provider', ''), data.get('api_key', ''))
        return jsonify({'ok': True, 'balance': svc.get_balance()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/gmail-campaign/captcha-balance', methods=['POST'])
def gmail_captcha_balance():
    data = request.get_json(force=True, silent=True) or {}
    try:
        from shared.captcha_solver import CaptchaSolver
        solver = CaptchaSolver(data.get('provider', ''), data.get('api_key', ''))
        return jsonify({'ok': True, 'balance': solver.get_balance()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/gmail-campaign/generate-identities', methods=['POST'])
def gmail_generate_identities():
    data = request.get_json(force=True, silent=True) or {}
    count = min(int(data.get('count', 5)), 100)
    try:
        from shared.username_generator import generate_batch
        return jsonify({'ok': True, 'identities': generate_batch(count)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


def run_app():
    """Called by main_entry.py (frozen mode) to start the Flask server."""
    print("=" * 60)
    print("Gmail Bot Backend Server")
    print("=" * 60)
    print("Server starting on http://localhost:5000")
    _lic = license_manager.get_license_info()
    if _lic.get('valid'):
        print(f"[AUTH] License OK (tier={_lic.get('tier')}, days_left={_lic.get('days_remaining')})")
    else:
        print(f"[AUTH] License INVALID — {_lic.get('reason')} — most endpoints blocked until activation")
    print(f"[AUTH_TOKEN] {_INTERNAL_TOKEN}", flush=True)
    print("Server started - Ready to accept requests")
    print("=" * 60)
    app.run(host='127.0.0.1', port=5000, debug=False, threaded=True, use_reloader=False)


if __name__ == '__main__':
    run_app()
